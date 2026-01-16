import os
import re
import io
import glob
import calendar
from datetime import datetime, timedelta
from .validation import int_to_roman
from .lotes import listar_lotes, obter_lote_por_id, salvar_novo_lote, editar_lote, deletar_lote, _load_lotes_data, normalizar_precos
from .unidades import _load_unidades_data
from .mapas import _load_all_mapas_partitioned, _load_mapas_data, _load_mapas_partitioned, calcular_metricas_lotes


def calcular_saldo_consumido(custo_acumulado, valor_contratual, data_inicio_str):
	"""
	Calcula a porcentagem do saldo consumido considerando o tempo decorrido desde o início do contrato.
	
	Args:
		custo_acumulado: Valor já gasto até o momento
		valor_contratual: Valor total do contrato
		data_inicio_str: Data de início do contrato (formato YYYY-MM-DD)
	
	Returns:
		float: Porcentagem consumida (0-100)
	"""
	if not valor_contratual or valor_contratual == 0:
		return 0.0
	
	if not data_inicio_str:
		# Se não tem data de início, retorna apenas o percentual simples
		return (custo_acumulado / valor_contratual) * 100
	
	try:
		# Parsear data de início
		data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d')
		data_atual = datetime.now()
		
		# Calcular meses decorridos desde o início
		meses_decorridos = (data_atual.year - data_inicio.year) * 12 + (data_atual.month - data_inicio.month)
		
		# Considerar que contratos geralmente são de 12 meses
		duracao_contrato_meses = 12
		
		if meses_decorridos <= 0:
			# Contrato ainda não começou
			return 0.0
		elif meses_decorridos >= duracao_contrato_meses:
			# Contrato já terminou ou está no último período
			meses_decorridos = duracao_contrato_meses
		
		# Calcular quanto deveria ter sido consumido até agora (proporcional ao tempo)
		consumo_esperado = (meses_decorridos / duracao_contrato_meses) * valor_contratual
		
		# Retornar o percentual real consumido
		percentual_consumido = (custo_acumulado / valor_contratual) * 100
		
		return percentual_consumido
	except Exception:
		# Em caso de erro, retorna cálculo simples
		return (custo_acumulado / valor_contratual) * 100


def categorizar_mapas_por_unidade_flags(mapas, unit_flags):
	"""
	Categoriza mapas em três categorias baseado nos flags da unidade (delegacia, sub_empresa)
	
	Args:
		mapas: Lista de mapas para categorizar
		unit_flags: Dict com flags de cada unidade {nome_unidade: {delegacia: bool, sub_empresa: bool}}
	
	Returns:
		Tupla: (mapas_default, mapas_delegacia, mapas_sub_empresa)
	"""
	mapas_default = []
	mapas_delegacia = []
	mapas_sub_empresa = []
	
	for mapa in mapas:
		unidade_nome = (mapa.get('unidade') or '').strip()
		flags = unit_flags.get(unidade_nome, {})
		
		delegacia = flags.get('delegacia', False)
		sub_empresa = flags.get('sub_empresa', False)
		
		if delegacia:
			mapas_delegacia.append(mapa)
		elif sub_empresa:
			mapas_sub_empresa.append(mapa)
		else:
			mapas_default.append(mapa)
	
	return mapas_default, mapas_delegacia, mapas_sub_empresa


def carregar_lotes_para_dashboard():
	"""
	Carrega dados de lotes e mapas formatados para o dashboard
	Integra informações de lotes, unidades e mapas
	"""
	lotes_raw = _load_lotes_data() or []
	unidades_raw = _load_unidades_data() or []
	
	# Carregar TODOS os mapas (particionados por mês/ano)
	mapas = _load_all_mapas_partitioned() or []

	unidades_list = []
	if isinstance(unidades_raw, dict) and isinstance(unidades_raw.get('unidades'), list):
		unidades_list = unidades_raw.get('unidades')
	elif isinstance(unidades_raw, list):
		unidades_list = unidades_raw

	unidades_map = {}
	for u in unidades_list:
		if isinstance(u, dict) and isinstance(u.get('id'), int):
			unidades_map[int(u.get('id'))] = u.get('nome')

	lotes = []
	if isinstance(lotes_raw, dict) and isinstance(lotes_raw.get('lotes'), list):
		src_lotes = lotes_raw.get('lotes')
	elif isinstance(lotes_raw, list):
		src_lotes = lotes_raw
	else:
		src_lotes = []
	
	# Calcular métricas (custo_mes, refeicoes_mes, etc.) baseado nos mapas reais
	calcular_metricas_lotes(src_lotes, mapas)

	for l in src_lotes:
		if not isinstance(l, dict):
			continue
		raw_unidades = l.get('unidades') or []
		unidades_final = []
		if isinstance(raw_unidades, list) and raw_unidades:
			if all(isinstance(x, int) or (isinstance(x, str) and x.isdigit()) for x in raw_unidades):
				for x in raw_unidades:
					try:
						uid = int(x)
						unidades_final.append(unidades_map.get(uid, str(uid)))
					except Exception:
						unidades_final.append(str(x))
			else:
				unidades_final = [str(x) for x in raw_unidades if x]

		try:
			refeicoes_mes = int(l.get('refeicoes_mes') if l.get('refeicoes_mes') is not None else (l.get('refeicoes') or 0))
		except Exception:
			try:
				refeicoes_mes = int(float(l.get('refeicoes') or 0))
			except Exception:
				refeicoes_mes = 0
		try:
			custo_mes = float(l.get('custo_mes') if l.get('custo_mes') is not None else l.get('custo') or 0.0)
		except Exception:
			try:
				custo_mes = float(str(l.get('custo') or 0).replace(',', '.'))
			except Exception:
				custo_mes = 0.0
		try:
			desvio_mes = float(l.get('desvio_mes') if l.get('desvio_mes') is not None else l.get('desvio') or 0.0)
		except Exception:
			try:
				desvio_mes = float(str(l.get('desvio') or 0).replace(',', '.'))
			except Exception:
				desvio_mes = 0.0
		try:
			meses_cadastrados = int(l.get('meses_cadastrados') if l.get('meses_cadastrados') is not None else l.get('meses') or 0)
		except Exception:
			meses_cadastrados = 0

		try:
			conformidade_val = l.get('conformidade')
			if conformidade_val is None:
				conformidade = 0.0
			else:
				conformidade = float(str(conformidade_val).replace(',', '.'))
		except Exception:
			conformidade = 0.0

		try:
			valor_contratual = float(l.get('valor_contratual') or 0)
		except Exception:
			valor_contratual = 0.0
		
		# Calcular saldo consumido
		data_inicio = l.get('data_inicio')
		saldo_consumido = calcular_saldo_consumido(custo_mes, valor_contratual, data_inicio)

		lote_obj = {
			'id': l.get('id'),
			'nome': l.get('nome') or l.get('nome_lote') or '',
			'empresa': l.get('empresa') or '',
			'numero_contrato': l.get('numero_contrato') or l.get('contrato') or '',
			'contrato': l.get('numero_contrato') or l.get('contrato') or '',
			'sub_empresa': l.get('sub_empresa') or '',
			'data_inicio': data_inicio,
			'data_fim': l.get('data_fim') or '',
			'valor_contratual': valor_contratual,
			'ativo': l.get('ativo', True),
			'unidades': unidades_final,
			'precos': normalizar_precos(l.get('precos')),
			'quantitativos': l.get('quantitativos', {}),  # Adicionar quantitativos
			'refeicoes_mes': refeicoes_mes,
			'custo_mes': custo_mes,
			'desvio_mes': desvio_mes,
			'meses_cadastrados': meses_cadastrados,
			'refeicoes': l.get('refeicoes'),
			'conformidade': conformidade,
			'alertas': l.get('alertas'),
			'saldo_consumido': saldo_consumido,
			'criado_em': l.get('criado_em'),  # Adicionar data de criação
			'atualizado_em': l.get('atualizado_em'),  # Adicionar data de atualização
			'percentual_executado': l.get('percentual_executado', 0.0),  # Adicionar percentual calculado
			'refeicoes_por_mes': l.get('refeicoes_por_mes', {}),  # Adicionar refeições por mês
			'lote_predecessor_id': l.get('lote_predecessor_id'),  # Adicionar predecessor
			'ultima_atualizacao': l.get('ultima_atualizacao', 'Sem registro')  # Adicionar última atividade
		}
		lotes.append(lote_obj)

	mapas_dados = []
	mapas_list_src = []
	
	mapas_particionados = _load_all_mapas_partitioned()
	if mapas_particionados:
		mapas_list_src = mapas_particionados
	else:
		mapas_raw = _load_mapas_data() or []
		if isinstance(mapas_raw, dict) and isinstance(mapas_raw.get('mapas'), list):
			mapas_list_src = mapas_raw.get('mapas')
		elif isinstance(mapas_raw, list):
			mapas_list_src = mapas_raw
		else:
			mapas_list_src = []

	for m in mapas_list_src:
		if not isinstance(m, dict):
			continue
		try:
			lote_id = int(m.get('lote_id') if m.get('lote_id') is not None else m.get('lote') or m.get('loteId'))
		except Exception:
			try:
				lote_id = int(str(m.get('lote_id') or m.get('lote') or m.get('loteId')).strip())
			except Exception:
				lote_id = None

		mes_val = m.get('mes') or m.get('month') or m.get('mes_num')
		ano_val = m.get('ano') or m.get('year')
		try:
			mes = int(mes_val)
		except Exception:
			try:
				mes = int(str(mes_val).strip())
			except Exception:
				mes = None
		try:
			ano = int(ano_val)
		except Exception:
			try:
				ano = int(str(ano_val).strip())
			except Exception:
				ano = None

		unidade_raw = m.get('unidade') or m.get('unidade_nome') or m.get('unidadeNome') or ''
		nome_unidade = None
		try:
			if isinstance(unidade_raw, int):
				nome_unidade = unidades_map.get(int(unidade_raw))
			else:
				ustr = str(unidade_raw).strip()
				if ustr.isdigit():
					uid = int(ustr)
					nome_unidade = unidades_map.get(uid) or ustr
				else:
					nome_unidade = ustr
		except Exception:
			nome_unidade = str(unidade_raw)

		datas = m.get('datas') if isinstance(m.get('datas'), list) else []

		def _coerce_list(name):
			v = m.get(name)
			if isinstance(v, list):
				return v
			return []

		cafe_interno = _coerce_list('cafe_interno')
		cafe_funcionario = _coerce_list('cafe_funcionario')
		almoco_interno = _coerce_list('almoco_interno')
		almoco_funcionario = _coerce_list('almoco_funcionario')
		lanche_interno = _coerce_list('lanche_interno')
		lanche_funcionario = _coerce_list('lanche_funcionario')
		jantar_interno = _coerce_list('jantar_interno')
		jantar_funcionario = _coerce_list('jantar_funcionario')
		dados_siisp = _coerce_list('dados_siisp')

		total_refeicoes = 0
		n_days = 0
		if isinstance(datas, list) and len(datas) > 0:
			n_days = len(datas)
		else:
			n_days = max(len(cafe_interno), len(cafe_funcionario), len(almoco_interno), len(almoco_funcionario), 
						 len(lanche_interno), len(lanche_funcionario), len(jantar_interno), len(jantar_funcionario))

		for i in range(n_days):
			vals = 0
			for arr in (cafe_interno, cafe_funcionario, almoco_interno, almoco_funcionario, 
						lanche_interno, lanche_funcionario, jantar_interno, jantar_funcionario):
				try:
					v = arr[i] if i < len(arr) and (arr[i] is not None) else 0
					vals += int(v)
				except Exception:
					try:
						vals += int(float(arr[i]))
					except Exception:
						pass
			total_refeicoes += vals

		n_siisp = _coerce_list('n_siisp')
		if not n_siisp and dados_siisp:
			n_siisp = dados_siisp

		cafe_interno_siisp = _coerce_list('cafe_interno_siisp') if 'cafe_interno_siisp' in m else []
		almoco_interno_siisp = _coerce_list('almoco_interno_siisp') if 'almoco_interno_siisp' in m else []
		lanche_interno_siisp = _coerce_list('lanche_interno_siisp') if 'lanche_interno_siisp' in m else []
		jantar_interno_siisp = _coerce_list('jantar_interno_siisp') if 'jantar_interno_siisp' in m else []
		
		cafe_funcionario_siisp = _coerce_list('cafe_funcionario_siisp') if 'cafe_funcionario_siisp' in m else []
		almoco_funcionario_siisp = _coerce_list('almoco_funcionario_siisp') if 'almoco_funcionario_siisp' in m else []
		lanche_funcionario_siisp = _coerce_list('lanche_funcionario_siisp') if 'lanche_funcionario_siisp' in m else []
		jantar_funcionario_siisp = _coerce_list('jantar_funcionario_siisp') if 'jantar_funcionario_siisp' in m else []
		
		if n_siisp and not cafe_interno_siisp:
			for i in range(max(len(n_siisp), n_days)):
				siisp_dia = n_siisp[i] if i < len(n_siisp) and n_siisp[i] is not None else 0
				
				cafe_int_dia = cafe_interno[i] if i < len(cafe_interno) and cafe_interno[i] is not None else 0
				almoco_int_dia = almoco_interno[i] if i < len(almoco_interno) and almoco_interno[i] is not None else 0
				lanche_int_dia = lanche_interno[i] if i < len(lanche_interno) and lanche_interno[i] is not None else 0
				jantar_int_dia = jantar_interno[i] if i < len(jantar_interno) and jantar_interno[i] is not None else 0
				
				try:
					cafe_interno_siisp.append(int(cafe_int_dia) - int(siisp_dia))
					almoco_interno_siisp.append(int(almoco_int_dia) - int(siisp_dia))
					lanche_interno_siisp.append(int(lanche_int_dia) - int(siisp_dia))
					jantar_interno_siisp.append(int(jantar_int_dia) - int(siisp_dia))
				except Exception:
					cafe_interno_siisp.append(0)
					almoco_interno_siisp.append(0)
					lanche_interno_siisp.append(0)
					jantar_interno_siisp.append(0)
				
				cafe_func_dia = cafe_funcionario[i] if i < len(cafe_funcionario) and cafe_funcionario[i] is not None else 0
				almoco_func_dia = almoco_funcionario[i] if i < len(almoco_funcionario) and almoco_funcionario[i] is not None else 0
				lanche_func_dia = lanche_funcionario[i] if i < len(lanche_funcionario) and lanche_funcionario[i] is not None else 0
				jantar_func_dia = jantar_funcionario[i] if i < len(jantar_funcionario) and jantar_funcionario[i] is not None else 0
				
				try:
					cafe_funcionario_siisp.append(int(cafe_func_dia))
					almoco_funcionario_siisp.append(int(almoco_func_dia))
					lanche_funcionario_siisp.append(int(lanche_func_dia))
					jantar_funcionario_siisp.append(int(jantar_func_dia))
				except Exception:
					cafe_funcionario_siisp.append(0)
					almoco_funcionario_siisp.append(0)
					lanche_funcionario_siisp.append(0)
					jantar_funcionario_siisp.append(0)

		mapa_obj = {
			'id': m.get('id'),
			'lote_id': lote_id,
			'nome_unidade': nome_unidade,
			'mes': mes,
			'ano': ano,
			'data': datas,
			'linhas': int(m.get('linhas') or 0),
			'colunas_count': int(m.get('colunas_count') or 0),
			'cafe_interno': cafe_interno,
			'cafe_funcionario': cafe_funcionario,
			'almoco_interno': almoco_interno,
			'almoco_funcionario': almoco_funcionario,
			'lanche_interno': lanche_interno,
			'lanche_funcionario': lanche_funcionario,
			'jantar_interno': jantar_interno,
			'jantar_funcionario': jantar_funcionario,
			'dados_siisp': dados_siisp,
			'n_siisp': n_siisp,
			'cafe_interno_siisp': cafe_interno_siisp,
			'almoco_interno_siisp': almoco_interno_siisp,
			'lanche_interno_siisp': lanche_interno_siisp,
			'jantar_interno_siisp': jantar_interno_siisp,
			'cafe_funcionario_siisp': cafe_funcionario_siisp,
			'almoco_funcionario_siisp': almoco_funcionario_siisp,
			'lanche_funcionario_siisp': lanche_funcionario_siisp,
			'jantar_funcionario_siisp': jantar_funcionario_siisp,
			'refeicoes_mes': total_refeicoes,
			'criado_em': m.get('criado_em'),
			'atualizado_em': m.get('atualizado_em')
		}
		mapas_dados.append(mapa_obj)

	return {'lotes': lotes, 'mapas_dados': mapas_dados}


def gerar_excel_exportacao(lote_id, unidades_list, data_inicio=None, data_fim=None):
	BASE_DIR = os.path.dirname(os.path.abspath(__file__))
	dados_dir = os.path.join(BASE_DIR, '..', 'dados')
	"""
	Gera arquivo Excel com dados de um lote específico
	"""
	try:
		from openpyxl import load_workbook
		from openpyxl.formatting.rule import CellIsRule
		from openpyxl.styles import PatternFill
		from copy import copy
	except ImportError as e:
		return {'success': False, 'error': f'Biblioteca não instalada: {str(e)}'}
	
	try:
		dashboard_data = carregar_lotes_para_dashboard()
		lotes = dashboard_data.get('lotes', [])
		
		lote = None
		for l in lotes:
			try:
				if int(l.get('id')) == int(lote_id):
					lote = l
					break
			except Exception:
				continue
		
		if not lote:
			return {'success': False, 'error': 'Lote não encontrado'}
		
		precos = normalizar_precos(lote.get('precos', {}))
		# Garantir que todos os preços são float (número) para exportação
		for tipo_refeicao in precos:
			if isinstance(precos[tipo_refeicao], dict):
				for subcampo in precos[tipo_refeicao]:
					valor = precos[tipo_refeicao][subcampo]
					try:
						precos[tipo_refeicao][subcampo] = float(str(valor).replace(',', '.'))
					except Exception:
						precos[tipo_refeicao][subcampo] = 0.0
			else:
				try:
					precos[tipo_refeicao] = float(str(precos[tipo_refeicao]).replace(',', '.'))
				except Exception:
					precos[tipo_refeicao] = 0.0

		# Converter strings de data para datetime se fornecidas
		data_inicio_dt = None
		data_fim_dt = None
		if data_inicio:
			try:
				data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
				print(f"📅 Filtro data início: {data_inicio_dt.strftime('%Y-%m-%d')}")
			except Exception as e:
				print(f"⚠️ Erro ao converter data_inicio: {e}")
		if data_fim:
			try:
				data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
				print(f"📅 Filtro data fim: {data_fim_dt.strftime('%Y-%m-%d')}")
			except Exception as e:
				print(f"⚠️ Erro ao converter data_fim: {e}")
		

		# Buscar mapas diretamente do banco de dados
		from functions.mapas import carregar_mapas_db
		filtros = {'lote_id': lote_id}
		mapas_db = carregar_mapas_db(filtros)
		
		# Se o lote tiver predecessor, buscar também os mapas do predecessor
		predecessor_id = lote.get('lote_predecessor_id')
		precos_predecessor = {}
		
		if predecessor_id:
			print(f"📋 Buscando mapas do predecessor {predecessor_id} para exportação")
			# Buscar dados do lote predecessor
			predecessor_lote = None
			for l in lotes:
				try:
					if int(l.get('id')) == int(predecessor_id):
						predecessor_lote = l
						break
				except Exception:
					continue
			
			if predecessor_lote:
				print(f"✅ Predecessor encontrado: {predecessor_lote.get('nome')}")
				# Normalizar preços do predecessor
				precos_predecessor = normalizar_precos(predecessor_lote.get('precos', {}))
				for tipo_refeicao in precos_predecessor:
					if isinstance(precos_predecessor[tipo_refeicao], dict):
						for subcampo in precos_predecessor[tipo_refeicao]:
							valor = precos_predecessor[tipo_refeicao][subcampo]
							try:
								precos_predecessor[tipo_refeicao][subcampo] = float(str(valor).replace(',', '.'))
							except Exception:
								precos_predecessor[tipo_refeicao][subcampo] = 0.0
					else:
						try:
							precos_predecessor[tipo_refeicao] = float(str(precos_predecessor[tipo_refeicao]).replace(',', '.'))
						except Exception:
							precos_predecessor[tipo_refeicao] = 0.0
				
				# Buscar mapas do predecessor
				mapas_predecessor = carregar_mapas_db({'lote_id': predecessor_id})
				print(f"📊 Encontrados {len(mapas_predecessor)} mapas do predecessor")
				# Marcar cada mapa do predecessor com uma flag para usar preços diferentes
				for m in mapas_predecessor:
					m['_is_predecessor'] = True
				# Adicionar à lista de mapas
				mapas_db.extend(mapas_predecessor)
				print(f"📊 Total de mapas após incluir predecessor: {len(mapas_db)}")
		
		# Carregar flags das unidades (delegacia, sub_empresa)
		unit_flags = {}
		try:
			from functions.unidades import api_listar_unidades
			from functions.models import Unidade
			resultado_unidades = api_listar_unidades(lote_id)
			if resultado_unidades.get('success'):
				for u in resultado_unidades.get('unidades', []):
					nome_subempresa = ''
					# Se é sub_empresa, buscar o nome da unidade principal
					if u.get('sub_empresa') and u.get('unidade_principal_id'):
						try:
							unidade_principal = Unidade.query.get(u.get('unidade_principal_id'))
							if unidade_principal:
								nome_subempresa = unidade_principal.nome
						except:
							nome_subempresa = ''
					
					unit_flags[u.get('nome', '')] = {
						'delegacia': u.get('delegacia', False),
						'sub_empresa': u.get('sub_empresa', False),
						'nome_subempresa': nome_subempresa
					}
		except Exception as e:
			print(f"⚠️ Erro ao carregar flags das unidades: {e}")

		mapas_filtrados = []
		for m in mapas_db:
			# Filtrar por unidade
			if unidades_list:
				unidade_nome = (m.get('unidade') or '').strip()
				if unidade_nome not in unidades_list:
					continue
			# Filtrar por intervalo de datas
			if data_inicio_dt or data_fim_dt:
				datas = m.get('datas', [])
				if not datas:
					continue
				datas_filtradas = []
				indices_filtrados = []
				for idx, data_str in enumerate(datas):
					try:
						data_dt = None
						for formato in ['%d/%m/%Y', '%Y-%m-%d']:
							try:
								data_dt = datetime.strptime(data_str, formato)
								break
							except:
								continue
						if not data_dt:
							continue
						if data_inicio_dt and data_dt < data_inicio_dt:
							continue
						if data_fim_dt and data_dt > data_fim_dt:
							continue
						datas_filtradas.append(data_str)
						indices_filtrados.append(idx)
					except Exception:
						continue
				if not datas_filtradas:
					continue
				m_filtrado = m.copy()
				m_filtrado['datas'] = datas_filtradas
				for campo in ['cafe_interno', 'cafe_funcionario', 'almoco_interno', 'almoco_funcionario',
							  'lanche_interno', 'lanche_funcionario', 'jantar_interno', 'jantar_funcionario',
							  'dados_siisp', 'n_siisp']:
					if campo in m_filtrado and isinstance(m_filtrado[campo], list):
						m_filtrado[campo] = [m_filtrado[campo][i] for i in indices_filtrados if i < len(m_filtrado[campo])]
				mapas_filtrados.append(m_filtrado)
			else:
				mapas_filtrados.append(m)

		if not mapas_filtrados:
			return {'success': False, 'error': 'Nenhum dado encontrado para os filtros selecionados'}

		# Agrupar mapas por mês/ano
		from collections import defaultdict
		mapas_por_mes = defaultdict(list)
		
		for mapa in mapas_filtrados:
			datas_mapa = mapa.get('datas', [])
			# Agrupar por cada data presente no mapa
			for idx, data_str in enumerate(datas_mapa):
				try:
					data_dt = None
					for formato in ['%d/%m/%Y', '%Y-%m-%d']:
						try:
							data_dt = datetime.strptime(data_str, formato)
							break
						except:
							continue
					if not data_dt:
						continue
					
					# Criar chave mês/ano
					mes_ano_chave = (data_dt.year, data_dt.month)
					
					# Criar cópia do mapa com apenas os dados deste índice
					mapa_dia = mapa.copy()
					mapa_dia['datas'] = [data_str]
					for campo in ['cafe_interno', 'cafe_funcionario', 'almoco_interno', 'almoco_funcionario',
								  'lanche_interno', 'lanche_funcionario', 'jantar_interno', 'jantar_funcionario',
								  'dados_siisp', 'n_siisp']:
						if campo in mapa_dia and isinstance(mapa_dia[campo], list) and idx < len(mapa_dia[campo]):
							mapa_dia[campo] = [mapa_dia[campo][idx]]
						else:
							mapa_dia[campo] = []
					
					mapas_por_mes[mes_ano_chave].append(mapa_dia)
				except Exception:
					continue
		
		# Ordenar os meses cronologicamente
		meses_ordenados = sorted(mapas_por_mes.keys())
		
		# Se houver apenas um mês, manter o comportamento antigo (sem sufixo)
		usar_sufixo_mes = len(meses_ordenados) > 1

		modelo_path = os.path.join(dados_dir, 'modelo.xlsx')
		
		if not os.path.exists(modelo_path):
			return {'success': False, 'error': 'Arquivo modelo.xlsx não encontrado'}

		wb = load_workbook(modelo_path)
		
		# Obter as planilhas modelo
		ws1_modelo = None
		ws_resumo_modelo = None
		
		if 'COMPARATIVO' in wb.sheetnames:
			ws1_modelo = wb['COMPARATIVO']
		else:
			ws1_modelo = wb.active
			
		if 'RESUMO' in wb.sheetnames:
			ws_resumo_modelo = wb['RESUMO']
		
		# Renomear as planilhas modelo para temporário (evitar conflito quando copiar)
		if ws1_modelo:
			ws1_modelo.title = '_TEMP_COMPARATIVO'
		if ws_resumo_modelo:
			ws_resumo_modelo.title = '_TEMP_RESUMO'
		
		# Nomes dos meses em português
		meses_pt = [
			'', 'JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO',
			'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO'
		]
		
		# Processar cada mês
		for idx_mes, (ano, mes) in enumerate(meses_ordenados):
			mes_nome = meses_pt[mes]
			sufixo = f" - {mes_nome}" if usar_sufixo_mes else ""
			mapas_do_mes = mapas_por_mes[(ano, mes)]
			
			# Copiar planilha COMPARATIVO para este mês
			ws1 = wb.copy_worksheet(ws1_modelo)
			ws1.title = f'COMPARATIVO{sufixo}'
			
			# Definir larguras de coluna
			ws1.column_dimensions['A'].width = 19.5
			ws1.column_dimensions['B'].width = 30
			ws1.column_dimensions['C'].width = 10
			ws1.column_dimensions['D'].width = 12.5
			# Colunas E a L: 13
			for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
				ws1.column_dimensions[col].width = 13
			# Colunas M a T: 19
			for col in ['M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
				ws1.column_dimensions[col].width = 19
			
			# Categorizar mapas por flags da unidade
			mapas_default, mapas_delegacia, mapas_sub = categorizar_mapas_por_unidade_flags(mapas_do_mes, unit_flags)
			
			# Dicionário com as três categorias a processar
			categorias_resumo = []
			if mapas_default:
				categorias_resumo.append({
					'nome': f'RESUMO{sufixo}',
					'titulo': 'RESUMO FINAL',
					'mapas': mapas_default,
					'tipo': 'default'
				})
			if mapas_delegacia:
				categorias_resumo.append({
					'nome': f'RESUMO DELEGACIA{sufixo}',
					'titulo': 'RESUMO FINAL DELEGACIA',
					'mapas': mapas_delegacia,
					'tipo': 'delegacia'
				})
			if mapas_sub:
				# Obter nome da subempresa do lote
				sub_empresa_nome = lote.get('sub_empresa', 'SUBEMPRESA')
				categorias_resumo.append({
					'nome': f'RESUMO ({sub_empresa_nome}){sufixo}',
					'titulo': f'RESUMO FINAL {sub_empresa_nome}',
					'mapas': mapas_sub,
					'tipo': 'sub_empresa'
				})
			
			# Criar e preencher cada planilha RESUMO
			for categoria in categorias_resumo:
				if ws_resumo_modelo:
					ws_resumo = wb.copy_worksheet(ws_resumo_modelo)
					ws_resumo.title = categoria['nome']
					
					# Definir larguras de coluna para RESUMO
					ws_resumo.column_dimensions['B'].width = 22
					ws_resumo.column_dimensions['C'].width = 50
					# Colunas D a K: 30
					for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
						ws_resumo.column_dimensions[col].width = 30
					
					precos_ordem = [
						('cafe', 'interno'),
						('cafe', 'funcionario'),
						('almoco', 'interno'),
						('almoco', 'funcionario'),
						('lanche', 'interno'),
						('lanche', 'funcionario'),
						('jantar', 'interno'),
						('jantar', 'funcionario')
					]
					
					# Variáveis compartilhadas entre RESUMO e COMPARATIVO
					empresa_nome = lote.get('empresa', '')
					lote_nome = lote.get('nome', f"LOTE {lote_id}")
					
					# Preencher informações de contrato
					contrato_numero = lote.get('contrato', '')
					ws_resumo['B8'] = f"CONTRATO : {contrato_numero}"
					
					# Determinar qual tabela de preços usar para este mês
					mapas_categoria = categoria['mapas']
					todos_predecessor = all(m.get('_is_predecessor', False) for m in mapas_categoria)
					precos_para_resumo = precos_predecessor if todos_predecessor else precos
					
					# Determinar o formato da data baseado nos dados desta categoria
					periodo_texto = ''
					if mapas_categoria:
						todas_datas = []
						for mapa in mapas_categoria:
							datas_mapa = mapa.get('datas', [])
							for data_str in datas_mapa:
								try:
									data_dt = None
									for formato in ['%d/%m/%Y', '%Y-%m-%d']:
										try:
											data_dt = datetime.strptime(data_str, formato)
											break
										except:
											continue
									if data_dt:
										todas_datas.append(data_dt)
								except:
									continue
						
						if todas_datas:
							todas_datas.sort()
							mes_nome_periodo = meses_pt[mes]
							periodo_texto = f"{mes_nome_periodo} - {ano}"
					
					# Formatar título baseado no tipo de categoria
					if categoria['tipo'] == 'delegacia':
						texto_resumo = f"RESUMO FINAL {lote_nome} - DELEGACIA - {periodo_texto}".upper()
					elif categoria['tipo'] == 'sub_empresa':
						sub_empresa_nome = lote.get('sub_empresa', 'SUBEMPRESA')
						texto_resumo = f"RESUMO {lote_nome} - EMPRESA {sub_empresa_nome} - {periodo_texto}".upper()
					else:  # default
						texto_resumo = f"RESUMO FINAL {lote_nome} - EMPRESA {empresa_nome} - {periodo_texto}".upper()
					
					ws_resumo['B7'] = texto_resumo
					
					for merged_range in list(ws_resumo.merged_cells.ranges):
						min_col = merged_range.min_col
						max_col = merged_range.max_col
						min_row = merged_range.min_row
						max_row = merged_range.max_row
						if max_row >= 13 and min_col >= 2 and max_col <= 11:
							ws_resumo.unmerge_cells(str(merged_range))
					
					if unidades_list:
						nomes_unidades = [u for u in unidades_list if any(m.get('unidade') == u for m in mapas_categoria)]
					else:
						nomes_unidades = list(set(m.get('unidade', '') for m in mapas_categoria if m.get('unidade')))
						nomes_unidades.sort()
					
					quantidade_unidades = len(nomes_unidades)
					
					# Pegar o estilo da célula B11 como referência
					estilo_b11 = ws_resumo['B11']
					
					# Pegar o estilo da célula C11 como referência para os nomes das unidades
					estilo_c11 = ws_resumo['C11']
				
					# Pegar o estilo da célula D11 como referência para os totais
					estilo_d11 = ws_resumo['D11']
					
					# Pegar estilos das células E11-K11 para os totais
					estilo_e11 = ws_resumo['E11']
					estilo_f11 = ws_resumo['F11']
					estilo_g11 = ws_resumo['G11']
					estilo_h11 = ws_resumo['H11']
					estilo_i11 = ws_resumo['I11']
					estilo_j11 = ws_resumo['J11']
				estilo_k11 = ws_resumo['K11']
				
				# Pegar estilos das células D13-K13 para os valores unitários
				estilo_d13 = ws_resumo['D13']
				estilo_e13 = ws_resumo['E13']
				estilo_f13 = ws_resumo['F13']
				estilo_g13 = ws_resumo['G13']
				estilo_h13 = ws_resumo['H13']
				estilo_i13 = ws_resumo['I13']
				estilo_j13 = ws_resumo['J13']
				estilo_k13 = ws_resumo['K13']
				
				# Pegar estilos das células D14-K14 para os totais
				estilo_d14 = ws_resumo['D14']
				estilo_e14 = ws_resumo['E14']
				estilo_f14 = ws_resumo['F14']
				estilo_g14 = ws_resumo['G14']
				estilo_h14 = ws_resumo['H14']
				estilo_i14 = ws_resumo['I14']
				estilo_j14 = ws_resumo['J14']
				estilo_k14 = ws_resumo['K14']
				
				# Pegar estilos das células D15-K15 para os valores parciais
				estilo_d15 = ws_resumo['D15']
				estilo_e15 = ws_resumo['E15']
				estilo_f15 = ws_resumo['F15']
				estilo_g15 = ws_resumo['G15']
				estilo_h15 = ws_resumo['H15']
				estilo_i15 = ws_resumo['I15']
				estilo_j15 = ws_resumo['J15']
				estilo_k15 = ws_resumo['K15']
				
				# Pegar estilo da célula D16 para o valor total
				estilo_d16 = ws_resumo['D16']
				
				# Pegar estilo da célula B13 para a coluna mesclada
				estilo_b13 = ws_resumo['B13']
				# Preencher valores unitários na linha 13 (preços do lote - usar preços corretos para o mês)
				# D13: Café interno
				cell_d13 = ws_resumo.cell(row=13, column=4, value=precos_para_resumo.get('cafe', {}).get('interno', 0))
				if estilo_d13.has_style:
					cell_d13.font = copy(estilo_d13.font)
					cell_d13.border = copy(estilo_d13.border)
					cell_d13.fill = copy(estilo_d13.fill)
					cell_d13.number_format = copy(estilo_d13.number_format)
					cell_d13.protection = copy(estilo_d13.protection)
					cell_d13.alignment = copy(estilo_d13.alignment)
				
				# E13: Café funcionário
				cell_e13 = ws_resumo.cell(row=13, column=5, value=precos_para_resumo.get('cafe', {}).get('funcionario', 0))
				if estilo_e13.has_style:
					cell_e13.font = copy(estilo_e13.font)
					cell_e13.border = copy(estilo_e13.border)
					cell_e13.fill = copy(estilo_e13.fill)
					cell_e13.number_format = copy(estilo_e13.number_format)
					cell_e13.protection = copy(estilo_e13.protection)
					cell_e13.alignment = copy(estilo_e13.alignment)
				
				# F13: Almoço interno
				cell_f13 = ws_resumo.cell(row=13, column=6, value=precos_para_resumo.get('almoco', {}).get('interno', 0))
				if estilo_f13.has_style:
					cell_f13.font = copy(estilo_f13.font)
					cell_f13.border = copy(estilo_f13.border)
					cell_f13.fill = copy(estilo_f13.fill)
					cell_f13.number_format = copy(estilo_f13.number_format)
					cell_f13.protection = copy(estilo_f13.protection)
					cell_f13.alignment = copy(estilo_f13.alignment)
				
				# G13: Almoço funcionário
				cell_g13 = ws_resumo.cell(row=13, column=7, value=precos_para_resumo.get('almoco', {}).get('funcionario', 0))
				if estilo_g13.has_style:
					cell_g13.font = copy(estilo_g13.font)
					cell_g13.border = copy(estilo_g13.border)
					cell_g13.fill = copy(estilo_g13.fill)
					cell_g13.number_format = copy(estilo_g13.number_format)
					cell_g13.protection = copy(estilo_g13.protection)
					cell_g13.alignment = copy(estilo_g13.alignment)
				
				# H13: Lanche interno
				cell_h13 = ws_resumo.cell(row=13, column=8, value=precos_para_resumo.get('lanche', {}).get('interno', 0))
				if estilo_h13.has_style:
					cell_h13.font = copy(estilo_h13.font)
					cell_h13.border = copy(estilo_h13.border)
					cell_h13.fill = copy(estilo_h13.fill)
					cell_h13.number_format = copy(estilo_h13.number_format)
					cell_h13.protection = copy(estilo_h13.protection)
					cell_h13.alignment = copy(estilo_h13.alignment)
				
				# I13: Lanche funcionário
				cell_i13 = ws_resumo.cell(row=13, column=9, value=precos_para_resumo.get('lanche', {}).get('funcionario', 0))
				if estilo_i13.has_style:
					cell_i13.font = copy(estilo_i13.font)
					cell_i13.border = copy(estilo_i13.border)
					cell_i13.fill = copy(estilo_i13.fill)
					cell_i13.number_format = copy(estilo_i13.number_format)
					cell_i13.protection = copy(estilo_i13.protection)
					cell_i13.alignment = copy(estilo_i13.alignment)
				
				# J13: Jantar interno
				cell_j13 = ws_resumo.cell(row=13, column=10, value=precos_para_resumo.get('jantar', {}).get('interno', 0))
				if estilo_j13.has_style:
					cell_j13.font = copy(estilo_j13.font)
					cell_j13.border = copy(estilo_j13.border)
					cell_j13.fill = copy(estilo_j13.fill)
					cell_j13.number_format = copy(estilo_j13.number_format)
					cell_j13.protection = copy(estilo_j13.protection)
					cell_j13.alignment = copy(estilo_j13.alignment)
				
				# K13: Jantar funcionário
				cell_k13 = ws_resumo.cell(row=13, column=11, value=precos_para_resumo.get('jantar', {}).get('funcionario', 0))
				if estilo_k13.has_style:
					cell_k13.font = copy(estilo_k13.font)
					cell_k13.border = copy(estilo_k13.border)
					cell_k13.fill = copy(estilo_k13.fill)
					cell_k13.number_format = copy(estilo_k13.number_format)
					cell_k13.protection = copy(estilo_k13.protection)
					cell_k13.alignment = copy(estilo_k13.alignment)
				
				# Adicionar linhas se necessário (quantidade_unidades - 1)
				if quantidade_unidades > 1:
					linhas_para_adicionar = quantidade_unidades - 1
					# Inserir linhas após a linha 11
					ws_resumo.insert_rows(12, linhas_para_adicionar)
				
				# Calcular totais por unidade (VALORES MONETÁRIOS, não quantidades)
				# Para lidar com preços diferentes (lote atual vs predecessor)
				totais_valor_cafe_interno = {}
				totais_valor_cafe_funcionario = {}
				totais_valor_almoco_interno = {}
				totais_valor_almoco_funcionario = {}
				totais_valor_lanche_interno = {}
				totais_valor_lanche_funcionario = {}
				totais_valor_jantar_interno = {}
				totais_valor_jantar_funcionario = {}
				
				# Também manter totais de quantidades para referência
				totais_cafe_interno = {}
				totais_cafe_funcionario = {}
				totais_almoco_interno = {}
				totais_almoco_funcionario = {}
				totais_lanche_interno = {}
				totais_lanche_funcionario = {}
				totais_jantar_interno = {}
				totais_jantar_funcionario = {}
				
				for mapa in mapas_do_mes:
					unidade_nome = (mapa.get('unidade') or '').strip()
					
					# Determinar qual preço usar (lote atual ou predecessor)
					is_predecessor = mapa.get('_is_predecessor', False)
					precos_usados = precos_predecessor if is_predecessor else precos
					
					# Extrair preços individuais
					preco_cafe_interno = precos_usados.get('cafe', {}).get('interno', 0) if isinstance(precos_usados.get('cafe'), dict) else 0
					preco_cafe_funcionario = precos_usados.get('cafe', {}).get('funcionario', 0) if isinstance(precos_usados.get('cafe'), dict) else 0
					preco_almoco_interno = precos_usados.get('almoco', {}).get('interno', 0) if isinstance(precos_usados.get('almoco'), dict) else 0
					preco_almoco_funcionario = precos_usados.get('almoco', {}).get('funcionario', 0) if isinstance(precos_usados.get('almoco'), dict) else 0
					preco_lanche_interno = precos_usados.get('lanche', {}).get('interno', 0) if isinstance(precos_usados.get('lanche'), dict) else 0
					preco_lanche_funcionario = precos_usados.get('lanche', {}).get('funcionario', 0) if isinstance(precos_usados.get('lanche'), dict) else 0
					preco_jantar_interno = precos_usados.get('jantar', {}).get('interno', 0) if isinstance(precos_usados.get('jantar'), dict) else 0
					preco_jantar_funcionario = precos_usados.get('jantar', {}).get('funcionario', 0) if isinstance(precos_usados.get('jantar'), dict) else 0
					
					# Inicializar dicionários se necessário
					if unidade_nome not in totais_cafe_interno:
						totais_cafe_interno[unidade_nome] = 0
						totais_cafe_funcionario[unidade_nome] = 0
						totais_almoco_interno[unidade_nome] = 0
						totais_almoco_funcionario[unidade_nome] = 0
						totais_lanche_interno[unidade_nome] = 0
						totais_lanche_funcionario[unidade_nome] = 0
						totais_jantar_interno[unidade_nome] = 0
						totais_jantar_funcionario[unidade_nome] = 0
						totais_valor_cafe_interno[unidade_nome] = 0
						totais_valor_cafe_funcionario[unidade_nome] = 0
						totais_valor_almoco_interno[unidade_nome] = 0
						totais_valor_almoco_funcionario[unidade_nome] = 0
						totais_valor_lanche_interno[unidade_nome] = 0
						totais_valor_lanche_funcionario[unidade_nome] = 0
						totais_valor_jantar_interno[unidade_nome] = 0
						totais_valor_jantar_funcionario[unidade_nome] = 0
					
					# Somar café interno (quantidade e valor)
					for valor in mapa.get('cafe_interno', []):
						try:
							qtd = int(valor) if valor is not None else 0
							totais_cafe_interno[unidade_nome] += qtd
							totais_valor_cafe_interno[unidade_nome] += qtd * preco_cafe_interno
						except:
							pass
					
					# Somar café funcionário
					for valor in mapa.get('cafe_funcionario', []):
						try:
							qtd = int(valor) if valor is not None else 0
							totais_cafe_funcionario[unidade_nome] += qtd
							totais_valor_cafe_funcionario[unidade_nome] += qtd * preco_cafe_funcionario
						except:
							pass
					
					# Somar almoço interno
					for valor in mapa.get('almoco_interno', []):
						try:
							qtd = int(valor) if valor is not None else 0
							totais_almoco_interno[unidade_nome] += qtd
							totais_valor_almoco_interno[unidade_nome] += qtd * preco_almoco_interno
						except:
							pass
					
					# Somar almoço funcionário
					for valor in mapa.get('almoco_funcionario', []):
						try:
							qtd = int(valor) if valor is not None else 0
							totais_almoco_funcionario[unidade_nome] += qtd
							totais_valor_almoco_funcionario[unidade_nome] += qtd * preco_almoco_funcionario
						except:
							pass
					
					# Somar lanche interno
					for valor in mapa.get('lanche_interno', []):
						try:
							qtd = int(valor) if valor is not None else 0
							totais_lanche_interno[unidade_nome] += qtd
							totais_valor_lanche_interno[unidade_nome] += qtd * preco_lanche_interno
						except:
							pass
					
					# Somar lanche funcionário
					for valor in mapa.get('lanche_funcionario', []):
						try:
							qtd = int(valor) if valor is not None else 0
							totais_lanche_funcionario[unidade_nome] += qtd
							totais_valor_lanche_funcionario[unidade_nome] += qtd * preco_lanche_funcionario
						except:
							pass
					
					# Somar jantar interno
					for valor in mapa.get('jantar_interno', []):
						try:
							qtd = int(valor) if valor is not None else 0
							totais_jantar_interno[unidade_nome] += qtd
							totais_valor_jantar_interno[unidade_nome] += qtd * preco_jantar_interno
						except:
							pass
					
					# Somar jantar funcionário
					for valor in mapa.get('jantar_funcionario', []):
						try:
							qtd = int(valor) if valor is not None else 0
							totais_jantar_funcionario[unidade_nome] += qtd
							totais_valor_jantar_funcionario[unidade_nome] += qtd * preco_jantar_funcionario
						except:
							pass
				
				# Preencher a coluna ORDEM (coluna B), UNIDADE (coluna C) e CAFÉ INTERNO (coluna D) começando da linha 11
				for i in range(quantidade_unidades):
					linha_atual = 11 + i
					nome_unidade_atual = nomes_unidades[i]
					
					# Coluna B: ORDEM (número sequencial)
					cell_ordem = ws_resumo.cell(row=linha_atual, column=2, value=i + 1)
					
					# Copiar o estilo de B11 para a célula atual
					if estilo_b11.has_style:
						cell_ordem.font = copy(estilo_b11.font)
						cell_ordem.border = copy(estilo_b11.border)
						cell_ordem.fill = copy(estilo_b11.fill)
						cell_ordem.number_format = copy(estilo_b11.number_format)
						cell_ordem.protection = copy(estilo_b11.protection)
						cell_ordem.alignment = copy(estilo_b11.alignment)
					
					# Coluna C: Nome da unidade
					cell_unidade = ws_resumo.cell(row=linha_atual, column=3, value=nomes_unidades[i])
					
					# Copiar o estilo de C11 para a célula atual
					if estilo_c11.has_style:
						cell_unidade.font = copy(estilo_c11.font)
						cell_unidade.border = copy(estilo_c11.border)
						cell_unidade.fill = copy(estilo_c11.fill)
						cell_unidade.number_format = copy(estilo_c11.number_format)
						cell_unidade.protection = copy(estilo_c11.protection)
						cell_unidade.alignment = copy(estilo_c11.alignment)
					
					# Coluna D: Total de café interno
					total_cafe = totais_cafe_interno.get(nome_unidade_atual, 0)
					cell_cafe = ws_resumo.cell(row=linha_atual, column=4, value=total_cafe)
					
					# Copiar o estilo de D11 para a célula atual
					if estilo_d11.has_style:
						cell_cafe.font = copy(estilo_d11.font)
						cell_cafe.border = copy(estilo_d11.border)
						cell_cafe.fill = copy(estilo_d11.fill)
						cell_cafe.number_format = copy(estilo_d11.number_format)
						cell_cafe.protection = copy(estilo_d11.protection)
						cell_cafe.alignment = copy(estilo_d11.alignment)
					
					# Coluna E: Total de café funcionário
					cell_e = ws_resumo.cell(row=linha_atual, column=5, value=totais_cafe_funcionario.get(nome_unidade_atual, 0))
					if estilo_e11.has_style:
						cell_e.font = copy(estilo_e11.font)
						cell_e.border = copy(estilo_e11.border)
						cell_e.fill = copy(estilo_e11.fill)
						cell_e.number_format = copy(estilo_e11.number_format)
						cell_e.protection = copy(estilo_e11.protection)
						cell_e.alignment = copy(estilo_e11.alignment)
					
					# Coluna F: Total de almoço interno
					cell_f = ws_resumo.cell(row=linha_atual, column=6, value=totais_almoco_interno.get(nome_unidade_atual, 0))
					if estilo_f11.has_style:
						cell_f.font = copy(estilo_f11.font)
						cell_f.border = copy(estilo_f11.border)
						cell_f.fill = copy(estilo_f11.fill)
						cell_f.number_format = copy(estilo_f11.number_format)
						cell_f.protection = copy(estilo_f11.protection)
						cell_f.alignment = copy(estilo_f11.alignment)
					
					# Coluna G: Total de almoço funcionário
					cell_g = ws_resumo.cell(row=linha_atual, column=7, value=totais_almoco_funcionario.get(nome_unidade_atual, 0))
					if estilo_g11.has_style:
						cell_g.font = copy(estilo_g11.font)
						cell_g.border = copy(estilo_g11.border)
						cell_g.fill = copy(estilo_g11.fill)
						cell_g.number_format = copy(estilo_g11.number_format)
						cell_g.protection = copy(estilo_g11.protection)
						cell_g.alignment = copy(estilo_g11.alignment)
					
					# Coluna H: Total de lanche interno
					cell_h = ws_resumo.cell(row=linha_atual, column=8, value=totais_lanche_interno.get(nome_unidade_atual, 0))
					if estilo_h11.has_style:
						cell_h.font = copy(estilo_h11.font)
						cell_h.border = copy(estilo_h11.border)
						cell_h.fill = copy(estilo_h11.fill)
						cell_h.number_format = copy(estilo_h11.number_format)
						cell_h.protection = copy(estilo_h11.protection)
						cell_h.alignment = copy(estilo_h11.alignment)
					
					# Coluna I: Total de lanche funcionário
					cell_i = ws_resumo.cell(row=linha_atual, column=9, value=totais_lanche_funcionario.get(nome_unidade_atual, 0))
					if estilo_i11.has_style:
						cell_i.font = copy(estilo_i11.font)
						cell_i.border = copy(estilo_i11.border)
						cell_i.fill = copy(estilo_i11.fill)
						cell_i.number_format = copy(estilo_i11.number_format)
						cell_i.protection = copy(estilo_i11.protection)
						cell_i.alignment = copy(estilo_i11.alignment)
					
					# Coluna J: Total de jantar interno
					cell_j = ws_resumo.cell(row=linha_atual, column=10, value=totais_jantar_interno.get(nome_unidade_atual, 0))
					if estilo_j11.has_style:
						cell_j.font = copy(estilo_j11.font)
						cell_j.border = copy(estilo_j11.border)
						cell_j.fill = copy(estilo_j11.fill)
						cell_j.number_format = copy(estilo_j11.number_format)
						cell_j.protection = copy(estilo_j11.protection)
						cell_j.alignment = copy(estilo_j11.alignment)
					
					# Coluna K: Total de jantar funcionário
					cell_k = ws_resumo.cell(row=linha_atual, column=11, value=totais_jantar_funcionario.get(nome_unidade_atual, 0))
					if estilo_k11.has_style:
						cell_k.font = copy(estilo_k11.font)
						cell_k.border = copy(estilo_k11.border)
						cell_k.fill = copy(estilo_k11.fill)
						cell_k.number_format = copy(estilo_k11.number_format)
						cell_k.protection = copy(estilo_k11.protection)
						cell_k.alignment = copy(estilo_k11.alignment)
				
				# Calcular linha para totais (3 linhas abaixo da última unidade)
				linha_totais = 11 + quantidade_unidades + 2  # +2 porque queremos 3 células abaixo (11+n = última, +1 pula uma, +2 pula duas)
				
				# Linha inicial e final para a fórmula de soma
				linha_inicial_dados = 11
				linha_final_dados = 11 + quantidade_unidades - 1
				
				# Linha de preços unitários (sempre será 13 após inserção de linhas, mas vamos calcular dinamicamente)
				linha_precos = 11 + quantidade_unidades + 1  # Uma linha abaixo da última unidade
				
				# Linha de valores parciais (4 linhas abaixo da última unidade)
				linha_valores_parciais = 11 + quantidade_unidades + 3  # +3 porque queremos 4 células abaixo
				
				# Linha de valor total (5 linhas abaixo da última unidade)
				linha_valor_total = 11 + quantidade_unidades + 4  # +4 porque queremos 5 células abaixo
				
				# Pegar estilos das células da linha de totais (vamos usar os mesmos estilos das colunas)
				# Coluna D: Soma de café interno
				cell_d_total = ws_resumo.cell(row=linha_totais, column=4)
				cell_d_total.value = f'=SUM(D{linha_inicial_dados}:D{linha_final_dados})'
				if estilo_d14.has_style:
					cell_d_total.font = copy(estilo_d14.font)
					cell_d_total.border = copy(estilo_d14.border)
					cell_d_total.fill = copy(estilo_d14.fill)
					cell_d_total.number_format = copy(estilo_d14.number_format)
					cell_d_total.protection = copy(estilo_d14.protection)
					cell_d_total.alignment = copy(estilo_d14.alignment)
				
				# Coluna E: Soma de café funcionário
				cell_e_total = ws_resumo.cell(row=linha_totais, column=5)
				cell_e_total.value = f'=SUM(E{linha_inicial_dados}:E{linha_final_dados})'
				if estilo_e14.has_style:
					cell_e_total.font = copy(estilo_e14.font)
					cell_e_total.border = copy(estilo_e14.border)
					cell_e_total.fill = copy(estilo_e14.fill)
					cell_e_total.number_format = copy(estilo_e14.number_format)
					cell_e_total.protection = copy(estilo_e14.protection)
					cell_e_total.alignment = copy(estilo_e14.alignment)
				
				# Coluna F: Soma de almoço interno
				cell_f_total = ws_resumo.cell(row=linha_totais, column=6)
				cell_f_total.value = f'=SUM(F{linha_inicial_dados}:F{linha_final_dados})'
				if estilo_f14.has_style:
					cell_f_total.font = copy(estilo_f14.font)
					cell_f_total.border = copy(estilo_f14.border)
					cell_f_total.fill = copy(estilo_f14.fill)
					cell_f_total.number_format = copy(estilo_f14.number_format)
					cell_f_total.protection = copy(estilo_f14.protection)
					cell_f_total.alignment = copy(estilo_f14.alignment)
				
				# Coluna G: Soma de almoço funcionário
				cell_g_total = ws_resumo.cell(row=linha_totais, column=7)
				cell_g_total.value = f'=SUM(G{linha_inicial_dados}:G{linha_final_dados})'
				if estilo_g14.has_style:
					cell_g_total.font = copy(estilo_g14.font)
					cell_g_total.border = copy(estilo_g14.border)
					cell_g_total.fill = copy(estilo_g14.fill)
					cell_g_total.number_format = copy(estilo_g14.number_format)
					cell_g_total.protection = copy(estilo_g14.protection)
					cell_g_total.alignment = copy(estilo_g14.alignment)
				
				# Coluna H: Soma de lanche interno
				cell_h_total = ws_resumo.cell(row=linha_totais, column=8)
				cell_h_total.value = f'=SUM(H{linha_inicial_dados}:H{linha_final_dados})'
				if estilo_h14.has_style:
					cell_h_total.font = copy(estilo_h14.font)
					cell_h_total.border = copy(estilo_h14.border)
					cell_h_total.fill = copy(estilo_h14.fill)
					cell_h_total.number_format = copy(estilo_h14.number_format)
					cell_h_total.protection = copy(estilo_h14.protection)
					cell_h_total.alignment = copy(estilo_h14.alignment)
				
				# Coluna I: Soma de lanche funcionário
				cell_i_total = ws_resumo.cell(row=linha_totais, column=9)
				cell_i_total.value = f'=SUM(I{linha_inicial_dados}:I{linha_final_dados})'
				if estilo_i14.has_style:
					cell_i_total.font = copy(estilo_i14.font)
					cell_i_total.border = copy(estilo_i14.border)
					cell_i_total.fill = copy(estilo_i14.fill)
					cell_i_total.number_format = copy(estilo_i14.number_format)
					cell_i_total.protection = copy(estilo_i14.protection)
					cell_i_total.alignment = copy(estilo_i14.alignment)
				
				# Coluna J: Soma de jantar interno
				cell_j_total = ws_resumo.cell(row=linha_totais, column=10)
				cell_j_total.value = f'=SUM(J{linha_inicial_dados}:J{linha_final_dados})'
				if estilo_j14.has_style:
					cell_j_total.font = copy(estilo_j14.font)
					cell_j_total.border = copy(estilo_j14.border)
					cell_j_total.fill = copy(estilo_j14.fill)
					cell_j_total.number_format = copy(estilo_j14.number_format)
					cell_j_total.protection = copy(estilo_j14.protection)
					cell_j_total.alignment = copy(estilo_j14.alignment)
				
				# Coluna K: Soma de jantar funcionário
				cell_k_total = ws_resumo.cell(row=linha_totais, column=11)
				cell_k_total.value = f'=SUM(K{linha_inicial_dados}:K{linha_final_dados})'
				if estilo_k14.has_style:
					cell_k_total.font = copy(estilo_k14.font)
					cell_k_total.border = copy(estilo_k14.border)
					cell_k_total.fill = copy(estilo_k14.fill)
					cell_k_total.number_format = copy(estilo_k14.number_format)
					cell_k_total.protection = copy(estilo_k14.protection)
					cell_k_total.alignment = copy(estilo_k14.alignment)
				
				# Adicionar valores parciais usando FÓRMULAS (preço unitário × quantidade total)
				# Coluna D: Valor parcial café interno = D{linha_precos} * D{linha_totais}
				cell_d_parcial = ws_resumo.cell(row=linha_valores_parciais, column=4)
				cell_d_parcial.value = f'=D{linha_precos}*D{linha_totais}'
				if estilo_d15.has_style:
					cell_d_parcial.font = copy(estilo_d15.font)
					cell_d_parcial.border = copy(estilo_d15.border)
					cell_d_parcial.fill = copy(estilo_d15.fill)
					cell_d_parcial.number_format = copy(estilo_d15.number_format)
					cell_d_parcial.protection = copy(estilo_d15.protection)
					cell_d_parcial.alignment = copy(estilo_d15.alignment)
				
				# Coluna E: Valor parcial café funcionário = E{linha_precos} * E{linha_totais}
				cell_e_parcial = ws_resumo.cell(row=linha_valores_parciais, column=5)
				cell_e_parcial.value = f'=E{linha_precos}*E{linha_totais}'
				if estilo_e15.has_style:
					cell_e_parcial.font = copy(estilo_e15.font)
					cell_e_parcial.border = copy(estilo_e15.border)
					cell_e_parcial.fill = copy(estilo_e15.fill)
					cell_e_parcial.number_format = copy(estilo_e15.number_format)
					cell_e_parcial.protection = copy(estilo_e15.protection)
					cell_e_parcial.alignment = copy(estilo_e15.alignment)
				
				# Coluna F: Valor parcial almoço interno = F{linha_precos} * F{linha_totais}
				cell_f_parcial = ws_resumo.cell(row=linha_valores_parciais, column=6)
				cell_f_parcial.value = f'=F{linha_precos}*F{linha_totais}'
				if estilo_f15.has_style:
					cell_f_parcial.font = copy(estilo_f15.font)
					cell_f_parcial.border = copy(estilo_f15.border)
					cell_f_parcial.fill = copy(estilo_f15.fill)
					cell_f_parcial.number_format = copy(estilo_f15.number_format)
					cell_f_parcial.protection = copy(estilo_f15.protection)
					cell_f_parcial.alignment = copy(estilo_f15.alignment)
				
				# Coluna G: Valor parcial almoço funcionário = G{linha_precos} * G{linha_totais}
				cell_g_parcial = ws_resumo.cell(row=linha_valores_parciais, column=7)
				cell_g_parcial.value = f'=G{linha_precos}*G{linha_totais}'
				if estilo_g15.has_style:
					cell_g_parcial.font = copy(estilo_g15.font)
					cell_g_parcial.border = copy(estilo_g15.border)
					cell_g_parcial.fill = copy(estilo_g15.fill)
					cell_g_parcial.number_format = copy(estilo_g15.number_format)
					cell_g_parcial.protection = copy(estilo_g15.protection)
					cell_g_parcial.alignment = copy(estilo_g15.alignment)
				
				# Coluna H: Valor parcial lanche interno = H{linha_precos} * H{linha_totais}
				cell_h_parcial = ws_resumo.cell(row=linha_valores_parciais, column=8)
				cell_h_parcial.value = f'=H{linha_precos}*H{linha_totais}'
				if estilo_h15.has_style:
					cell_h_parcial.font = copy(estilo_h15.font)
					cell_h_parcial.border = copy(estilo_h15.border)
					cell_h_parcial.fill = copy(estilo_h15.fill)
					cell_h_parcial.number_format = copy(estilo_h15.number_format)
					cell_h_parcial.protection = copy(estilo_h15.protection)
					cell_h_parcial.alignment = copy(estilo_h15.alignment)
				
				# Coluna I: Valor parcial lanche funcionário = I{linha_precos} * I{linha_totais}
				cell_i_parcial = ws_resumo.cell(row=linha_valores_parciais, column=9)
				cell_i_parcial.value = f'=I{linha_precos}*I{linha_totais}'
				if estilo_i15.has_style:
					cell_i_parcial.font = copy(estilo_i15.font)
					cell_i_parcial.border = copy(estilo_i15.border)
					cell_i_parcial.fill = copy(estilo_i15.fill)
					cell_i_parcial.number_format = copy(estilo_i15.number_format)
					cell_i_parcial.protection = copy(estilo_i15.protection)
					cell_i_parcial.alignment = copy(estilo_i15.alignment)
				
				# Coluna J: Valor parcial jantar interno = J{linha_precos} * J{linha_totais}
				cell_j_parcial = ws_resumo.cell(row=linha_valores_parciais, column=10)
				cell_j_parcial.value = f'=J{linha_precos}*J{linha_totais}'
				if estilo_j15.has_style:
					cell_j_parcial.font = copy(estilo_j15.font)
					cell_j_parcial.border = copy(estilo_j15.border)
					cell_j_parcial.fill = copy(estilo_j15.fill)
					cell_j_parcial.number_format = copy(estilo_j15.number_format)
					cell_j_parcial.protection = copy(estilo_j15.protection)
					cell_j_parcial.alignment = copy(estilo_j15.alignment)
				
				# Coluna K: Valor parcial jantar funcionário = K{linha_precos} * K{linha_totais}
				cell_k_parcial = ws_resumo.cell(row=linha_valores_parciais, column=11)
				cell_k_parcial.value = f'=K{linha_precos}*K{linha_totais}'
				if estilo_k15.has_style:
					cell_k_parcial.font = copy(estilo_k15.font)
					cell_k_parcial.border = copy(estilo_k15.border)
					cell_k_parcial.fill = copy(estilo_k15.fill)
					cell_k_parcial.number_format = copy(estilo_k15.number_format)
					cell_k_parcial.protection = copy(estilo_k15.protection)
					cell_k_parcial.alignment = copy(estilo_k15.alignment)
				
				# Adicionar valor total (soma de todos os valores parciais D-K)
				# Mesclar células D:K na linha do valor total
				ws_resumo.merge_cells(start_row=linha_valor_total, start_column=4, end_row=linha_valor_total, end_column=11)
				
				# Criar fórmula de soma dos valores parciais
				cell_valor_total = ws_resumo.cell(row=linha_valor_total, column=4)
				cell_valor_total.value = f'=D{linha_valores_parciais}+E{linha_valores_parciais}+F{linha_valores_parciais}+G{linha_valores_parciais}+H{linha_valores_parciais}+I{linha_valores_parciais}+J{linha_valores_parciais}+K{linha_valores_parciais}'
				
				# Aplicar formatação da célula D16 original
				if estilo_d16.has_style:
					cell_valor_total.font = copy(estilo_d16.font)
					cell_valor_total.border = copy(estilo_d16.border)
					cell_valor_total.fill = copy(estilo_d16.fill)
					cell_valor_total.number_format = copy(estilo_d16.number_format)
					cell_valor_total.protection = copy(estilo_d16.protection)
					cell_valor_total.alignment = copy(estilo_d16.alignment)
				
				# Mesclar células B na coluna de preços até valor total
				# A linha inicial da mesclagem é sempre linha_precos (que começa em 11 + quantidade_unidades + 1)
				# A linha final é sempre linha_valor_total
				linha_inicial_mescla_b = linha_precos
				linha_final_mescla_b = linha_valor_total
				ws_resumo.merge_cells(start_row=linha_inicial_mescla_b, start_column=2, end_row=linha_final_mescla_b, end_column=2)
				
				# Aplicar formatação da célula B13 original à célula mesclada
				cell_b_mesclada = ws_resumo.cell(row=linha_inicial_mescla_b, column=2)
				if estilo_b13.has_style:
					cell_b_mesclada.font = copy(estilo_b13.font)
					cell_b_mesclada.border = copy(estilo_b13.border)
					cell_b_mesclada.fill = copy(estilo_b13.fill)
					cell_b_mesclada.number_format = copy(estilo_b13.number_format)
					cell_b_mesclada.protection = copy(estilo_b13.protection)
					cell_b_mesclada.alignment = copy(estilo_b13.alignment)
				
				if quantidade_unidades == 1:
					# Processamento para 1 unidade (código simplificado - mantido do original)
					pass
				else:
					# Processamento para múltiplas unidades (código simplificado - mantido do original)
					pass
			
			# Continua com o preenchimento da planilha COMPARATIVO
			# Determinar qual tabela de preços usar para este mês
			# Se todos os mapas do mês são do predecessor, usar preços do predecessor
			todos_predecessor = all(m.get('_is_predecessor', False) for m in mapas_do_mes)
			precos_para_planilha = precos_predecessor if todos_predecessor else precos
			
			col_inicio = 13
			for idx, (ref, tipo) in enumerate(precos_ordem):
				col = col_inicio + idx
				valor_preco = precos_para_planilha.get(ref, {}).get(tipo, 0)
				cell_preco = ws1.cell(row=6, column=col, value=valor_preco)
			
			# Adicionar fórmulas SUBTOTAL na linha 5 (colunas E até T)
			ws1['E5'] = '=SUBTOTAL(9,E12:E10000)'
			ws1['F5'] = '=SUBTOTAL(9,F12:F10000)'
			ws1['G5'] = '=SUBTOTAL(9,G12:G10000)'
			ws1['H5'] = '=SUBTOTAL(9,H12:H10000)'
			ws1['I5'] = '=SUBTOTAL(9,I12:I10000)'
			ws1['J5'] = '=SUBTOTAL(9,J12:J10000)'
			ws1['K5'] = '=SUBTOTAL(9,K12:K10000)'
			ws1['L5'] = '=SUBTOTAL(9,L12:L10000)'
			ws1['M5'] = '=SUBTOTAL(9,M12:M10000)'
			ws1['N5'] = '=SUBTOTAL(9,N12:N10000)'
			ws1['O5'] = '=SUBTOTAL(9,O12:O10000)'
			ws1['P5'] = '=SUBTOTAL(9,P12:P10000)'
			ws1['Q5'] = '=SUBTOTAL(9,Q12:Q10000)'
			ws1['R5'] = '=SUBTOTAL(9,R12:R10000)'
			ws1['S5'] = '=SUBTOTAL(9,S12:S10000)'
			ws1['T5'] = '=SUBTOTAL(9,T12:T10000)'
	
			# Processa dados e preenche células
			linha = 12
			
			# Pegar estilo das células da linha 12 como referência
			estilo_a12 = ws1['A12']
			estilo_b12 = ws1['B12']
			estilo_c12 = ws1['C12']
			estilo_d12 = ws1['D12']
			estilo_e12 = ws1['E12']
			estilo_m12 = ws1['M12']
			estilo_n12 = ws1['N12']
			estilo_o12 = ws1['O12']
			estilo_p12 = ws1['P12']
			estilo_q12 = ws1['Q12']
			estilo_r12 = ws1['R12']
			estilo_s12 = ws1['S12']
			estilo_t12 = ws1['T12']
			
			# Remover preenchimento cinza das células M12-T12 para permitir formatação condicional
			from openpyxl.styles import PatternFill
			no_fill = PatternFill(fill_type=None)
			for col in range(13, 21):  # Colunas M(13) até T(20)
				ws1.cell(row=12, column=col).fill = no_fill
			
			# Aplicar formatação condicional na linha 12 (M12-T12)
			green_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
			red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
			blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
			
			# Todas as colunas M, N, O, P, Q, R, S, T: Verde para "OK", Azul para 1-5, Vermelho para >5
			# Aplicar regras na ordem inversa (última adicionada = primeira avaliada)
			colunas_comparativo_linha12 = ['M12', 'N12', 'O12', 'P12', 'Q12', 'R12', 'S12', 'T12']
			for celula in colunas_comparativo_linha12:
				# Criar todas as regras
				rule_red = CellIsRule(operator='greaterThan', formula=['5'], fill=red_fill, stopIfTrue=True)
				rule_blue = CellIsRule(operator='between', formula=['1', '5'], fill=blue_fill, stopIfTrue=True)
				rule_green = CellIsRule(operator='equal', formula=['"OK"'], fill=green_fill, stopIfTrue=True)
				
				# Adicionar na ordem inversa: green (OK), blue (1-5), red (>5)
				ws1.conditional_formatting.add(celula, rule_green)
				ws1.conditional_formatting.add(celula, rule_blue)
				ws1.conditional_formatting.add(celula, rule_red)
			
			tem_dados = False
			for mapa in mapas_do_mes:
				unidade_nome = (mapa.get('unidade') or '').strip()
				dados_siisp = mapa.get('dados_siisp', [])
				datas = mapa.get('datas', [])
				cafe_interno = mapa.get('cafe_interno', [])
				cafe_funcionario = mapa.get('cafe_funcionario', [])
				almoco_interno = mapa.get('almoco_interno', [])
				almoco_funcionario = mapa.get('almoco_funcionario', [])
				lanche_interno = mapa.get('lanche_interno', [])
				lanche_funcionario = mapa.get('lanche_funcionario', [])
				jantar_interno = mapa.get('jantar_interno', [])
				jantar_funcionario = mapa.get('jantar_funcionario', [])
				
				for i in range(len(datas)):
					# Coluna A: Nome do lote (copiando formatação de A12)
					cell_a = ws1.cell(row=linha, column=1, value=lote_nome)
					if estilo_a12.has_style:
						cell_a.font = copy(estilo_a12.font)
						cell_a.border = copy(estilo_a12.border)
						cell_a.fill = copy(estilo_a12.fill)
						cell_a.number_format = copy(estilo_a12.number_format)
						cell_a.protection = copy(estilo_a12.protection)
						cell_a.alignment = copy(estilo_a12.alignment)
					
					# Coluna B: Nome da unidade (copiando formatação de B12)
					cell_b = ws1.cell(row=linha, column=2, value=unidade_nome)
					if estilo_b12.has_style:
						cell_b.font = copy(estilo_b12.font)
						cell_b.border = copy(estilo_b12.border)
						cell_b.fill = copy(estilo_b12.fill)
						cell_b.number_format = copy(estilo_b12.number_format)
						cell_b.protection = copy(estilo_b12.protection)
						cell_b.alignment = copy(estilo_b12.alignment)
					
					# Coluna C: Dados SIISP (copiando formatação de C12)
					siisp_valor = dados_siisp[i] if i < len(dados_siisp) else 0
					cell_c = ws1.cell(row=linha, column=3, value=siisp_valor)
					if estilo_c12.has_style:
						cell_c.font = copy(estilo_c12.font)
						cell_c.border = copy(estilo_c12.border)
						cell_c.fill = copy(estilo_c12.fill)
						cell_c.number_format = copy(estilo_c12.number_format)
						cell_c.protection = copy(estilo_c12.protection)
						cell_c.alignment = copy(estilo_c12.alignment)
					
					# Coluna D: Data (copiando formatação de D12)
					data_valor = datas[i] if i < len(datas) else ''
					cell_d = ws1.cell(row=linha, column=4, value=data_valor)
					if estilo_d12.has_style:
						cell_d.font = copy(estilo_d12.font)
						cell_d.border = copy(estilo_d12.border)
						cell_d.fill = copy(estilo_d12.fill)
						cell_d.number_format = copy(estilo_d12.number_format)
						cell_d.protection = copy(estilo_d12.protection)
						cell_d.alignment = copy(estilo_d12.alignment)
					
					# Coluna E: Café interno (copiando formatação de E12)
					cafe_int_valor = cafe_interno[i] if i < len(cafe_interno) else 0
					cell_e = ws1.cell(row=linha, column=5, value=cafe_int_valor)
					if estilo_e12.has_style:
						cell_e.font = copy(estilo_e12.font)
						cell_e.border = copy(estilo_e12.border)
						cell_e.fill = copy(estilo_e12.fill)
						cell_e.number_format = copy(estilo_e12.number_format)
						cell_e.protection = copy(estilo_e12.protection)
						cell_e.alignment = copy(estilo_e12.alignment)
					
					# Coluna F: Café funcionário (copiando formatação de E12)
					cafe_func_valor = cafe_funcionario[i] if i < len(cafe_funcionario) else 0
					cell_f = ws1.cell(row=linha, column=6, value=cafe_func_valor)
					if estilo_e12.has_style:
						cell_f.font = copy(estilo_e12.font)
						cell_f.border = copy(estilo_e12.border)
						cell_f.fill = copy(estilo_e12.fill)
						cell_f.number_format = copy(estilo_e12.number_format)
						cell_f.protection = copy(estilo_e12.protection)
						cell_f.alignment = copy(estilo_e12.alignment)
					
					# Coluna G: Almoço interno (copiando formatação de E12)
					almoco_int_valor = almoco_interno[i] if i < len(almoco_interno) else 0
					cell_g = ws1.cell(row=linha, column=7, value=almoco_int_valor)
					if estilo_e12.has_style:
						cell_g.font = copy(estilo_e12.font)
						cell_g.border = copy(estilo_e12.border)
						cell_g.fill = copy(estilo_e12.fill)
						cell_g.number_format = copy(estilo_e12.number_format)
						cell_g.protection = copy(estilo_e12.protection)
						cell_g.alignment = copy(estilo_e12.alignment)
					
					# Coluna H: Almoço funcionário (copiando formatação de E12)
					almoco_func_valor = almoco_funcionario[i] if i < len(almoco_funcionario) else 0
					cell_h = ws1.cell(row=linha, column=8, value=almoco_func_valor)
					if estilo_e12.has_style:
						cell_h.font = copy(estilo_e12.font)
						cell_h.border = copy(estilo_e12.border)
						cell_h.fill = copy(estilo_e12.fill)
						cell_h.number_format = copy(estilo_e12.number_format)
						cell_h.protection = copy(estilo_e12.protection)
						cell_h.alignment = copy(estilo_e12.alignment)
					
					# Coluna I: Lanche interno (copiando formatação de E12)
					lanche_int_valor = lanche_interno[i] if i < len(lanche_interno) else 0
					cell_i = ws1.cell(row=linha, column=9, value=lanche_int_valor)
					if estilo_e12.has_style:
						cell_i.font = copy(estilo_e12.font)
						cell_i.border = copy(estilo_e12.border)
						cell_i.fill = copy(estilo_e12.fill)
						cell_i.number_format = copy(estilo_e12.number_format)
						cell_i.protection = copy(estilo_e12.protection)
						cell_i.alignment = copy(estilo_e12.alignment)
					
					# Coluna J: Lanche funcionário (copiando formatação de E12)
					lanche_func_valor = lanche_funcionario[i] if i < len(lanche_funcionario) else 0
					cell_j = ws1.cell(row=linha, column=10, value=lanche_func_valor)
					if estilo_e12.has_style:
						cell_j.font = copy(estilo_e12.font)
						cell_j.border = copy(estilo_e12.border)
						cell_j.fill = copy(estilo_e12.fill)
						cell_j.number_format = copy(estilo_e12.number_format)
						cell_j.protection = copy(estilo_e12.protection)
						cell_j.alignment = copy(estilo_e12.alignment)
					
					# Coluna K: Jantar interno (copiando formatação de E12)
					jantar_int_valor = jantar_interno[i] if i < len(jantar_interno) else 0
					cell_k = ws1.cell(row=linha, column=11, value=jantar_int_valor)
					if estilo_e12.has_style:
						cell_k.font = copy(estilo_e12.font)
						cell_k.border = copy(estilo_e12.border)
						cell_k.fill = copy(estilo_e12.fill)
						cell_k.number_format = copy(estilo_e12.number_format)
						cell_k.protection = copy(estilo_e12.protection)
						cell_k.alignment = copy(estilo_e12.alignment)
					
					# Coluna L: Jantar funcionário (copiando formatação de E12)
					jantar_func_valor = jantar_funcionario[i] if i < len(jantar_funcionario) else 0
					cell_l = ws1.cell(row=linha, column=12, value=jantar_func_valor)
					if estilo_e12.has_style:
						cell_l.font = copy(estilo_e12.font)
						cell_l.border = copy(estilo_e12.border)
						cell_l.fill = copy(estilo_e12.fill)
						cell_l.number_format = copy(estilo_e12.number_format)
						cell_l.protection = copy(estilo_e12.protection)
						cell_l.alignment = copy(estilo_e12.alignment)
					
					# Coluna M: Fórmula =IF(E{linha}<=C{linha},"OK",E{linha}-C{linha})
					cell_m = ws1.cell(row=linha, column=13)
					cell_m.value = f'=IF(E{linha}<=C{linha},"OK",E{linha}-C{linha})'
					if estilo_m12.has_style:
						cell_m.font = copy(estilo_m12.font)
						cell_m.border = copy(estilo_m12.border)
						cell_m.fill = copy(estilo_m12.fill)
						cell_m.number_format = 'General'
						cell_m.protection = copy(estilo_m12.protection)
						cell_m.alignment = copy(estilo_m12.alignment)
					
					# Coluna N: Fórmula =IF(F{linha}<=C{linha},"OK",F{linha}-C{linha})
					cell_n = ws1.cell(row=linha, column=14)
					cell_n.value = f'=IF(F{linha}<=C{linha},"OK",F{linha}-C{linha})'
					if estilo_n12.has_style:
						cell_n.font = copy(estilo_n12.font)
						cell_n.border = copy(estilo_n12.border)
						cell_n.fill = copy(estilo_n12.fill)
						cell_n.number_format = 'General'
						cell_n.protection = copy(estilo_n12.protection)
						cell_n.alignment = copy(estilo_n12.alignment)
					
					# Coluna O: Fórmula =IF(G{linha}<=C{linha},"OK",G{linha}-C{linha})
					cell_o = ws1.cell(row=linha, column=15)
					cell_o.value = f'=IF(G{linha}<=C{linha},"OK",G{linha}-C{linha})'
					if estilo_o12.has_style:
						cell_o.font = copy(estilo_o12.font)
						cell_o.border = copy(estilo_o12.border)
						cell_o.fill = copy(estilo_o12.fill)
						cell_o.number_format = 'General'
						cell_o.protection = copy(estilo_o12.protection)
						cell_o.alignment = copy(estilo_o12.alignment)
					
					# Coluna P: Fórmula =IF(H{linha}<=C{linha},"OK",H{linha}-C{linha})
					cell_p = ws1.cell(row=linha, column=16)
					cell_p.value = f'=IF(H{linha}<=C{linha},"OK",H{linha}-C{linha})'
					if estilo_p12.has_style:
						cell_p.font = copy(estilo_p12.font)
						cell_p.border = copy(estilo_p12.border)
						cell_p.fill = copy(estilo_p12.fill)
						cell_p.number_format = 'General'
						cell_p.protection = copy(estilo_p12.protection)
						cell_p.alignment = copy(estilo_p12.alignment)
					
					# Coluna Q: Fórmula =IF(I{linha}<=C{linha},"OK",I{linha}-C{linha})
					cell_q = ws1.cell(row=linha, column=17)
					cell_q.value = f'=IF(I{linha}<=C{linha},"OK",I{linha}-C{linha})'
					if estilo_q12.has_style:
						cell_q.font = copy(estilo_q12.font)
						cell_q.border = copy(estilo_q12.border)
						cell_q.fill = copy(estilo_q12.fill)
						cell_q.number_format = 'General'
						cell_q.protection = copy(estilo_q12.protection)
						cell_q.alignment = copy(estilo_q12.alignment)
					
					# Coluna R: Fórmula =IF(J{linha}<=C{linha},"OK",J{linha}-C{linha})
					cell_r = ws1.cell(row=linha, column=18)
					cell_r.value = f'=IF(J{linha}<=C{linha},"OK",J{linha}-C{linha})'
					if estilo_r12.has_style:
						cell_r.font = copy(estilo_r12.font)
						cell_r.border = copy(estilo_r12.border)
						cell_r.fill = copy(estilo_r12.fill)
						cell_r.number_format = 'General'
						cell_r.protection = copy(estilo_r12.protection)
						cell_r.alignment = copy(estilo_r12.alignment)
					
					# Coluna S: Fórmula =IF(K{linha}<=C{linha},"OK",K{linha}-C{linha})
					cell_s = ws1.cell(row=linha, column=19)
					cell_s.value = f'=IF(K{linha}<=C{linha},"OK",K{linha}-C{linha})'
					if estilo_s12.has_style:
						cell_s.font = copy(estilo_s12.font)
						cell_s.border = copy(estilo_s12.border)
						cell_s.fill = copy(estilo_s12.fill)
						cell_s.number_format = 'General'
						cell_s.protection = copy(estilo_s12.protection)
						cell_s.alignment = copy(estilo_s12.alignment)
					
					# Coluna T: Fórmula =IF(L{linha}<=C{linha},"OK",L{linha}-C{linha})
					cell_t = ws1.cell(row=linha, column=20)
					cell_t.value = f'=IF(L{linha}<=C{linha},"OK",L{linha}-C{linha})'
					if estilo_t12.has_style:
						cell_t.font = copy(estilo_t12.font)
						cell_t.border = copy(estilo_t12.border)
						cell_t.fill = copy(estilo_t12.fill)
						cell_t.number_format = 'General'
						cell_t.protection = copy(estilo_t12.protection)
						cell_t.alignment = copy(estilo_t12.alignment)
					
					# Aplicar formatação condicional nas células M-T desta linha
					green_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
					red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
					blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
					
					# Aplicar formatação condicional para todas as colunas M-T
					# Verde: "OK", Azul: 1-5, Vermelho: >5
					colunas_comparativo = ['M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']
					for col in colunas_comparativo:
						celula = f'{col}{linha}'
						# Criar todas as regras
						rule_red = CellIsRule(operator='greaterThan', formula=['5'], fill=red_fill, stopIfTrue=True)
						rule_blue = CellIsRule(operator='between', formula=['1', '5'], fill=blue_fill, stopIfTrue=True)
						rule_green = CellIsRule(operator='equal', formula=['"OK"'], fill=green_fill, stopIfTrue=True)
						
						# Adicionar na ordem inversa: green primeiro (maior prioridade), depois blue, depois red
						ws1.conditional_formatting.add(celula, rule_green)
						ws1.conditional_formatting.add(celula, rule_blue)
						ws1.conditional_formatting.add(celula, rule_red)
					
					linha += 1
					tem_dados = True
			if not tem_dados:
				return {'success': False, 'error': 'Nenhum dado para exportar'}
		
		# Remover as planilhas modelo temporárias
		if ws1_modelo and ws1_modelo.title in wb.sheetnames:
			wb.remove(ws1_modelo)
		if ws_resumo_modelo and ws_resumo_modelo.title in wb.sheetnames:
			wb.remove(ws_resumo_modelo)
		
		output = io.BytesIO()
		wb.save(output)
		output.seek(0)

		nome_arquivo = f"tabela_lote_{lote_id}"
		if data_inicio and data_fim:
			nome_arquivo += f"_{data_inicio}_a_{data_fim}"
		nome_arquivo += ".xlsx"

		return {
			'success': True,
			'output': output,
			'filename': nome_arquivo
		}
	
	except Exception as e:
		import traceback
		traceback.print_exc()
		return {'success': False, 'error': f'Erro ao gerar arquivo: {str(e)}'}


def gerar_excel_exportacao_multiplos_lotes(data_inicio, data_fim):
	"""
	Gera arquivo Excel com dados de TODOS os lotes para um período específico.
	- COMPARATIVOS: Um por lote (COMPARATIVO LOTE A, COMPARATIVO LOTE B, etc)
	- RESUMOS: Agrupados por empresa/subempresa/delegacia
	"""
	try:
		from openpyxl import Workbook
		from openpyxl.formatting.rule import CellIsRule
		from openpyxl.styles import PatternFill, Border, Side
		from copy import copy
	except ImportError as e:
		return {'success': False, 'error': f'Biblioteca não instalada: {str(e)}'}
	
	try:
		dashboard_data = carregar_lotes_para_dashboard()
		lotes = dashboard_data.get('lotes', [])
		
		if not lotes:
			return {'success': False, 'error': 'Nenhum lote encontrado'}
		
		# Filtrar apenas lotes ATIVOS
		lotes_ativos = [lote for lote in lotes if lote.get('ativo', True)]
		
		if not lotes_ativos:
			return {'success': False, 'error': 'Nenhum lote ativo encontrado'}
		
		print(f"📊 Total de lotes ativos: {len(lotes_ativos)}")
		
		# Converter strings de data para datetime
		data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
		data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
		print(f"📅 Exportando todos os lotes ativos - Período: {data_inicio} a {data_fim}")
		
		# Criar um workbook novo
		wb = Workbook()
		wb.remove(wb.active)  # Remove a planilha padrão
		
		# Dicionários para agrupar os dados dos lotes
		comparativos = {}  # {lote_nome: wb_lote}
		resumos_por_categoria = {
			'empresa': {},      # {empresa_nome: [(lote_nome, ws_resumo), ...]}
			'sub_empresa': {},  # {sub_empresa_nome: [(lote_nome, ws_resumo), ...]}
			'delegacia': []     # [(lote_nome, ws_resumo), ...]
		}
		
		# Variável para rastrear se pelo menos um lote foi exportado
		algum_lote_exportado = False
		
		# Iterar sobre cada lote ATIVO e agrupar suas planilhas
		for lote in lotes_ativos:
			lote_id = lote.get('id')
			lote_nome = lote.get('nome', f'Lote {lote_id}')
			
			print(f"\n📊 Processando {lote_nome} (ID: {lote_id})")
			
			# Gerar dados para este lote usando a função existente
			resultado_lote = gerar_excel_exportacao(lote_id, [], data_inicio, data_fim)
			
			if not resultado_lote.get('success'):
				print(f"⚠️ Erro ao gerar dados para {lote_nome}: {resultado_lote.get('error')}")
				continue
			
			# Carregar o workbook gerado para este lote
			from openpyxl import load_workbook
			resultado_lote['output'].seek(0)
			wb_lote = load_workbook(resultado_lote['output'])
			
			# Separar COMPARATIVOS e RESUMOS
			for sheet_name in wb_lote.sheetnames:
				ws_origem = wb_lote[sheet_name]
				
				if 'COMPARATIVO' in sheet_name:
					# Armazenar COMPARATIVO com nome do lote
					comparativos[f"COMPARATIVO {lote_nome}"] = ws_origem
					print(f"  ✓ COMPARATIVO armazenado: COMPARATIVO {lote_nome}")
					
				elif 'RESUMO' in sheet_name:
					# Determinar categoria do RESUMO pelo nome da sheet
					if 'DELEGACIA' in sheet_name:
						# RESUMO DELEGACIA - tudo junto
						resumos_por_categoria['delegacia'].append((lote_nome, ws_origem))
						print(f"  ✓ RESUMO DELEGACIA armazenado: {lote_nome}")
						
					elif 'SUBEMPRESA' in sheet_name or 'SUBEMPRESA' in sheet_name or '(' in sheet_name:
						# RESUMO (SUBEMPRESA) - extrair nome da subempresa do nome da sheet
						# Formato: "RESUMO (Nome_da_Subempresa)" ou "RESUMO (Nome_da_Subempresa) - MÊS"
						import re
						match = re.search(r'RESUMO\s*\(([^)]+)\)', sheet_name)
						if match:
							sub_empresa_nome = match.group(1).strip()
						else:
							sub_empresa_nome = 'Sub-Empresa Padrão'
						
						if sub_empresa_nome not in resumos_por_categoria['sub_empresa']:
							resumos_por_categoria['sub_empresa'][sub_empresa_nome] = []
						resumos_por_categoria['sub_empresa'][sub_empresa_nome].append((lote_nome, ws_origem))
						print(f"  ✓ RESUMO (SUBEMPRESA) armazenado ({sub_empresa_nome}): {lote_nome}")
						
					else:
						# RESUMO padrão - agrupar por empresa do lote
						empresa_nome = lote.get('empresa', 'Empresa Padrão')
						if empresa_nome not in resumos_por_categoria['empresa']:
							resumos_por_categoria['empresa'][empresa_nome] = []
						resumos_por_categoria['empresa'][empresa_nome].append((lote_nome, ws_origem))
						print(f"  ✓ RESUMO armazenado ({empresa_nome}): {lote_nome}")
			
			algum_lote_exportado = True
		
		if not algum_lote_exportado:
			return {'success': False, 'error': 'Nenhum lote com dados para o período selecionado'}
		
		# FASE 2: Adicionar COMPARATIVOS primeiro (um por lote)
		print("\n📋 Adicionando COMPARATIVOS...")
		for nome_comparativo, ws_comparativo in comparativos.items():
			_copiar_sheet_para_workbook(wb, ws_comparativo, nome_comparativo)
			print(f"  ✓ {nome_comparativo} adicionado")
		
		# FASE 3: Adicionar RESUMOS agrupados por categoria
		print("\n📋 Adicionando RESUMOS agrupados...")
		
		# RESUMOS por Empresa
		for empresa_nome, resumos in resumos_por_categoria['empresa'].items():
			if resumos:
				nome_sheet = f"RESUMO - {empresa_nome}"
				_criar_resumo_agrupado(wb, resumos, nome_sheet)
				print(f"  ✓ {nome_sheet} criado com {len(resumos)} lote(s)")
		
		# RESUMOS por Sub-Empresa
		for sub_empresa_nome, resumos in resumos_por_categoria['sub_empresa'].items():
			if resumos:
				nome_sheet = f"RESUMO - {sub_empresa_nome}"
				_criar_resumo_agrupado(wb, resumos, nome_sheet)
				print(f"  ✓ {nome_sheet} criado com {len(resumos)} lote(s)")
		
		# RESUMO DELEGACIA
		if resumos_por_categoria['delegacia']:
			_criar_resumo_agrupado(wb, resumos_por_categoria['delegacia'], "RESUMO DELEGACIA")
			print(f"  ✓ RESUMO DELEGACIA criado com {len(resumos_por_categoria['delegacia'])} lote(s)")
		
		# Salvar workbook consolidado
		output = io.BytesIO()
		wb.save(output)
		output.seek(0)
		
		# Extrair mês e ano do data_inicio para nome do arquivo
		mes_numero = data_inicio_dt.month
		ano = data_inicio_dt.year
		meses_pt = ['', 'janeiro', 'fevereiro', 'marco', 'abril', 'maio', 'junho',
		            'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro']
		mes_nome = meses_pt[mes_numero]
		
		nome_arquivo = f"exportacao_todos_lotes_{mes_nome}_{ano}.xlsx"
		
		print(f"\n✅ Arquivo consolidado gerado: {nome_arquivo}")
		print(f"📦 Total de planilhas: {len(wb.sheetnames)}")
		
		return {
			'success': True,
			'output': output,
			'filename': nome_arquivo
		}
	
	except Exception as e:
		print(f"❌ Erro na exportação de múltiplos lotes: {str(e)}")
		import traceback
		traceback.print_exc()
		return {'success': False, 'error': str(e)}


def _copiar_sheet_para_workbook(wb_destino, ws_origem, novo_nome):
	"""Copia uma sheet completamente com toda formatação"""
	from copy import copy
	
	# Garantir que o nome não ultrapasse 31 caracteres
	if len(novo_nome) > 31:
		novo_nome = novo_nome[:31]
	
	# Criar nova planilha
	ws_destino = wb_destino.create_sheet(title=novo_nome)
	
	# Copiar dados e formatação
	for row in ws_origem.iter_rows():
		for cell in row:
			ws_destino[cell.coordinate].value = cell.value
			
			# Copiar formatação
			if cell.has_style:
				ws_destino[cell.coordinate].font = copy(cell.font)
				ws_destino[cell.coordinate].border = copy(cell.border)
				ws_destino[cell.coordinate].fill = copy(cell.fill)
				ws_destino[cell.coordinate].number_format = copy(cell.number_format)
				ws_destino[cell.coordinate].protection = copy(cell.protection)
				ws_destino[cell.coordinate].alignment = copy(cell.alignment)
	
	# Copiar larguras das colunas
	for col_letter in ws_origem.column_dimensions:
		ws_destino.column_dimensions[col_letter].width = ws_origem.column_dimensions[col_letter].width
	
	# Copiar alturas das linhas
	for row_num in ws_origem.row_dimensions:
		ws_destino.row_dimensions[row_num].height = ws_origem.row_dimensions[row_num].height
	
	# Copiar células mescladas
	for merged_cell_range in ws_origem.merged_cells.ranges:
		ws_destino.merge_cells(str(merged_cell_range))
	
	# Copiar regras de formatação condicional
	try:
		for range_string, rules in ws_origem.conditional_formatting._cf_rules.items():
			for rule in rules:
				ws_destino.conditional_formatting.add(range_string, rule)
	except:
		pass
	
	return ws_destino


def _criar_resumo_agrupado(wb_destino, resumos_list, nome_sheet):
	"""
	Cria uma sheet única com múltiplos resumos, um embaixo do outro,
	separados por uma linha divisória.
	resumos_list: lista de tuplas (lote_nome, ws_resumo)
	"""
	from copy import copy
	from openpyxl.styles import Border, Side, Font, PatternFill
	
	# Garantir que o nome não ultrapasse 31 caracteres
	if len(nome_sheet) > 31:
		nome_sheet = nome_sheet[:31]
	
	# Criar nova planilha
	ws_destino = wb_destino.create_sheet(title=nome_sheet)
	
	linha_atual = 1
	primeiro = True
	
	for lote_nome, ws_origem in resumos_list:
		if not primeiro:
			# Adicionar linha divisória entre resumos
			linha_atual += 1  # Pular uma linha
			
			# Preencher toda a linha com cor de fundo (divisória)
			cinza_claro = "CCCCCC"
			linha_divisoria = linha_atual
			for col in range(1, 15):  # Colunas A até N
				cell = ws_destino.cell(row=linha_divisoria, column=col)
				cell.fill = PatternFill(start_color=cinza_claro, end_color=cinza_claro, fill_type="solid")
				cell.border = Border(
					top=Side(style='thin', color='000000'),
					bottom=Side(style='thin', color='000000'),
					left=Side(style='thin', color='000000'),
					right=Side(style='thin', color='000000')
				)
			
			linha_atual += 1
		
		# Copiar dados do resumo
		inicio_dados = linha_atual  # Guardar linha inicial para calcular offset dos merged cells
		linha_origem = 0  # Rastrear linha da origem
		for row in ws_origem.iter_rows():
			linha_origem += 1
			for col_idx, cell in enumerate(row, 1):
				cell_value = cell.value
				
				# Se é uma fórmula com SUM/SOMA, ajustar as referências de linha
				if isinstance(cell_value, str) and ('=SUM(' in cell_value.upper()):
					# Encontrar o padrão =SUM(COLUNA{linha1}:COLUNA{linha2})
					import re
					match = re.search(r'=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', cell_value, re.IGNORECASE)
					if match:
						col1 = match.group(1)
						lin1_orig = int(match.group(2))
						col2 = match.group(3)
						lin2_orig = int(match.group(4))
						
						# Calcular novo offset baseado na posição da origem
						offset = linha_atual - linha_origem
						lin1_novo = lin1_orig + offset
						lin2_novo = lin2_orig + offset
						
						# Reconstruir a fórmula
						cell_value = f'=SUM({col1}{lin1_novo}:{col2}{lin2_novo})'
				
				# Se é outra fórmula (não SUM), ajustar as referências simples como =D18*D19
				elif isinstance(cell_value, str) and cell_value.startswith('=') and 'SUM' not in cell_value.upper():
					import re
					
					# Calcular offset
					offset = linha_atual - linha_origem
					
					# Substituir referências simples como D18, D19, etc
					def ajustar_referencia(match):
						ref = match.group(0)  # ex: "D18"
						col = ''.join([c for c in ref if c.isalpha()])
						lin = int(''.join([c for c in ref if c.isdigit()]))
						nova_lin = lin + offset
						return f"{col}{nova_lin}"
					
					# Encontrar e substituir todas as referências de célula (letras + dígitos)
					cell_value = re.sub(r'[A-Z]+\d+', ajustar_referencia, cell_value)
				
				ws_destino.cell(row=linha_atual, column=col_idx).value = cell_value
				
				# Copiar formatação
				if cell.has_style:
					ws_destino.cell(row=linha_atual, column=col_idx).font = copy(cell.font)
					ws_destino.cell(row=linha_atual, column=col_idx).border = copy(cell.border)
					ws_destino.cell(row=linha_atual, column=col_idx).fill = copy(cell.fill)
					ws_destino.cell(row=linha_atual, column=col_idx).number_format = copy(cell.number_format)
					ws_destino.cell(row=linha_atual, column=col_idx).alignment = copy(cell.alignment)
			
			linha_atual += 1
		
		# Copiar células mescladas da sheet de origem, recalculando as linhas
		offset_linhas = inicio_dados - 1  # Offset para recalcular os números de linha
		for merged_range in ws_origem.merged_cells.ranges:
			# Parsear o range mesclado (ex: "B2:K2")
			range_str = str(merged_range)
			try:
				# Recalcular as linhas do range
				from openpyxl.utils import get_column_letter, column_index_from_string
				partes = range_str.split(':')
				celula_inicio = partes[0]  # ex: "B2"
				celula_fim = partes[1] if len(partes) > 1 else partes[0]  # ex: "K2"
				
				# Extrair coluna e linha
				col_inicio_str = ''.join([c for c in celula_inicio if c.isalpha()])
				lin_inicio = int(''.join([c for c in celula_inicio if c.isdigit()]))
				
				col_fim_str = ''.join([c for c in celula_fim if c.isalpha()])
				lin_fim = int(''.join([c for c in celula_fim if c.isdigit()]))
				
				# Recalcular linhas com offset
				lin_inicio_nova = lin_inicio + offset_linhas
				lin_fim_nova = lin_fim + offset_linhas
				
				# Criar novo range
				novo_range = f"{col_inicio_str}{lin_inicio_nova}:{col_fim_str}{lin_fim_nova}"
				ws_destino.merge_cells(novo_range)
			except Exception as e:
				print(f"⚠️ Erro ao copiar merged cell {range_str}: {e}")
		
		primeiro = False
	
	# Copiar larguras das colunas da primeira sheet
	if resumos_list:
		ws_primeira = resumos_list[0][1]
		for col_letter in ws_primeira.column_dimensions:
			ws_destino.column_dimensions[col_letter].width = ws_primeira.column_dimensions[col_letter].width
	
	return ws_destino
