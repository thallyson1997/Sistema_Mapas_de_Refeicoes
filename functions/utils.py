import json
import re
import os
from datetime import datetime
try:
	import bcrypt
except Exception:
	bcrypt = None
import calendar


# ----- SIISP Comparison Helpers -----
def _calcular_campos_comparativos_siisp(record):
	"""Calcula campos comparativos SIISP para um registro de mapa.
	
	Calcula:
	- cafe_interno_siisp = cafe_interno - dados_siisp
	- cafe_funcionario_siisp = cafe_funcionario - dados_siisp
	- almoco_interno_siisp = almoco_interno - dados_siisp
	- almoco_funcionario_siisp = almoco_funcionario - dados_siisp
	- lanche_interno_siisp = lanche_interno - dados_siisp
	- lanche_funcionario_siisp = lanche_funcionario - dados_siisp
	- jantar_interno_siisp = jantar_interno - dados_siisp
	- jantar_funcionario_siisp = jantar_funcionario - dados_siisp
	
	Modifica o record in-place adicionando esses campos.
	"""
	if not isinstance(record, dict):
		return
	
	# Campos de refeições
	meal_fields = [
		'cafe_interno', 'cafe_funcionario',
		'almoco_interno', 'almoco_funcionario', 
		'lanche_interno', 'lanche_funcionario',
		'jantar_interno', 'jantar_funcionario'
	]
	
	# Obter dados_siisp
	dados_siisp = record.get('dados_siisp', [])
	if not isinstance(dados_siisp, list):
		dados_siisp = []
	
	# Para cada campo de refeição, calcular a diferença
	for field in meal_fields:
		field_data = record.get(field, [])
		if not isinstance(field_data, list):
			field_data = []
		
		# Campo comparativo correspondente
		siisp_field = f"{field}_siisp"
		
		# Calcular diferenças
		comparativo = []
		max_len = max(len(field_data), len(dados_siisp))
		
		for i in range(max_len):
			# Valor da refeição (0 se índice não existir)
			refeicao_val = field_data[i] if i < len(field_data) else 0
			# Valor SIISP (0 se índice não existir)
			siisp_val = dados_siisp[i] if i < len(dados_siisp) else 0
			
			# Garantir que são números
			try:
				refeicao_num = int(refeicao_val) if refeicao_val is not None else 0
			except (ValueError, TypeError):
				refeicao_num = 0
			
			try:
				siisp_num = int(siisp_val) if siisp_val is not None else 0
			except (ValueError, TypeError):
				siisp_num = 0
			
			# Calcular diferença
			diferenca = refeicao_num - siisp_num
			comparativo.append(diferenca)
		
		# Adicionar campo comparativo ao record
		record[siisp_field] = comparativo


# ----- Security / Password helpers -----
def _hash_password(senha):
	"""Retorna o hash bcrypt da senha (string). Se bcrypt não estiver disponível, retorna a senha em texto (fallback)."""
	if not senha:
		return ''
	if bcrypt is None:
		# fallback — prefer instalar bcrypt
		return str(senha)
	try:
		pw = str(senha).encode('utf-8')
		hashed = bcrypt.hashpw(pw, bcrypt.gensalt())
		return hashed.decode('utf-8')
	except Exception:
		return str(senha)

# ----- Small form / utility helpers -----
def _first_present(form_data, *names):
	"""Retorna o primeiro valor presente em form_data para os nomes fornecidos.

	Se nenhum estiver presente, retorna None. Não faz strip nem normalização —
	quem chama decide como tratar o valor.
	"""
	if not isinstance(form_data, dict):
		return None
	for n in names:
		if n in form_data:
			return form_data.get(n)
	return None

# ----- User registration & persistence -----
def cadastrar_novo_usuario(form_data=None):
	r = validar_cadastro_no_usuario(form_data)
	if not r.get('valido'):
		return {'ok': False, 'mensagem': r.get('mensagem', 'Validação falhou'), 'campo': r.get('campo')}

	base_dir = os.path.dirname(os.path.dirname(__file__))
	usuarios_path = os.path.join(base_dir, 'dados', 'usuarios.json')
	usuarios = None
	data_wrapped = None
	try:
		data = _load_usuarios_data()
		if isinstance(data, list):
			usuarios = data
		elif isinstance(data, dict) and isinstance(data.get('usuarios'), list):
			usuarios = data.get('usuarios')
			data_wrapped = data
		else:
			usuarios = []
	except Exception:
		usuarios = []

	existing_ids = [u.get('id') for u in usuarios if isinstance(u, dict) and isinstance(u.get('id'), int)]
	new_id = (max(existing_ids) + 1) if existing_ids else 1

	registro = {
		'id': new_id,
		# data de criação do cadastro (ISO 8601)
		'data_criacao': datetime.now().isoformat(),
		'cpf': re.sub(r'\D', '', str(form_data.get('cpf') or '')),
		'email': str(form_data.get('email') or '').strip(),
		'telefone': re.sub(r'\D', '', str(form_data.get('telefone') or '')),
		'matricula': str(form_data.get('matricula') or '').strip(),
		'usuario': str(form_data.get('usuario') or '').strip(),
		'nome': str(form_data.get('nome') or form_data.get('nome_completo') or '').strip(),
		'cargo': str(form_data.get('cargo') or '').strip(),
		'unidade': str(form_data.get('unidade') or '').strip(),
		'motivo': str(
			_first_present(form_data, 'motivo', 'motivo_solicitacao', 'justificativa', 'justificativa_acesso') or ''
		).strip(),
		'concordo': False,
		'ativo': False,
		'senha': _hash_password(form_data.get('senha') or '')
	}

	# normalizar valor do checkbox "concordo" (vários nomes possíveis vindos do form)
	# normalizar aliases do checkbox de aceite: buscar o primeiro nome presente
	_concordo_raw = _first_present(
		form_data,
		'concordo',
		'concordo_termos',
		'aceito',
		'aceito_termos',
		'aceitarTermos',
		'aceitar_termos'
	)
	if _concordo_raw is not None:
		v = str(_concordo_raw).strip().lower()
		if v in ('1', 'true', 'on', 'yes', 'sim', 'y'):
			registro['concordo'] = True

	try:
		usuarios.append(registro)
		os.makedirs(os.path.dirname(usuarios_path), exist_ok=True)
		if data_wrapped is not None:
			data_wrapped['usuarios'] = usuarios
			to_write = data_wrapped
		else:
			to_write = usuarios
		tmp_path = usuarios_path + '.tmp'
		with open(tmp_path, 'w', encoding='utf-8') as f:
			json.dump(to_write, f, ensure_ascii=False, indent=2)
		os.replace(tmp_path, usuarios_path)
		return {'ok': True, 'mensagem': 'Usuário cadastrado com sucesso. Aguarde a aprovação do seu cadastro.', 'id': new_id}
	except Exception as e:
		try:
			print('Erro ao salvar usuário:', e)
		except Exception:
			pass
		return {'ok': False, 'mensagem': 'Erro ao salvar usuário'}


# ----- User validators & lookup -----
def validar_cpf(cpf):
	if not cpf:
		return {'valido': False, 'mensagem': 'CPF inválido'}
	num = re.sub(r'\D', '', str(cpf))

	if len(num) != 11:
		return {'valido': False, 'mensagem': 'CPF inválido'}
	if re.match(r'^(\d)\1{10}$', num):
		return {'valido': False, 'mensagem': 'CPF inválido'}

	s = 0
	for i in range(9):
		s += int(num[i]) * (10 - i)
	d1 = 11 - (s % 11)
	if d1 >= 10:
		d1 = 0
	if d1 != int(num[9]):
		return {'valido': False, 'mensagem': 'CPF inválido'}

	s = 0
	for i in range(10):
		s += int(num[i]) * (11 - i)
	d2 = 11 - (s % 11)
	if d2 >= 10:
		d2 = 0
	if d2 != int(num[10]):
		return {'valido': False, 'mensagem': 'CPF inválido'}

	if _exists_in_usuarios(num, normalize=lambda x: re.sub(r'\D', '', x)):
		return {'valido': False, 'mensagem': 'CPF já cadastrado'}

	return {'valido': True, 'mensagem': 'OK'}

def _load_usuarios_data():
	base_dir = os.path.dirname(os.path.dirname(__file__))
	usuarios_path = os.path.join(base_dir, 'dados', 'usuarios.json')
	if not os.path.isfile(usuarios_path):
		return None
	try:
		with open(usuarios_path, 'r', encoding='utf-8') as f:
			return json.load(f)
	except Exception:
		return None

def _exists_in_usuarios(target, normalize=lambda x: x, active_only=True):
	if target is None:
		return False
	data = _load_usuarios_data()
	if not data:
		return False

	def _search(obj):
		if isinstance(obj, dict):
			# se este dict representa um registro de usuário com flag 'ativo',
			# respeitar active_only: ignorar o registro quando ativo == False
			if active_only and ('ativo' in obj) and isinstance(obj.get('ativo'), bool) and not obj.get('ativo'):
				return False
			for v in obj.values():
				if _search(v):
					return True
		elif isinstance(obj, list):
			for item in obj:
				if _search(item):
					return True
		else:
			try:
				val = str(obj)
			except Exception:
				return False
			if normalize(val) == normalize(target):
				return True
		return False

	return _search(data)

def _find_user(login_value, active_only=True):
	"""Retorna o dicionário do usuário cujo email OU usuario (username) corresponde a login_value.

	A busca por email é case-insensitive; por usuário também. Se nenhum encontrado, retorna None.
	"""
	if login_value is None:
		return None
	data = _load_usuarios_data()
	if not data:
		return None
	# normalizar a lista de usuários
	usuarios = None
	if isinstance(data, list):
		usuarios = data
	elif isinstance(data, dict) and isinstance(data.get('usuarios'), list):
		usuarios = data.get('usuarios')
	else:
		return None

	lv = str(login_value).strip()
	is_email = ('@' in lv)
	for u in usuarios:
		if not isinstance(u, dict):
			continue
		if active_only and isinstance(u.get('ativo'), bool) and not u.get('ativo'):
			continue
		# comparar email
		email = u.get('email')
		if is_email and isinstance(email, str) and email.strip().lower() == lv.lower():
			return u
		# comparar username
		usuario = u.get('usuario')
		if (not is_email) and isinstance(usuario, str) and usuario.strip().lower() == lv.lower():
			return u
	return None

def _check_password(stored_pw, provided_pw):
	"""Verifica se `provided_pw` corresponde a `stored_pw`.

	- Se bcrypt estiver disponível e stored_pw parece ser um hash bcrypt, usa bcrypt.checkpw.
	- Caso contrário, faz comparação direta (fallback).
	"""
	if stored_pw is None:
		return False
	sp = str(stored_pw)
	if not provided_pw:
		return False
	pp = str(provided_pw)
	# detectar hash bcrypt comum que começa com $2b$ ou $2a$ ou $2y$
	if bcrypt is not None and sp.startswith('$2'):
		try:
			return bcrypt.checkpw(pp.encode('utf-8'), sp.encode('utf-8'))
		except Exception:
			return False
	# fallback inseguro
	return sp == pp

def validar_login(login_value, senha):
	"""Valida credenciais de login.

	Retorna um dict:
	  - {'ok': True, 'mensagem': 'OK', 'user': <user_sanitized>} em sucesso
	  - {'ok': False, 'mensagem': '...'} em falha

	Mensagens específicas:
	  - se input parece e-mail e não existe: 'E-mail não cadastrado'
	  - se input parece username e não existe: 'Usuário não cadastrado'
	  - se senha incorreta: 'Senha incorreta'
	"""
	if not login_value:
		return {'ok': False, 'mensagem': 'Informe usuário ou e-mail'}

	is_email = ('@' in str(login_value))
	user = _find_user(login_value, active_only=True)
	if not user:
		return {'ok': False, 'mensagem': 'E-mail não cadastrado' if is_email else 'Usuário não cadastrado'}

	stored = user.get('senha')
	if not _check_password(stored, senha):
		return {'ok': False, 'mensagem': 'Senha incorreta'}

	# sucesso: não retornar a senha
	sanitized = {k: v for k, v in user.items() if k != 'senha'}
	return {'ok': True, 'mensagem': 'Login efetuado com sucesso', 'user': sanitized}

def validar_email(email):
	if not email:
		return {'valido': False, 'mensagem': 'Email inválido'}
	email = email.strip()
	email_regex = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')
	if not email_regex.match(email):
		return {'valido': False, 'mensagem': 'Email inválido'}
	# verificar duplicidade (case-insensitive)
	if _exists_in_usuarios(email.lower(), normalize=lambda x: x.lower()):
		return {'valido': False, 'mensagem': 'Email já cadastrado'}
	return {'valido': True, 'mensagem': 'OK'}

def validar_telefone(telefone):
	if not telefone:
		return {'valido': False, 'mensagem': 'Telefone inválido'}
	num = re.sub(r'\D', '', str(telefone))
	if len(num) < 10 or len(num) > 11:
		return {'valido': False, 'mensagem': 'Telefone inválido'}
	if re.match(r'^(\d)\1{9,10}$', num):
		return {'valido': False, 'mensagem': 'Telefone inválido'}
	if _exists_in_usuarios(num, normalize=lambda x: re.sub(r'\D', '', x)):
		return {'valido': False, 'mensagem': 'Telefone já cadastrado'}
	return {'valido': True, 'mensagem': 'OK'}

def validar_matricula(matricula):
	if not matricula:
		return {'valido': False, 'mensagem': 'Matrícula inválida'}
	mat = str(matricula).strip()
	if _exists_in_usuarios(mat, normalize=lambda x: x.strip()):
		return {'valido': False, 'mensagem': 'Matrícula já cadastrada'}
	return {'valido': True, 'mensagem': 'OK'}

def validar_username(username):
	if not username:
		return {'valido': False, 'mensagem': 'Nome de usuário inválido'}
	user = str(username).strip()
	if _exists_in_usuarios(user.lower(), normalize=lambda x: x.lower()):
		return {'valido': False, 'mensagem': 'Nome de usuário já existe'}
	return {'valido': True, 'mensagem': 'OK'}

def validar_senha(senha, confirmar):
	if senha is None or confirmar is None:
		return {'valido': False, 'mensagem': 'Senha inválida'}
	if str(senha) != str(confirmar):
		return {'valido': False, 'mensagem': 'Senhas não coincidem'}
	return {'valido': True, 'mensagem': 'OK'}

def validar_cadastro_no_usuario(form_data):
	if not isinstance(form_data, dict):
		return {'valido': False, 'mensagem': 'Dados do formulário inválidos'}

	cpf = form_data.get('cpf')
	email = form_data.get('email')
	telefone = form_data.get('telefone')
	matricula = form_data.get('matricula')
	usuario = form_data.get('usuario')
	senha = form_data.get('senha')
	confirmar = form_data.get('confirmarSenha') or form_data.get('confirmar')

	r = validar_cpf(cpf)
	if not r.get('valido'):
		return {'valido': False, 'mensagem': r.get('mensagem', 'CPF inválido'), 'campo': 'cpf'}

	r = validar_email(email)
	if not r.get('valido'):
		return {'valido': False, 'mensagem': r.get('mensagem', 'Email inválido'), 'campo': 'email'}

	r = validar_telefone(telefone)
	if not r.get('valido'):
		return {'valido': False, 'mensagem': r.get('mensagem', 'Telefone inválido'), 'campo': 'telefone'}

	r = validar_matricula(matricula)
	if not r.get('valido'):
		return {'valido': False, 'mensagem': r.get('mensagem', 'Matrícula inválida'), 'campo': 'matricula'}

	r = validar_username(usuario)
	if not r.get('valido'):
		return {'valido': False, 'mensagem': r.get('mensagem', 'Nome de usuário inválido'), 'campo': 'usuario'}

	r = validar_senha(senha, confirmar)
	if not r.get('valido'):
		return {'valido': False, 'mensagem': r.get('mensagem', 'Senhas não coincidem'), 'campo': 'senha'}

	# Todas as validações passaram — não salvar aqui, apenas indicar sucesso
	return {'valido': True, 'mensagem': 'Validação OK'}

# ----- Lotes helpers -----
def _load_lotes_data():
	base_dir = os.path.dirname(os.path.dirname(__file__))
	lotes_path = os.path.join(base_dir, 'dados', 'lotes.json')
	if not os.path.isfile(lotes_path):
		return None
	try:
		with open(lotes_path, 'r', encoding='utf-8') as f:
			return json.load(f)
	except Exception:
		return None

def _save_lotes_data(data):
	base_dir = os.path.dirname(os.path.dirname(__file__))
	lotes_path = os.path.join(base_dir, 'dados', 'lotes.json')
	try:
		os.makedirs(os.path.dirname(lotes_path), exist_ok=True)
		tmp_path = lotes_path + '.tmp'
		with open(tmp_path, 'w', encoding='utf-8') as f:
			json.dump(data, f, ensure_ascii=False, indent=2)
		os.replace(tmp_path, lotes_path)
		return True
	except Exception:
		return False

def _load_unidades_data():
	base_dir = os.path.dirname(os.path.dirname(__file__))
	unidades_path = os.path.join(base_dir, 'dados', 'unidades.json')
	if not os.path.isfile(unidades_path):
		return None
	try:
		with open(unidades_path, 'r', encoding='utf-8') as f:
			return json.load(f)
	except Exception:
		return None

def _save_unidades_data(data):
	base_dir = os.path.dirname(os.path.dirname(__file__))
	unidades_path = os.path.join(base_dir, 'dados', 'unidades.json')
	try:
		os.makedirs(os.path.dirname(unidades_path), exist_ok=True)
		tmp_path = unidades_path + '.tmp'
		with open(tmp_path, 'w', encoding='utf-8') as f:
			json.dump(data, f, ensure_ascii=False, indent=2)
		os.replace(tmp_path, unidades_path)
		return True
	except Exception:
		return False

def salvar_novo_lote(payload: dict):
	if not isinstance(payload, dict):
		return {'success': False, 'error': 'Payload inválido'}

	nome = payload.get('nome_lote') or payload.get('nome') or payload.get('nomeLote') or ''
	empresa = payload.get('nome_empresa') or payload.get('empresa') or payload.get('empresa_nome') or ''
	numero_contrato = payload.get('numero_contrato') or payload.get('contrato') or ''
	data_inicio = payload.get('data_inicio') or payload.get('inicio_contrato') or ''
	unidades = payload.get('unidades') or payload.get('unidades[]') or []

	if unidades is None:
		unidades = []
	if isinstance(unidades, str):
		unidades = [u.strip() for u in unidades.split(',') if u.strip()]
	if not isinstance(unidades, list):
		unidades = list(unidades)

	precos = {}
	raw_precos = payload.get('precos') or {}
	if isinstance(raw_precos, dict):
		precos['cafe_interno'] = raw_precos.get('cafe', {}).get('interno') if isinstance(raw_precos.get('cafe'), dict) else raw_precos.get('cafe_interno') or raw_precos.get('cafeInterno')
		precos['cafe_funcionario'] = raw_precos.get('cafe', {}).get('funcionario') if isinstance(raw_precos.get('cafe'), dict) else raw_precos.get('cafe_funcionario') or raw_precos.get('cafeFuncionario')
		precos['almoco_interno'] = raw_precos.get('almoco', {}).get('interno') if isinstance(raw_precos.get('almoco'), dict) else raw_precos.get('almoco_interno') or raw_precos.get('almocoInterno')
		precos['almoco_funcionario'] = raw_precos.get('almoco', {}).get('funcionario') if isinstance(raw_precos.get('almoco'), dict) else raw_precos.get('almoco_funcionario') or raw_precos.get('almocoFuncionario')
		precos['lanche_interno'] = raw_precos.get('lanche', {}).get('interno') if isinstance(raw_precos.get('lanche'), dict) else raw_precos.get('lanche_interno') or raw_precos.get('lancheInterno')
		precos['lanche_funcionario'] = raw_precos.get('lanche', {}).get('funcionario') if isinstance(raw_precos.get('lanche'), dict) else raw_precos.get('lanche_funcionario') or raw_precos.get('lancheFuncionario')
		precos['jantar_interno'] = raw_precos.get('jantar', {}).get('interno') if isinstance(raw_precos.get('jantar'), dict) else raw_precos.get('jantar_interno') or raw_precos.get('jantarInterno')
		precos['jantar_funcionario'] = raw_precos.get('jantar', {}).get('funcionario') if isinstance(raw_precos.get('jantar'), dict) else raw_precos.get('jantar_funcionario') or raw_precos.get('jantarFuncionario')
	else:
		for k in ['cafe_interno','cafe_funcionario','almoco_interno','almoco_funcionario','lanche_interno','lanche_funcionario','jantar_interno','jantar_funcionario']:
			precos[k] = payload.get(k)

	if not nome or not empresa or not numero_contrato or not data_inicio:
		return {'success': False, 'error': 'Campos obrigatórios faltando'}
	if not unidades or not isinstance(unidades, list) or len(unidades) == 0:
		return {'success': False, 'error': 'Adicione pelo menos uma unidade'}

	data = _load_lotes_data()
	lotes = None
	wrapped = None
	if isinstance(data, list):
		lotes = data
	elif isinstance(data, dict) and isinstance(data.get('lotes'), list):
		lotes = data.get('lotes')
		wrapped = data
	else:
		lotes = []

	existing_ids = [l.get('id') for l in lotes if isinstance(l, dict) and isinstance(l.get('id'), int)]
	new_id = (max(existing_ids) + 1) if existing_ids else 0

	input_unidades = unidades
	unit_ids = []
	created_unit_ids = []

	units_data = _load_unidades_data()
	units_list = []
	units_wrapped = None
	if isinstance(units_data, list):
		units_list = units_data
	elif isinstance(units_data, dict) and isinstance(units_data.get('unidades'), list):
		units_list = units_data.get('unidades')
		units_wrapped = units_data
	else:
		units_list = []

	existing_unit_ids = [u.get('id') for u in units_list if isinstance(u, dict) and isinstance(u.get('id'), int)]
	next_unit_id = (max(existing_unit_ids) + 1) if existing_unit_ids else 0

	def _is_int_like(x):
		try:
			int(x)
			return True
		except Exception:
			return False

	if isinstance(input_unidades, list) and input_unidades and all(_is_int_like(u) for u in input_unidades):
		unit_ids = [int(u) for u in input_unidades]
	else:
		for raw in (input_unidades or []):
			name = str(raw).strip()
			if not name:
				continue
			found = None
			for u in units_list:
				if not isinstance(u, dict):
					continue
				if isinstance(u.get('nome'), str) and u.get('nome').strip().lower() == name.lower():
					found = u
					break
			if found:
				found['lote_id'] = new_id
				unit_ids.append(found.get('id'))
			else:
				uid = next_unit_id
				new_unit = {
					'id': uid,
					'nome': name,
					'lote_id': new_id,
					'criado_em': datetime.now().isoformat()
				}
				units_list.append(new_unit)
				unit_ids.append(uid)
				created_unit_ids.append(uid)
				next_unit_id += 1

	if units_wrapped is not None:
		units_wrapped['unidades'] = units_list
		to_write_units = units_wrapped
	else:
		to_write_units = units_list

	if not _save_unidades_data(to_write_units):
		return {'success': False, 'error': 'Erro ao salvar unidades'}

	lote_record = {
		'id': new_id,
		'nome': str(nome),
		'empresa': str(empresa),
		'numero_contrato': str(numero_contrato),
		'data_inicio': str(data_inicio),
		'unidades': unit_ids,
		'precos': precos,
		'ativo': True,
		'criado_em': datetime.now().isoformat()
	}

	lotes.append(lote_record)
	if wrapped is not None:
		wrapped['lotes'] = lotes
		to_write = wrapped
	else:
		to_write = lotes

	ok = _save_lotes_data(to_write)
	if not ok:
		if created_unit_ids:
			try:
				ud = _load_unidades_data() or []
				if isinstance(ud, dict) and isinstance(ud.get('unidades'), list):
					lst = ud.get('unidades')
					lst = [u for u in lst if not (isinstance(u, dict) and u.get('id') in created_unit_ids)]
					ud['unidades'] = lst
				else:
					lst = [u for u in (ud if isinstance(ud, list) else []) if not (isinstance(u, dict) and u.get('id') in created_unit_ids)]
					ud = lst
				_save_unidades_data(ud)
			except Exception:
				pass
		return {'success': False, 'error': 'Erro ao salvar lote'}
	return {'success': True, 'id': new_id}


# ----- Preços normalization helper (exported) -----
def normalizar_precos(raw_precos):
	"""Normaliza diferentes formatos de `precos` para o formato nested esperado

	Entrada aceita:
	  - dict nested: {'cafe': {'interno': '1.2', 'funcionario': '0.8'}, ...}
	  - dict plano: {'cafe_interno': '1.2', 'cafe_funcionario': '0.8', ...}
	  - string JSON ou string com pares tipo "cafe_interno:1.2,cafe_funcionario=0.8"

	Retorna:
	  {'cafe': {'interno': float, 'funcionario': float}, ...}
	"""
	meals = ('cafe', 'almoco', 'lanche', 'jantar')

	def _to_float(v):
		try:
			return float(str(v).replace(',', '.'))
		except Exception:
			return 0.0

	res = {m: {'interno': 0.0, 'funcionario': 0.0} for m in meals}
	if raw_precos is None:
		return res

	# string -> tentar decodificar JSON ou extrair pares
	if isinstance(raw_precos, str):
		txt = raw_precos.strip()
		try:
			parsed = json.loads(txt)
		except Exception:
			try:
				parsed = json.loads(txt.replace("'", '"'))
			except Exception:
				parsed = {}
				for m in re.finditer(r"([a-zA-Z0-9_]+)\s*[:=]\s*['\"]?([0-9\.,]+)['\"]?", txt):
					k = m.group(1)
					v = m.group(2)
					parsed[k] = v
		raw = parsed
	elif isinstance(raw_precos, dict):
		raw = raw_precos
	else:
		return res

	if isinstance(raw, dict):
		for meal in meals:
			val = raw.get(meal)
			if isinstance(val, dict):
				res[meal]['interno'] = _to_float(val.get('interno') or val.get('interno_val') or 0)
				res[meal]['funcionario'] = _to_float(val.get('funcionario') or val.get('funcionario_val') or 0)
			else:
				int_key = f"{meal}_interno"
				func_key = f"{meal}_funcionario"
				if int_key in raw or func_key in raw:
					res[meal]['interno'] = _to_float(raw.get(int_key) or raw.get(int_key.replace('_', '')))
					res[meal]['funcionario'] = _to_float(raw.get(func_key) or raw.get(func_key.replace('_', '')))
				int_key2 = f"{meal}Interno"
				func_key2 = f"{meal}Funcionario"
				if (res[meal]['interno'] == 0.0) and int_key2 in raw:
					res[meal]['interno'] = _to_float(raw.get(int_key2))
				if (res[meal]['funcionario'] == 0.0) and func_key2 in raw:
					res[meal]['funcionario'] = _to_float(raw.get(func_key2))
		for m in meals:
			res[m]['interno'] = _to_float(res[m]['interno'])
			res[m]['funcionario'] = _to_float(res[m]['funcionario'])
		return res
	return res


# ----- Parser tabular para campos de texto -----
def parse_texto_tabular(texto):
	"""Analisa texto tabular (separado por tab ou por espaços) e retorna
	um dicionário com listas numéricas por coluna.

	Retorno exemplo quando bem-sucedido:
	  {'ok': True, 'colunas': {0: [1,2,3], 1: [128,127,...]}, 'linhas': N, 'colunas_count': M}

	Se falhar, retorna {'ok': False, 'error': '...'}.
	"""
	if texto is None:
		return {'ok': False, 'error': 'Texto vazio'}
	if not isinstance(texto, str):
		try:
			texto = str(texto)
		except Exception:
			return {'ok': False, 'error': 'Texto não serializável'}

	lines = [ln.strip() for ln in texto.splitlines() if ln.strip()]
	if not lines:
		return {'ok': True, 'colunas': {}, 'linhas': 0, 'colunas_count': 0}

	# determinar delimitador: preferir tab, senão espaço(s)
	delimiter = '\t' if any('\t' in ln for ln in lines) else None

	rows = []
	for ln in lines:
		if delimiter:
			parts = [p.strip() for p in ln.split('\t')]
		else:
			# separar por espaços múltiplos
			parts = [p.strip() for p in re.split(r"\s+", ln) if p.strip()]
		rows.append(parts)

	# número máximo de colunas
	max_cols = max(len(r) for r in rows)

	# inicializar colunas — usar chaves legíveis: 'coluna_0', 'coluna_1', ...
	cols = {f'coluna_{i}': [] for i in range(max_cols)}

	def _to_number(token):
		if token is None:
			return None
		t = str(token).strip()
		if t == '':
			return None
		# substituir vírgula decimal
		t2 = t.replace(',', '.')
		# permitir sinais e pontos
		m = re.match(r'^[-+]?\d+(?:\.\d+)?$', t2)
		if m:
			# inteiro ou float
			if '.' in t2:
				try:
					return float(t2)
				except Exception:
					return None
			else:
				try:
					return int(t2)
				except Exception:
					try:
						return float(t2)
					except Exception:
						return None
		# tentar extrair primeiro número presente
		m2 = re.search(r'[-+]?\d+[\.,]?\d*', t)
		if m2:
			s = m2.group(0).replace(',', '.')
			try:
				if '.' in s:
					return float(s)
				return int(s)
			except Exception:
				try:
					return float(s)
				except Exception:
					return None
		return None

	for r in rows:
		for idx in range(max_cols):
			token = r[idx] if idx < len(r) else ''
			num = _to_number(token)
			cols[f'coluna_{idx}'].append(num)

	return {'ok': True, 'colunas': cols, 'linhas': len(rows), 'colunas_count': max_cols}


def _normalizar_datas_coluna(col0_values, entry):
	"""Normaliza valores da coluna_0 para uma lista de strings DD/MM/YYYY baseada em mes/ano do registro.

	Aceita tokens numéricos (1, 2), strings como 'dia 1', '01/10' etc. Se não for possível
	normalizar um token, coloca None na posição correspondente.
	"""
	if not isinstance(col0_values, list):
		return None

	# extrair mês e ano do entry com várias chaves possíveis
	mes_keys = ('mes', 'month', 'mes_num', 'mesNumero', 'month_num')
	ano_keys = ('ano', 'year')
	mes = None
	ano = None
	for k in mes_keys:
		if k in entry and entry.get(k) is not None:
			try:
				mes = int(entry.get(k))
				break
			except Exception:
				try:
					mes = int(str(entry.get(k)).strip())
					break
				except Exception:
					mes = None
	for k in ano_keys:
		if k in entry and entry.get(k) is not None:
			try:
				ano = int(entry.get(k))
				break
			except Exception:
				try:
					ano = int(str(entry.get(k)).strip())
					break
				except Exception:
					ano = None

	# fallback para ano/mês atual quando não fornecidos
	now = datetime.now()
	if mes is None:
		mes = now.month
	if ano is None:
		ano = now.year

	# número de dias no mês para validação
	try:
		days_in_month = calendar.monthrange(ano, mes)[1]
	except Exception:
		days_in_month = 31

	out = []
	for v in col0_values:
		if v is None:
			out.append(None)
			continue
		# aceitar inteiros já convertidos pelo parser
		if isinstance(v, (int,)):
			day = int(v)
		else:
			s = str(v).strip()
			if not s:
				out.append(None)
				continue
			# formatos com barra: '01/10' -> extrair primeira parte como dia
			if '/' in s or '-' in s:
				parts = re.split(r'[\/\-]', s)
				# procurar o primeiro numeric part
				day = None
				for p in parts:
					m = re.search(r'(\d{1,2})', p)
					if m:
						try:
							day = int(m.group(1))
							break
						except Exception:
							day = None
				if day is None:
					# fallback: look for any number
					nm = re.search(r'(\d{1,2})', s)
					day = int(nm.group(1)) if nm else None
			else:
				# 'dia 1' or '1' or '01' etc. buscar primeiro número de 1-2 dígitos
				m = re.search(r'(\d{1,2})', s)
				if m:
					try:
						day = int(m.group(1))
					except Exception:
						day = None
				else:
					day = None

		# validar dia
		try:
			if day is None or day < 1 or day > days_in_month:
				out.append(None)
			else:
				dt = datetime(year=ano, month=mes, day=day)
				out.append(dt.strftime('%d/%m/%Y'))
		except Exception:
			out.append(None)

	return out


def _get_days_in_month_from_entry(entry):
	"""Retorna número de dias do mês/ano declarado no registro, ou None se não puder extrair."""
	mes_keys = ('mes', 'month', 'mes_num', 'mesNumero', 'month_num')
	ano_keys = ('ano', 'year')
	mes = None
	ano = None
	for k in mes_keys:
		if k in entry and entry.get(k) is not None:
			try:
				mes = int(entry.get(k))
				break
			except Exception:
				try:
					mes = int(str(entry.get(k)).strip())
					break
				except Exception:
					mes = None
	for k in ano_keys:
		if k in entry and entry.get(k) is not None:
			try:
				ano = int(entry.get(k))
				break
			except Exception:
				try:
					ano = int(str(entry.get(k)).strip())
					break
				except Exception:
					ano = None
	if mes is None or ano is None:
		return None
	try:
		return calendar.monthrange(ano, mes)[1]
	except Exception:
		return None


def _validate_map_day_lengths(entry):
	"""Valida que todas as listas diárias presentes no registro tenham comprimento igual ao número de dias do mês/ano.

	Retorna (True, None) quando ok, ou (False, mensagem) quando inválido.
	"""
	days = _get_days_in_month_from_entry(entry)
	if days is None:
		return (False, 'Mês ou ano inválido ou ausente no registro')

	expected = int(days)
	# campos que representam séries diárias — se presentes, devem ter length == expected
	daily_fields = [
		'dados_siisp',
		'cafe_interno', 'cafe_funcionario',
		'almoco_interno', 'almoco_funcionario',
		'lanche_interno', 'lanche_funcionario',
		'jantar_interno', 'jantar_funcionario',
		'datas'
	]
	errors = []
	for f in daily_fields:
		if f in entry:
			v = entry.get(f)
			if not isinstance(v, list):
				errors.append(f"{f} não é uma lista")
			else:
				# dados_siisp é opcional: pode ser lista vazia (nenhum dado) ou ter exatamente 'expected' elementos
				if f == 'dados_siisp':
					if len(v) not in (0, expected):
						errors.append(f"{f} tem {len(v)} elementos; esperado 0 ou {expected}")
				else:
					if len(v) != expected:
						errors.append(f"{f} tem {len(v)} elementos; esperado {expected}")
	if errors:
		return (False, '; '.join(errors))
	return (True, None)


# ----- Dashboard loader (reusable) -----
def calcular_metricas_lotes(lotes, mapas):
	"""Calcula métricas agregadas para cada lote baseado nos mapas associados.
	
	Modifica os lotes in-place adicionando os campos:
	- meses_cadastrados: número de meses únicos com dados
	- refeicoes_mes: média de refeições por mês
	- custo_mes: custo médio por mês
	- desvio_mes: desvio médio por mês (diferença entre refeições e SIISP)
	
	Args:
		lotes: lista de dicionários representando os lotes
		mapas: lista de dicionários representando os mapas de refeições
	"""
	from collections import defaultdict
	
	# calcular meses cadastrados por lote: conjunto único de (mes, ano)
	meses_por_lote = defaultdict(set)
	totais_refeicoes_por_lote = {}
	totais_custos_por_lote = {}
	totais_desvios_por_lote = {}
	
	for m in (mapas or []):
		try:
			lote_id = int(m.get('lote_id'))
		except Exception:
			continue
		
		mes = m.get('mes') or m.get('month') or m.get('mes_num') or m.get('month_num')
		ano = m.get('ano') or m.get('year')
		
		# tentar extrair mês/ano a partir de datas quando faltarem
		if (mes is None or ano is None) and isinstance(m.get('datas'), list) and len(m.get('datas')) > 0:
			try:
				# formato esperado DD/MM/YYYY
				parts = str(m.get('datas')[0]).split('/')
				if len(parts) >= 3:
					mes = int(parts[1])
					ano = int(parts[2])
			except Exception:
				pass
		
		try:
			mes_i = int(mes)
			ano_i = int(ano)
		except Exception:
			# não foi possível extrair mês/ano válidos
			continue
		
		meses_por_lote[lote_id].add((mes_i, ano_i))
		
		# acumular refeições totais por lote (usar campo pré-calculado do mapa quando disponível)
		try:
			total = int(m.get('refeicoes_mes') or 0)
		except Exception:
			try:
				total = int(float(m.get('refeicoes_mes') or 0))
			except Exception:
				total = 0
		
		totais_refeicoes_por_lote[lote_id] = totais_refeicoes_por_lote.get(lote_id, 0) + total
		
		# acumular custos totais por lote
		custo_mapa = 0.0
		# calcular custo do mapa atual multiplicando quantidades por preços
		# buscar o lote correspondente para obter os preços
		lote_do_mapa = None
		for l_temp in lotes:
			try:
				if int(l_temp.get('id')) == lote_id:
					lote_do_mapa = l_temp
					break
			except Exception:
				continue
		
		if lote_do_mapa and isinstance(lote_do_mapa.get('precos'), dict):
			precos = lote_do_mapa.get('precos', {})
			# somar todos os tipos de refeição (cafe, almoco, lanche, jantar) x (interno, funcionario)
			meal_fields = [
				('cafe_interno', precos.get('cafe', {}).get('interno', 0)),
				('cafe_funcionario', precos.get('cafe', {}).get('funcionario', 0)),
				('almoco_interno', precos.get('almoco', {}).get('interno', 0)),
				('almoco_funcionario', precos.get('almoco', {}).get('funcionario', 0)),
				('lanche_interno', precos.get('lanche', {}).get('interno', 0)),
				('lanche_funcionario', precos.get('lanche', {}).get('funcionario', 0)),
				('jantar_interno', precos.get('jantar', {}).get('interno', 0)),
				('jantar_funcionario', precos.get('jantar', {}).get('funcionario', 0))
			]
			
			for field_name, preco_unitario in meal_fields:
				if field_name in m:
					try:
						quantidade = sum(int(x or 0) for x in m[field_name] if x is not None)
						custo_mapa += quantidade * float(preco_unitario or 0)
					except Exception:
						pass
		
		totais_custos_por_lote[lote_id] = totais_custos_por_lote.get(lote_id, 0.0) + custo_mapa
		
		# acumular desvios totais por lote (baseado nos campos _siisp)
		# Desvio = valor monetário dos EXCEDENTES (apenas valores positivos)
		desvio_mapa = 0.0
		desvio_fields = [
			('cafe_interno_siisp', precos.get('cafe', {}).get('interno', 0)),
			('cafe_funcionario_siisp', precos.get('cafe', {}).get('funcionario', 0)),
			('almoco_interno_siisp', precos.get('almoco', {}).get('interno', 0)),
			('almoco_funcionario_siisp', precos.get('almoco', {}).get('funcionario', 0)),
			('lanche_interno_siisp', precos.get('lanche', {}).get('interno', 0)),
			('lanche_funcionario_siisp', precos.get('lanche', {}).get('funcionario', 0)),
			('jantar_interno_siisp', precos.get('jantar', {}).get('interno', 0)),
			('jantar_funcionario_siisp', precos.get('jantar', {}).get('funcionario', 0))
		]
		
		for field_name, preco_unitario in desvio_fields:
			if field_name in m:
				try:
					# somar apenas os excedentes (valores positivos > 0)
					for valor_dia in m[field_name]:
						if valor_dia is not None:
							val = int(valor_dia)
							if val > 0:  # apenas excedentes
								desvio_mapa += val * float(preco_unitario or 0)
				except Exception:
					pass
		
		totais_desvios_por_lote[lote_id] = totais_desvios_por_lote.get(lote_id, 0.0) + desvio_mapa
	
	# anexar métricas calculadas a cada lote
	for l in lotes:
		try:
			lid = int(l.get('id'))
		except Exception:
			lid = None
		
		count = len(meses_por_lote.get(lid, set())) if lid is not None else 0
		l['meses_cadastrados'] = count
		
		# calcular média mensal de refeições (total refeicoes / meses_cadastrados)
		total_ref = totais_refeicoes_por_lote.get(lid, 0) if lid is not None else 0
		avg = 0.0
		if count > 0:
			try:
				avg = round(float(total_ref) / float(count), 2)
			except Exception:
				avg = 0.0
		l['refeicoes_mes'] = avg
		
		# calcular custo médio mensal (total custos / meses_cadastrados)
		total_custo = totais_custos_por_lote.get(lid, 0.0) if lid is not None else 0.0
		avg_custo = 0.0
		if count > 0:
			try:
				avg_custo = round(total_custo / float(count), 2)
			except Exception:
				avg_custo = 0.0
		l['custo_mes'] = avg_custo
		
		# calcular desvio médio mensal (total desvios / meses_cadastrados)
		total_desvio = totais_desvios_por_lote.get(lid, 0.0) if lid is not None else 0.0
		avg_desvio = 0.0
		if count > 0:
			try:
				avg_desvio = round(total_desvio / float(count), 2)
			except Exception:
				avg_desvio = 0.0
		l['desvio_mes'] = avg_desvio


def carregar_lotes_para_dashboard():
	"""Carrega e normaliza os lotes e unidades para uso no dashboard.

	Retorna um dicionário: { 'lotes': [...], 'mapas_dados': [...] }
	onde cada lote é um dict compatível com o que o template espera
	(campos: id, nome, empresa, contrato, data_inicio, ativo, unidades, precos, ...).
	"""
	lotes_raw = _load_lotes_data() or []
	unidades_raw = _load_unidades_data() or []

	# normalizar lista de unidades como lista de objetos
	unidades_list = []
	if isinstance(unidades_raw, dict) and isinstance(unidades_raw.get('unidades'), list):
		unidades_list = unidades_raw.get('unidades')
	elif isinstance(unidades_raw, list):
		unidades_list = unidades_raw

	# construir mapa id -> nome
	unidades_map = {}
	for u in unidades_list:
		if isinstance(u, dict) and isinstance(u.get('id'), int):
			unidades_map[int(u.get('id'))] = u.get('nome')

	lotes = []
	if isinstance(lotes_raw, dict) and isinstance(lotes_raw.get('lotes'), list):
		src_lotes = lotes_raw.get('lotes')
	elif isinstance(lotes_raw, list):
		src_lotes = lotes_raw
	else:
		src_lotes = []

	for l in src_lotes:
		if not isinstance(l, dict):
			continue
		raw_unidades = l.get('unidades') or []
		unidades_final = []
		if isinstance(raw_unidades, list) and raw_unidades:
			# detect numeric ids
			if all(isinstance(x, int) or (isinstance(x, str) and x.isdigit()) for x in raw_unidades):
				for x in raw_unidades:
					try:
						uid = int(x)
						unidades_final.append(unidades_map.get(uid, str(uid)))
					except Exception:
						unidades_final.append(str(x))
			else:
				unidades_final = [str(x) for x in raw_unidades if x]

		# usar função pública de normalização de preços
		# (normalizar_precos já lida com strings, dicts planos e nested)
		# removemos a implementação local e delegamos à função exportada

		# garantir campos numéricos usados pelo template com valores padrão
		try:
			refeicoes_mes = int(l.get('refeicoes_mes') if l.get('refeicoes_mes') is not None else (l.get('refeicoes') or 0))
		except Exception:
			try:
				refeicoes_mes = int(float(l.get('refeicoes') or 0))
			except Exception:
				refeicoes_mes = 0
		try:
			custo_mes = float(l.get('custo_mes') if l.get('custo_mes') is not None else l.get('custo') or 0.0)
		except Exception:
			try:
				custo_mes = float(str(l.get('custo') or 0).replace(',', '.'))
			except Exception:
				custo_mes = 0.0
		try:
			desvio_mes = float(l.get('desvio_mes') if l.get('desvio_mes') is not None else l.get('desvio') or 0.0)
		except Exception:
			try:
				desvio_mes = float(str(l.get('desvio') or 0).replace(',', '.'))
			except Exception:
				desvio_mes = 0.0
		try:
			meses_cadastrados = int(l.get('meses_cadastrados') if l.get('meses_cadastrados') is not None else l.get('meses') or 0)
		except Exception:
			meses_cadastrados = 0

		# padronizar conformidade: garantir float e default 0.0
		try:
			conformidade_val = l.get('conformidade')
			if conformidade_val is None:
				conformidade = 0.0
			else:
				# aceitar strings com vírgula ou ponto, ints etc.
				conformidade = float(str(conformidade_val).replace(',', '.'))
		except Exception:
			conformidade = 0.0

		lote_obj = {
			'id': l.get('id'),
			'nome': l.get('nome') or l.get('nome_lote') or '',
			'empresa': l.get('empresa') or '',
			'contrato': l.get('numero_contrato') or l.get('contrato') or '',
			'data_inicio': l.get('data_inicio'),
			'ativo': l.get('ativo', True),
			'unidades': unidades_final,
			'precos': normalizar_precos(l.get('precos')),
			'refeicoes_mes': refeicoes_mes,
			'custo_mes': custo_mes,
			'desvio_mes': desvio_mes,
			'meses_cadastrados': meses_cadastrados,
			'refeicoes': l.get('refeicoes'),
			'conformidade': conformidade,
			'alertas': l.get('alertas')
		}
		lotes.append(lote_obj)

	mapas_dados = []

	# Carregar mapas salvos - NOVO: priorizar arquivos particionados
	mapas_list_src = []
	
	# Primeiro: tentar carregar arquivos particionados
	mapas_particionados = _load_all_mapas_partitioned()
	if mapas_particionados:
		mapas_list_src = mapas_particionados
		print(f"✅ Usando {len(mapas_list_src)} mapas de arquivos particionados")
	else:
		# Fallback: carregar arquivo legado mapas.json
		print("📂 Nenhum arquivo particionado encontrado, tentando mapas.json legado...")
		mapas_raw = _load_mapas_data() or []
		if isinstance(mapas_raw, dict) and isinstance(mapas_raw.get('mapas'), list):
			mapas_list_src = mapas_raw.get('mapas')
		elif isinstance(mapas_raw, list):
			mapas_list_src = mapas_raw
		else:
			mapas_list_src = []
		print(f"📂 Carregados {len(mapas_list_src)} mapas do arquivo legado")

	for m in mapas_list_src:
		if not isinstance(m, dict):
			continue
		# normalizar campos básicos
		try:
			lote_id = int(m.get('lote_id') if m.get('lote_id') is not None else m.get('lote') or m.get('loteId'))
		except Exception:
			try:
				lote_id = int(str(m.get('lote_id') or m.get('lote') or m.get('loteId')).strip())
			except Exception:
				lote_id = None

		mes_val = m.get('mes') or m.get('month') or m.get('mes_num')
		ano_val = m.get('ano') or m.get('year')
		try:
			mes = int(mes_val)
		except Exception:
			try:
				mes = int(str(mes_val).strip())
			except Exception:
				mes = None
		try:
			ano = int(ano_val)
		except Exception:
			try:
				ano = int(str(ano_val).strip())
			except Exception:
				ano = None

		unidade_raw = m.get('unidade') or m.get('unidade_nome') or m.get('unidadeNome') or ''
		# if unidade looks like an id, map to name
		nome_unidade = None
		try:
			if isinstance(unidade_raw, int):
				nome_unidade = unidades_map.get(int(unidade_raw))
			else:
				ustr = str(unidade_raw).strip()
				if ustr.isdigit():
					uid = int(ustr)
					nome_unidade = unidades_map.get(uid) or ustr
				else:
					nome_unidade = ustr
		except Exception:
			nome_unidade = str(unidade_raw)

		# datas
		datas = m.get('datas') if isinstance(m.get('datas'), list) else []

		# helper to coerce list fields to lists of numbers
		def _coerce_list(name):
			v = m.get(name)
			if isinstance(v, list):
				return v
			return []

		cafe_interno = _coerce_list('cafe_interno')
		cafe_funcionario = _coerce_list('cafe_funcionario')
		almoco_interno = _coerce_list('almoco_interno')
		almoco_funcionario = _coerce_list('almoco_funcionario')
		lanche_interno = _coerce_list('lanche_interno')
		lanche_funcionario = _coerce_list('lanche_funcionario')
		jantar_interno = _coerce_list('jantar_interno')
		jantar_funcionario = _coerce_list('jantar_funcionario')
		dados_siisp = _coerce_list('dados_siisp')

		# calcular total de refeições no mês (somando interno+funcionario de cada refeição por dia)
		total_refeicoes = 0
		# determinar número de dias como comprimento máximo das listas de datas ou das listas de refeições
		n_days = 0
		if isinstance(datas, list) and len(datas) > 0:
			n_days = len(datas)
		else:
			n_days = max(len(cafe_interno), len(cafe_funcionario), len(almoco_interno), len(almoco_funcionario), len(lanche_interno), len(lanche_funcionario), len(jantar_interno), len(jantar_funcionario))

		for i in range(n_days):
			vals = 0
			for arr in (cafe_interno, cafe_funcionario, almoco_interno, almoco_funcionario, lanche_interno, lanche_funcionario, jantar_interno, jantar_funcionario):
				try:
					v = arr[i] if i < len(arr) and (arr[i] is not None) else 0
					vals += int(v)
				except Exception:
					try:
						vals += int(float(arr[i]))
					except Exception:
						pass
			total_refeicoes += vals

		# Processar dados SIISP e calcular diferenças
		n_siisp = _coerce_list('n_siisp')  # Array de números SIISP por dia
		
		# Se não há n_siisp mas há dados_siisp, tentar extrair números SIISP de dados_siisp
		if not n_siisp and dados_siisp:
			n_siisp = dados_siisp  # dados_siisp pode conter os números SIISP
		
		# Calcular diferenças SIISP para internos (refeições - n_siisp)
		# IMPORTANTE: Verificar se os campos já existem no JSON antes de recalcular
		cafe_interno_siisp = _coerce_list('cafe_interno_siisp') if 'cafe_interno_siisp' in m else []
		almoco_interno_siisp = _coerce_list('almoco_interno_siisp') if 'almoco_interno_siisp' in m else []
		lanche_interno_siisp = _coerce_list('lanche_interno_siisp') if 'lanche_interno_siisp' in m else []
		jantar_interno_siisp = _coerce_list('jantar_interno_siisp') if 'jantar_interno_siisp' in m else []
		
		# Calcular diferenças SIISP para funcionários
		cafe_funcionario_siisp = _coerce_list('cafe_funcionario_siisp') if 'cafe_funcionario_siisp' in m else []
		almoco_funcionario_siisp = _coerce_list('almoco_funcionario_siisp') if 'almoco_funcionario_siisp' in m else []
		lanche_funcionario_siisp = _coerce_list('lanche_funcionario_siisp') if 'lanche_funcionario_siisp' in m else []
		jantar_funcionario_siisp = _coerce_list('jantar_funcionario_siisp') if 'jantar_funcionario_siisp' in m else []
		
		# Só recalcular se não existirem no JSON
		if n_siisp and not cafe_interno_siisp:
			for i in range(max(len(n_siisp), n_days)):
				# Obter valores para o dia i
				siisp_dia = n_siisp[i] if i < len(n_siisp) and n_siisp[i] is not None else 0
				
				# Diferenças para internos (positivo = mais refeições que internos SIISP)
				cafe_int_dia = cafe_interno[i] if i < len(cafe_interno) and cafe_interno[i] is not None else 0
				almoco_int_dia = almoco_interno[i] if i < len(almoco_interno) and almoco_interno[i] is not None else 0
				lanche_int_dia = lanche_interno[i] if i < len(lanche_interno) and lanche_interno[i] is not None else 0
				jantar_int_dia = jantar_interno[i] if i < len(jantar_interno) and jantar_interno[i] is not None else 0
				
				try:
					cafe_interno_siisp.append(int(cafe_int_dia) - int(siisp_dia))
					almoco_interno_siisp.append(int(almoco_int_dia) - int(siisp_dia))
					lanche_interno_siisp.append(int(lanche_int_dia) - int(siisp_dia))
					jantar_interno_siisp.append(int(jantar_int_dia) - int(siisp_dia))
				except Exception:
					cafe_interno_siisp.append(0)
					almoco_interno_siisp.append(0)
					lanche_interno_siisp.append(0)
					jantar_interno_siisp.append(0)
				
				# Para funcionários, diferença é simplesmente o número de funcionários (SIISP = 0 para funcionários)
				cafe_func_dia = cafe_funcionario[i] if i < len(cafe_funcionario) and cafe_funcionario[i] is not None else 0
				almoco_func_dia = almoco_funcionario[i] if i < len(almoco_funcionario) and almoco_funcionario[i] is not None else 0
				lanche_func_dia = lanche_funcionario[i] if i < len(lanche_funcionario) and lanche_funcionario[i] is not None else 0
				jantar_func_dia = jantar_funcionario[i] if i < len(jantar_funcionario) and jantar_funcionario[i] is not None else 0
				
				try:
					cafe_funcionario_siisp.append(int(cafe_func_dia))
					almoco_funcionario_siisp.append(int(almoco_func_dia))
					lanche_funcionario_siisp.append(int(lanche_func_dia))
					jantar_funcionario_siisp.append(int(jantar_func_dia))
				except Exception:
					cafe_funcionario_siisp.append(0)
					almoco_funcionario_siisp.append(0)
					lanche_funcionario_siisp.append(0)
					jantar_funcionario_siisp.append(0)

		mapa_obj = {
			'id': m.get('id'),
			'lote_id': lote_id,
			'nome_unidade': nome_unidade,
			'mes': mes,
			'ano': ano,
			'data': datas,
			'linhas': int(m.get('linhas') or 0),
			'colunas_count': int(m.get('colunas_count') or 0),
			'cafe_interno': cafe_interno,
			'cafe_funcionario': cafe_funcionario,
			'almoco_interno': almoco_interno,
			'almoco_funcionario': almoco_funcionario,
			'lanche_interno': lanche_interno,
			'lanche_funcionario': lanche_funcionario,
			'jantar_interno': jantar_interno,
			'jantar_funcionario': jantar_funcionario,
			'dados_siisp': dados_siisp,
			'n_siisp': n_siisp,  # Array de números SIISP
			'cafe_interno_siisp': cafe_interno_siisp,
			'almoco_interno_siisp': almoco_interno_siisp,
			'lanche_interno_siisp': lanche_interno_siisp,
			'jantar_interno_siisp': jantar_interno_siisp,
			'cafe_funcionario_siisp': cafe_funcionario_siisp,
			'almoco_funcionario_siisp': almoco_funcionario_siisp,
			'lanche_funcionario_siisp': lanche_funcionario_siisp,
			'jantar_funcionario_siisp': jantar_funcionario_siisp,
			'refeicoes_mes': total_refeicoes,
			'criado_em': m.get('criado_em'),
			'atualizado_em': m.get('atualizado_em')
		}
		mapas_dados.append(mapa_obj)

	return {'lotes': lotes, 'mapas_dados': mapas_dados}


def _load_mapas_data():
	base_dir = os.path.dirname(os.path.dirname(__file__))
	mapas_path = os.path.join(base_dir, 'dados', 'mapas.json')
	print(f"🔍 DEBUG: Carregando mapas de: {mapas_path}")
	if not os.path.isfile(mapas_path):
		print("❌ Arquivo mapas.json não encontrado!")
		return None
	try:
		with open(mapas_path, 'r', encoding='utf-8') as f:
			data = json.load(f)
			# Debug: mostrar primeiro valor de cafe_funcionario_siisp
			if isinstance(data, list) and len(data) > 0:
				primeiro_mapa = data[0]
				if 'cafe_funcionario_siisp' in primeiro_mapa:
					print(f"✅ Primeiro valor cafe_funcionario_siisp no arquivo: {primeiro_mapa['cafe_funcionario_siisp'][0] if primeiro_mapa['cafe_funcionario_siisp'] else 'VAZIO'}")
				else:
					print("⚠️ Campo cafe_funcionario_siisp NÃO EXISTE no JSON!")
			return data
	except Exception as e:
		print(f"❌ Erro ao ler mapas.json: {e}")
		return None


def _save_mapas_data(data):
	base_dir = os.path.dirname(os.path.dirname(__file__))
	mapas_path = os.path.join(base_dir, 'dados', 'mapas.json')
	try:
		os.makedirs(os.path.dirname(mapas_path), exist_ok=True)
		tmp_path = mapas_path + '.tmp'
		with open(tmp_path, 'w', encoding='utf-8') as f:
			json.dump(data, f, ensure_ascii=False, indent=2)
		os.replace(tmp_path, mapas_path)
		return True
	except Exception:
		return False


# ----- Partitioned Mapas Storage (by month/year) -----
def _get_mapas_filepath(mes, ano):
	"""Retorna o caminho do arquivo de mapas particionado por mês/ano.
	
	Formato: dados/mapas_2025_09.json (setembro/2025)
	"""
	base_dir = os.path.dirname(os.path.dirname(__file__))
	filename = f'mapas_{ano}_{mes:02d}.json'
	return os.path.join(base_dir, 'dados', filename)


def _load_mapas_partitioned(mes, ano):
	"""Carrega mapas de um arquivo particionado específico (mês/ano).
	
	Retorna lista de mapas ou None se arquivo não existir.
	"""
	filepath = _get_mapas_filepath(mes, ano)
	if not os.path.isfile(filepath):
		return None
	try:
		with open(filepath, 'r', encoding='utf-8') as f:
			data = json.load(f)
			# Normalizar: sempre retornar lista
			if isinstance(data, dict) and 'mapas' in data:
				return data['mapas']
			elif isinstance(data, list):
				return data
			return None
	except Exception as e:
		print(f"❌ Erro ao ler {filepath}: {e}")
		return None


def _save_mapas_partitioned(mapas_list, mes, ano):
	"""Salva lista de mapas em arquivo particionado (mês/ano).
	
	Args:
		mapas_list: lista de dicionários representando mapas
		mes: número do mês (1-12)
		ano: ano (ex: 2025)
	
	Returns:
		True se salvou com sucesso, False caso contrário
	"""
	filepath = _get_mapas_filepath(mes, ano)
	try:
		os.makedirs(os.path.dirname(filepath), exist_ok=True)
		tmp_path = filepath + '.tmp'
		# Salvar como lista direta (sem wrapper)
		with open(tmp_path, 'w', encoding='utf-8') as f:
			json.dump(mapas_list, f, ensure_ascii=False, indent=2)
		os.replace(tmp_path, filepath)
		print(f"✅ Mapas salvos em: {filepath} ({len(mapas_list)} registros)")
		return True
	except Exception as e:
		print(f"❌ Erro ao salvar {filepath}: {e}")
		return False


def _load_mapas_by_period(mes_inicio, ano_inicio, mes_fim, ano_fim):
	"""Carrega todos os mapas dentro de um período (range de meses).
	
	Retorna lista agregada de todos os mapas encontrados.
	"""
	mapas_agregados = []
	
	# Iterar através dos meses
	ano_atual = ano_inicio
	mes_atual = mes_inicio
	
	while (ano_atual < ano_fim) or (ano_atual == ano_fim and mes_atual <= mes_fim):
		# Tentar carregar arquivo particionado
		mapas_mes = _load_mapas_partitioned(mes_atual, ano_atual)
		if mapas_mes:
			mapas_agregados.extend(mapas_mes)
			print(f"📂 Carregados {len(mapas_mes)} mapas de {mes_atual:02d}/{ano_atual}")
		
		# Avançar para próximo mês
		mes_atual += 1
		if mes_atual > 12:
			mes_atual = 1
			ano_atual += 1
	
	print(f"✅ Total de mapas carregados no período: {len(mapas_agregados)}")
	return mapas_agregados


def _detect_mes_ano_from_entry(entry):
	"""Detecta mês e ano de um registro de mapa.
	
	Retorna tupla (mes, ano) ou (None, None) se não puder detectar.
	"""
	if not isinstance(entry, dict):
		return (None, None)
	
	# Tentar extrair de várias chaves possíveis
	mes_keys = ('mes', 'month', 'mes_num', 'mesNumero', 'month_num')
	ano_keys = ('ano', 'year')
	
	mes = None
	ano = None
	
	for k in mes_keys:
		if k in entry:
			try:
				mes = int(entry[k])
				break
			except (ValueError, TypeError):
				continue
	
	for k in ano_keys:
		if k in entry:
			try:
				ano = int(entry[k])
				break
			except (ValueError, TypeError):
				continue
	
	return (mes, ano)


def _load_all_mapas_partitioned():
	"""Carrega TODOS os arquivos de mapas particionados disponíveis.
	
	Varre o diretório dados/ procurando por arquivos mapas_YYYY_MM.json
	e carrega todos eles em uma única lista agregada.
	
	Retorna lista de mapas ou lista vazia se nenhum arquivo encontrado.
	"""
	import glob
	
	base_dir = os.path.dirname(os.path.dirname(__file__))
	dados_dir = os.path.join(base_dir, 'dados')
	
	# Padrão: mapas_2025_09.json, mapas_2025_10.json, etc
	pattern = os.path.join(dados_dir, 'mapas_????_??.json')
	
	arquivos = glob.glob(pattern)
	mapas_agregados = []
	
	for filepath in sorted(arquivos):
		try:
			with open(filepath, 'r', encoding='utf-8') as f:
				data = json.load(f)
				# Normalizar: sempre trabalhar com lista
				if isinstance(data, dict) and 'mapas' in data:
					mapas_agregados.extend(data['mapas'])
				elif isinstance(data, list):
					mapas_agregados.extend(data)
				
				filename = os.path.basename(filepath)
				print(f"📂 Carregado: {filename} ({len(data) if isinstance(data, list) else len(data.get('mapas', []))} mapas)")
		except Exception as e:
			print(f"❌ Erro ao carregar {filepath}: {e}")
			continue
	
	print(f"✅ Total de mapas particionados carregados: {len(mapas_agregados)}")
	return mapas_agregados


def salvar_mapas_raw(payload):
	"""Salva o payload recebido em arquivos particionados por mês/ano.

	Retorna dict simples: {'success': True} ou {'success': False, 'error': '...'}.
	"""
	# Aceitar dict ou lista (um ou vários mapas)
	try:
		entries = []
		if isinstance(payload, list):
			entries = payload
		else:
			entries = [payload or {}]
		
		# Detectar mês/ano dos novos registros
		periodos_afetados = set()
		for entry in entries:
			mes, ano = _detect_mes_ano_from_entry(entry)
			if mes and ano:
				periodos_afetados.add((mes, ano))
		
		if not periodos_afetados:
			return {'success': False, 'error': 'Não foi possível detectar mês/ano dos dados'}
		
		# Carregar mapas existentes dos períodos afetados
		mapas_list = []
		for (mes, ano) in periodos_afetados:
			mapas_periodo = _load_mapas_partitioned(mes, ano)
			if mapas_periodo:
				mapas_list.extend(mapas_periodo)
		
		# Também tentar carregar do arquivo legado (backward compatibility)
		legacy_data = _load_mapas_data()
		if legacy_data:
			if isinstance(legacy_data, dict) and isinstance(legacy_data.get('mapas'), list):
				mapas_list.extend(legacy_data.get('mapas'))
			elif isinstance(legacy_data, list):
				mapas_list.extend(legacy_data)

		# coletar ids existentes
		existing_ids = {int(m.get('id')) for m in mapas_list if isinstance(m, dict) and isinstance(m.get('id'), int)}
		next_id = (max(existing_ids) + 1) if existing_ids else 0

		saved_ids = []
		saved_records = []
		for entry in entries:
			if not isinstance(entry, dict):
				# envolver valores não-dict
				entry = {'data': entry}
			# decidir id
			provided = entry.get('id')

			# Função auxiliar: extrair a tupla de chave (lote_id, unidade, mes, ano)
			def _extract_key(obj):
				# possíveis nomes
				lote_keys = ('lote_id', 'loteId', 'lote', 'loteId')
				unidade_keys = ('unidade', 'unidade_id', 'unidadeId', 'unidade_nome', 'unidadeNome')
				mes_keys = ('mes', 'month', 'mes_num', 'mesNumero', 'month_num')
				ano_keys = ('ano', 'year')
				try:
					lote_val = None
					for k in lote_keys:
						if k in obj:
							lote_val = obj.get(k)
							break
					if lote_val is None:
						return None
					unidade_val = None
					for k in unidade_keys:
						if k in obj:
							unidade_val = obj.get(k)
							break
					if unidade_val is None:
						return None
					mes_val = None
					for k in mes_keys:
						if k in obj:
							mes_val = obj.get(k)
							break
					if mes_val is None:
						return None
					ano_val = None
					for k in ano_keys:
						if k in obj:
							ano_val = obj.get(k)
							break
					if ano_val is None:
						return None
					# normalizar tipos
					try:
						lote_n = int(lote_val)
					except Exception:
						try:
							lote_n = int(str(lote_val).strip())
						except Exception:
							return None
					try:
						mes_n = int(mes_val)
					except Exception:
						try:
							mes_n = int(str(mes_val).strip())
						except Exception:
							return None
					try:
						ano_n = int(ano_val)
					except Exception:
						try:
							ano_n = int(str(ano_val).strip())
						except Exception:
							return None
					unidade_s = str(unidade_val).strip().lower()
					return (lote_n, unidade_s, mes_n, ano_n)
				except Exception:
					return None

			entry_key = _extract_key(entry)
			matched_index = None
			matched_record = None
			if entry_key is not None:
				# procurar por mapa existente com a mesma tupla
				for idx, existing_map in enumerate(mapas_list):
					try:
						existing_key = _extract_key(existing_map) if isinstance(existing_map, dict) else None
						if existing_key is not None and existing_key == entry_key:
							matched_index = idx
							matched_record = existing_map
							break
					except Exception:
						continue

			if matched_index is not None:
				# sobrescrever: manter o mesmo id do registro existente
				assigned = int(matched_record.get('id')) if isinstance(matched_record.get('id'), int) else matched_record.get('id')
				rec = dict(entry)
				# Se o registro contém texto tabular, tentar parsear para colunas numéricas
				possible_text_keys = ('texto', 'conteudo', 'dados', 'texto_raw', 'texto_mapas', 'mapa_texto')
				text_val = None
				used_text_key = None
				for k in possible_text_keys:
					if k in rec and rec.get(k) is not None:
						text_val = rec.get(k)
						used_text_key = k
						break
				if text_val is not None:
					parsed = parse_texto_tabular(text_val)
					if parsed.get('ok'):
						cols = parsed.get('colunas') or {}
						col_count = int(parsed.get('colunas_count') or 0)
						# exigir ao menos 9 colunas quando o texto foi enviado como mapa (upload de mapa)
						if col_count < 9:
							return {'success': False, 'error': f'Texto tabular contém colunas insuficientes: {col_count} (<9)'}
						for ck, cv in cols.items():
							rec[ck] = cv
						# atribuir contagens
						rec['linhas'] = parsed.get('linhas')
						rec['colunas_count'] = parsed.get('colunas_count')
						col_count = int(parsed.get('colunas_count') or 0)
						# Se o parse gerou apenas 1 coluna, tratar como dados_siisp (lista numérica)
						if col_count == 1:
							# mover coluna_0 para dados_siisp
							if 'coluna_0' in rec:
								rec['dados_siisp'] = rec.pop('coluna_0')
							else:
								rec['dados_siisp'] = []
						else:
							# renomear colunas nutricionais: coluna_1..coluna_8 -> nomes semânticos
							col_rename_map = {
								'coluna_1': 'cafe_interno',
								'coluna_2': 'cafe_funcionario',
								'coluna_3': 'almoco_interno',
								'coluna_4': 'almoco_funcionario',
								'coluna_5': 'lanche_interno',
								'coluna_6': 'lanche_funcionario',
								'coluna_7': 'jantar_interno',
								'coluna_8': 'jantar_funcionario'
							}
							for oldk, newk in col_rename_map.items():
								if oldk in rec:
									rec[newk] = rec.pop(oldk)
							# normalizar coluna_0 para 'datas' e remover coluna_0
							if 'coluna_0' in rec:
								try:
									datas = _normalizar_datas_coluna(rec.get('coluna_0'), rec)
									rec.pop('coluna_0', None)
									rec['datas'] = datas
								except Exception:
									pass
					else:
						rec['colunas_parse_error'] = parsed.get('error')
					# remover o campo de texto cru antes de salvar
					if used_text_key:
						try:
							rec.pop(used_text_key, None)
						except Exception:
							pass
				rec['id'] = assigned
				# Normalizar campo dados_siisp recebido diretamente (pode ser string com linhas)
				if 'dados_siisp' in rec:
					val = rec.get('dados_siisp')
					if isinstance(val, str):
						parsed_ds = parse_texto_tabular(val)
						if parsed_ds.get('ok') and int(parsed_ds.get('colunas_count') or 0) == 1:
							rec['dados_siisp'] = parsed_ds.get('colunas', {}).get('coluna_0', [])
						else:
							# se não conseguiu parsear, tornar lista vazia
							rec['dados_siisp'] = []
					elif not isinstance(val, list):
						rec['dados_siisp'] = []
				# garantir que sempre exista lista em dados_siisp
				if 'dados_siisp' not in rec or rec.get('dados_siisp') is None:
					rec['dados_siisp'] = []
				
				# Se dados_siisp estiver vazio, preencher com zeros baseado no número de dias do mês
				if not rec.get('dados_siisp') or len(rec.get('dados_siisp', [])) == 0:
					try:
						mes_num = int(rec.get('mes') or datetime.now().month)
						ano_num = int(rec.get('ano') or datetime.now().year)
						dias_no_mes = calendar.monthrange(ano_num, mes_num)[1]
						rec['dados_siisp'] = [0] * dias_no_mes
					except Exception:
						# Fallback: usar 31 dias se não conseguir calcular
						rec['dados_siisp'] = [0] * 31
				# preservar criado_em se houver, e anotar atualizado_em
				if 'criado_em' not in rec and matched_record.get('criado_em'):
					rec['criado_em'] = matched_record.get('criado_em')
				rec['atualizado_em'] = datetime.now().isoformat()
				
				# Calcular campos comparativos SIISP
				_calcular_campos_comparativos_siisp(rec)
				
				# validar comprimentos das listas diárias antes de sobrescrever
				valid_ok, valid_msg = _validate_map_day_lengths(rec)
				if not valid_ok:
					return {'success': False, 'error': f'Validação de tamanho falhou: {valid_msg}'}
				# substituir no lugar
				mapas_list[matched_index] = rec
				saved_ids.append(assigned)
				saved_records.append(rec)
				# marcar operação para este registro
				if 'operacoes' not in locals():
					operacoes = []
				operacoes.append('overwritten')
				# atualizar existing_ids set (id já existia)
				existing_ids.add(int(assigned))
				continue

			# Se o registro contém texto tabular, tentar parsear para colunas numéricas
			# aceitar 'texto' como entrada mas não armazená-lo: iremos parsear e remover
			possible_text_keys = ('texto', 'conteudo', 'dados', 'texto_raw', 'texto_mapas', 'mapa_texto')
			text_val = None
			used_text_key = None
			for k in possible_text_keys:
				if k in entry and entry.get(k) is not None:
					text_val = entry.get(k)
					used_text_key = k
					break
			if text_val is not None:
				parsed = parse_texto_tabular(text_val)
				if parsed.get('ok'):
					# anexar colunas e contagens no registro salvo
					cols = parsed.get('colunas') or {}
					# mover cada coluna para o nível superior do registro (chaves 'coluna_0', 'coluna_1', ...)
					for ck, cv in cols.items():
						entry[ck] = cv
					entry['linhas'] = parsed.get('linhas')
					entry['colunas_count'] = parsed.get('colunas_count')
					col_count = int(parsed.get('colunas_count') or 0)
					# exigir ao menos 9 colunas quando o texto foi enviado como mapa (upload de mapa)
					if col_count < 9:
						return {'success': False, 'error': f'Texto tabular contém colunas insuficientes: {col_count} (<9)'}
					# Se houver apenas 1 coluna no texto, tratá-la como dados_siisp
					if col_count == 1:
						if 'coluna_0' in entry:
							entry['dados_siisp'] = entry.pop('coluna_0')
						else:
							entry['dados_siisp'] = []
					else:
						# renomear colunas nutricionais: coluna_1..coluna_8 -> nomes semânticos
						col_rename_map = {
							'coluna_1': 'cafe_interno',
							'coluna_2': 'cafe_funcionario',
							'coluna_3': 'almoco_interno',
							'coluna_4': 'almoco_funcionario',
							'coluna_5': 'lanche_interno',
							'coluna_6': 'lanche_funcionario',
							'coluna_7': 'jantar_interno',
							'coluna_8': 'jantar_funcionario'
						}
						for oldk, newk in col_rename_map.items():
							if oldk in entry:
								entry[newk] = entry.pop(oldk)
						# normalizar coluna_0 para 'datas' com padrão DD/MM/YYYY baseado em mes/ano do registro
						if 'coluna_0' in entry:
							try:
								datas = _normalizar_datas_coluna(entry.get('coluna_0'), entry)
								# remover coluna_0 após normalizar
								entry.pop('coluna_0', None)
								entry['datas'] = datas
							except Exception:
								# não interromper o salvamento em caso de falha na normalização
								pass
				else:
					# anotar erro de parse, mas prosseguir com o salvamento bruto
					entry['colunas_parse_error'] = parsed.get('error')
				# remover o campo de texto cru antes de salvar (usuário pediu não armazenar 'texto')
				if used_text_key:
					try:
						entry.pop(used_text_key, None)
					except Exception:
						pass
			if provided is None:
				assigned = next_id
				next_id += 1
			else:
				# validar id fornecido
				try:
					pid = int(provided)
				except Exception:
					return {'success': False, 'error': 'ID inválido fornecido'}
				if pid in existing_ids or pid in saved_ids:
					return {'success': False, 'error': f'ID já existe: {pid}'}
				assigned = pid
				if pid >= next_id:
					next_id = pid + 1

			rec = dict(entry)
			rec['id'] = assigned
			if 'criado_em' not in rec:
				rec['criado_em'] = datetime.now().isoformat()
			# Normalizar campo dados_siisp recebido diretamente (pode ser string com linhas)
			if 'dados_siisp' in rec:
				val = rec.get('dados_siisp')
				if isinstance(val, str):
					parsed_ds = parse_texto_tabular(val)
					if parsed_ds.get('ok') and int(parsed_ds.get('colunas_count') or 0) == 1:
						rec['dados_siisp'] = parsed_ds.get('colunas', {}).get('coluna_0', [])
					else:
						rec['dados_siisp'] = []
				elif not isinstance(val, list):
					rec['dados_siisp'] = []
			# garantir lista vazia quando não enviado
			if 'dados_siisp' not in rec or rec.get('dados_siisp') is None:
				rec['dados_siisp'] = []
			
			# Se dados_siisp estiver vazio, preencher com zeros baseado no número de dias do mês
			if not rec.get('dados_siisp') or len(rec.get('dados_siisp', [])) == 0:
				try:
					mes_num = int(rec.get('mes') or datetime.now().month)
					ano_num = int(rec.get('ano') or datetime.now().year)
					dias_no_mes = calendar.monthrange(ano_num, mes_num)[1]
					rec['dados_siisp'] = [0] * dias_no_mes
				except Exception:
					# Fallback: usar 31 dias se não conseguir calcular
					rec['dados_siisp'] = [0] * 31
			
			# Calcular campos comparativos SIISP
			_calcular_campos_comparativos_siisp(rec)
			
			# validar comprimentos das listas diárias antes de salvar
			valid_ok, valid_msg = _validate_map_day_lengths(rec)
			if not valid_ok:
				return {'success': False, 'error': f'Validação de tamanho falhou: {valid_msg}'}
			mapas_list.append(rec)
			saved_ids.append(assigned)
			saved_records.append(rec)
			# marcar operação de criação para este registro
			if 'operacoes' not in locals():
				operacoes = []
			operacoes.append('created')

		# NOVO: Sistema particionado - agrupar mapas por mês/ano e salvar separadamente
		mapas_por_periodo = {}  # {(mes, ano): [mapas...]}
		
		for mapa in mapas_list:
			mes, ano = _detect_mes_ano_from_entry(mapa)
			if mes and ano:
				key = (mes, ano)
				if key not in mapas_por_periodo:
					mapas_por_periodo[key] = []
				mapas_por_periodo[key].append(mapa)
			else:
				# Se não detectar mes/ano, tentar salvar no legado
				print(f"⚠️ Mapa sem mes/ano detectável: {mapa.get('id', 'sem id')}")
		
		# Salvar cada período em seu próprio arquivo
		all_saved = True
		for (mes, ano), mapas_periodo in mapas_por_periodo.items():
			ok = _save_mapas_partitioned(mapas_periodo, mes, ano)
			if not ok:
				all_saved = False
				print(f"❌ Falha ao salvar mapas de {mes:02d}/{ano}")
		
		if not all_saved:
			return {'success': False, 'error': 'Erro ao salvar alguns arquivos de mapas'}
		
		# Retorno enriquecido: id(s) e registro(s) salvos
		if len(saved_records) == 1:
			ret = {'success': True, 'id': saved_records[0]['id'], 'registro': saved_records[0]}
			if 'operacoes' in locals() and isinstance(operacoes, list) and len(operacoes) == 1:
				ret['operacao'] = operacoes[0]
			return ret
		# múltiplos registros: incluir lista de operações paralela aos ids/registros
		ret = {'success': True, 'ids': saved_ids, 'registros': saved_records}
		if 'operacoes' in locals() and isinstance(operacoes, list):
			ret['operacoes'] = operacoes
		return ret
	except Exception:
		return {'success': False, 'error': 'Erro ao salvar mapas'}


def preparar_dados_entrada_manual(data):
	"""Prepara dados de entrada manual para salvamento.
	
	Converte formato tabular, normaliza arrays, gera datas e adiciona metadados necessários.
	
	Retorna: {'success': True, 'data': dados_preparados} ou {'success': False, 'error': '...'}
	"""
	try:
		if not isinstance(data, dict):
			return {'success': False, 'error': 'Dados inválidos'}
		
		# Clonar dados para não modificar original
		import copy
		processed = copy.deepcopy(data)
		
		# Campos de refeições
		meal_fields = [
			'cafe_interno', 'cafe_funcionario',
			'almoco_interno', 'almoco_funcionario', 
			'lanche_interno', 'lanche_funcionario',
			'jantar_interno', 'jantar_funcionario'
		]
		
		# Verificar se os dados vêm no formato tabular (dados_tabela) e converter
		if 'dados_tabela' in processed and isinstance(processed['dados_tabela'], list):
			tabela = processed['dados_tabela']
			
			# Criar arrays vazios
			for field in meal_fields:
				processed[field] = []
			
			# Converter cada dia da tabela para arrays
			for dia_data in tabela:
				for field in meal_fields:
					valor = dia_data.get(field, 0)
					try:
						valor_int = int(valor) if valor is not None and valor != '' else 0
					except (ValueError, TypeError):
						valor_int = 0
					processed[field].append(valor_int)
			
			# Remover o campo dados_tabela (não é necessário)
			del processed['dados_tabela']
		
		else:
			# Normalizar arrays vazios/nulos para 0 (formato direto)
			def normalizar_array(arr):
				if not isinstance(arr, list):
					return []
				normalized = []
				for item in arr:
					if item is None or item == '' or item == 'null':
						normalized.append(0)
					else:
						try:
							normalized.append(int(item))
						except (ValueError, TypeError):
							normalized.append(0)
				return normalized
			
			for field in meal_fields:
				if field in processed:
					processed[field] = normalizar_array(processed.get(field))
		
		# Determinar número de dias baseado no maior array
		max_days = 0
		for field in meal_fields:
			if field in processed and isinstance(processed[field], list):
				max_days = max(max_days, len(processed[field]))
		
		# Gerar array de datas baseado em mes/ano e max_days
		mes = processed.get('mes')
		ano = processed.get('ano')
		datas = []
		
		if mes and ano:
			try:
				mes = int(mes)
				ano = int(ano)
				days_in_month = calendar.monthrange(ano, mes)[1]
				# Usar o menor entre max_days e dias no mês
				num_days = min(max_days, days_in_month) if max_days > 0 else days_in_month
				
				for dia in range(1, num_days + 1):
					data_str = f"{dia:02d}/{mes:02d}/{ano}"
					datas.append(data_str)
			except:
				# Fallback: gerar datas baseado apenas em max_days
				for dia in range(1, max_days + 1):
					data_str = f"{dia:02d}/01/2025"  # fallback genérico
					datas.append(data_str)
		
		# Adicionar campos de estrutura completa
		processed['datas'] = datas
		processed['linhas'] = len(datas)
		processed['colunas_count'] = 9  # Sempre 9 colunas (tipos de refeição)
		
		# Adicionar timestamp de criação
		processed['criado_em'] = datetime.now().isoformat()
		
		# Garantir que dados_siisp existe como array vazio
		if 'dados_siisp' not in processed:
			processed['dados_siisp'] = []
		
		# Se dados_siisp estiver vazio, preencher com zeros baseado no número de dias do mês
		if not processed.get('dados_siisp') or len(processed.get('dados_siisp', [])) == 0:
			try:
				mes = int(processed.get('mes') or datetime.now().month)
				ano = int(processed.get('ano') or datetime.now().year)
				dias_no_mes = calendar.monthrange(ano, mes)[1]
				processed['dados_siisp'] = [0] * dias_no_mes
			except Exception:
				# Fallback: usar o número de datas geradas
				processed['dados_siisp'] = [0] * len(datas)
		
		# Calcular campos comparativos SIISP
		_calcular_campos_comparativos_siisp(processed)
		
		return {'success': True, 'data': processed}
		
	except Exception as e:
		return {'success': False, 'error': f'Erro ao preparar dados: {str(e)}'}


def reordenar_registro_mapas(registro_id):
	"""Reordena um registro específico no arquivo mapas.json para garantir ordem correta dos campos.
	
	Args:
		registro_id: ID do registro a ser reordenado
	
	Retorna: True se sucesso, False se erro
	"""
	try:
		base_dir = os.path.dirname(os.path.dirname(__file__))
		mapas_file = os.path.join(base_dir, 'dados', 'mapas.json')
		
		if not os.path.exists(mapas_file):
			return False
		
		with open(mapas_file, 'r', encoding='utf-8') as f:
			mapas_data = json.load(f)
		
		if not isinstance(mapas_data, list):
			return False
		
		# Encontrar e reordenar o registro específico
		for i, mapa in enumerate(mapas_data):
			if isinstance(mapa, dict) and mapa.get('id') == registro_id:
				# Ordem específica dos campos
				field_order = [
					'lote_id', 'mes', 'ano', 'unidade', 'dados_siisp', 'linhas', 'colunas_count',
					'cafe_interno', 'cafe_funcionario', 'almoco_interno', 'almoco_funcionario',
					'lanche_interno', 'lanche_funcionario', 'jantar_interno', 'jantar_funcionario',
					'datas', 'id', 'criado_em', 'atualizado_em'
				]
				
				ordered_data = {}
				for field in field_order:
					if field in mapa:
						ordered_data[field] = mapa[field]
				
				# Adicionar campos não listados
				for field, value in mapa.items():
					if field not in ordered_data:
						ordered_data[field] = value
				
				mapas_data[i] = ordered_data
				break
		
		# Salvar arquivo reordenado
		with open(mapas_file, 'w', encoding='utf-8') as f:
			json.dump(mapas_data, f, indent=2, ensure_ascii=False)
		
		return True
		
	except Exception:
		return False


def adicionar_siisp_em_mapa(payload):
	"""Adiciona dados SIISP a um mapa existente.
	
	Args:
		payload: dict contendo:
			- unidade: nome da unidade
			- mes: número do mês (1-12)
			- ano: ano (ex: 2025)
			- lote_id: ID do lote
			- dados_siisp: texto tabular com dados SIISP ou lista de números
	
	Retorna:
		dict com 'success' (bool) e 'error' (str) ou 'registro' (dict)
	"""
	if not isinstance(payload, dict):
		return {'success': False, 'error': 'Payload inválido'}
	
	unidade = payload.get('unidade', '').strip()
	mes = payload.get('mes')
	ano = payload.get('ano')
	lote_id = payload.get('lote_id')
	dados_siisp_raw = payload.get('dados_siisp')
	
	# Validações básicas
	if not unidade or not mes or not ano or lote_id is None:
		return {'success': False, 'error': 'Campos obrigatórios ausentes: unidade, mes, ano, lote_id'}
	
	try:
		mes = int(mes)
		ano = int(ano)
		lote_id = int(lote_id)
	except Exception:
		return {'success': False, 'error': 'Valores inválidos para mes, ano ou lote_id'}
	
	if mes < 1 or mes > 12:
		return {'success': False, 'error': 'Mês deve estar entre 1 e 12'}
	
	# Calcular dias esperados do mês
	import calendar
	try:
		dias_esperados = calendar.monthrange(ano, mes)[1]
	except Exception:
		return {'success': False, 'error': 'Combinação inválida de mês/ano'}
	
	# Processar dados_siisp
	dados_siisp_list = []
	if isinstance(dados_siisp_raw, str):
		# Texto tabular - parsear
		parsed = parse_texto_tabular(dados_siisp_raw)
		if not parsed.get('ok'):
			return {'success': False, 'error': f'Erro ao processar dados SIISP: {parsed.get("error")}'}
		
		# Pegar primeira coluna
		colunas = parsed.get('colunas', {})
		primeira_coluna_key = 'coluna_0'
		if primeira_coluna_key in colunas:
			dados_siisp_list = colunas[primeira_coluna_key]
		else:
			return {'success': False, 'error': 'Dados SIISP não encontrados na primeira coluna'}
	elif isinstance(dados_siisp_raw, list):
		# Se é lista, tentar converter cada item para número
		for item in dados_siisp_raw:
			try:
				if isinstance(item, (int, float)):
					dados_siisp_list.append(int(item))
				elif isinstance(item, str):
					# Remover espaços e converter
					num = int(item.strip())
					dados_siisp_list.append(num)
				else:
					dados_siisp_list.append(0)
			except Exception:
				dados_siisp_list.append(0)
	elif dados_siisp_raw is None or dados_siisp_raw == '':
		return {'success': False, 'error': 'Dados SIISP não fornecidos'}
	else:
		return {'success': False, 'error': f'Formato de dados SIISP inválido: tipo {type(dados_siisp_raw).__name__}. Esperado: string ou lista'}
	
	# Validar quantidade de dias
	if len(dados_siisp_list) != dias_esperados:
		return {
			'success': False, 
			'error': f'Quantidade de dados SIISP ({len(dados_siisp_list)}) não corresponde aos dias do mês ({dias_esperados} dias em {mes:02d}/{ano})'
		}
	
	# Buscar mapa existente usando sistema particionado
	mapas_existentes = _load_mapas_partitioned(mes, ano)
	if mapas_existentes is None:
		return {
			'success': False,
			'error': f'Nenhum mapa encontrado para {mes:02d}/{ano}. Adicione dados de refeições primeiro.'
		}
	
	# Procurar mapa específico
	mapa_encontrado = None
	indice_mapa = None
	for i, m in enumerate(mapas_existentes):
		if not isinstance(m, dict):
			continue
		m_unidade = str(m.get('unidade', '')).strip()
		m_lote_id = m.get('lote_id')
		m_mes = m.get('mes')
		m_ano = m.get('ano')
		
		try:
			if (m_unidade.lower() == unidade.lower() and 
				int(m_lote_id) == int(lote_id) and
				int(m_mes) == int(mes) and
				int(m_ano) == int(ano)):
				mapa_encontrado = m
				indice_mapa = i
				break
		except Exception:
			continue
	
	if mapa_encontrado is None:
		return {
			'success': False,
			'error': f'Mapa não encontrado para Unidade "{unidade}", Lote {lote_id}, período {mes:02d}/{ano}. Adicione dados de refeições primeiro.'
		}
	
	# Atualizar dados_siisp no mapa
	mapa_encontrado['dados_siisp'] = dados_siisp_list
	mapa_encontrado['atualizado_em'] = datetime.now().isoformat()
	
	# Recalcular campos comparativos SIISP
	_calcular_campos_comparativos_siisp(mapa_encontrado)
	
	# Atualizar no array
	mapas_existentes[indice_mapa] = mapa_encontrado
	
	# Salvar arquivo particionado
	if not _save_mapas_partitioned(mapas_existentes, mes, ano):
		return {'success': False, 'error': 'Erro ao salvar dados'}
	
	return {
		'success': True,
		'registro': mapa_encontrado,
		'mensagem': f'Dados SIISP adicionados com sucesso ao mapa {mapa_encontrado.get("id")}'
	}


def excluir_mapa(payload):
	"""Remove um mapa específico do arquivo particionado.
	
	Args:
		payload: dict contendo:
			- unidade: nome da unidade
			- mes: número do mês (1-12)
			- ano: ano (ex: 2025)
			- lote_id: ID do lote (opcional, para validação adicional)
	
	Retorna:
		dict com 'success' (bool) e 'error' (str) ou 'mensagem' (str)
	"""
	if not isinstance(payload, dict):
		return {'success': False, 'error': 'Payload inválido'}
	
	unidade = payload.get('unidade', '').strip()
	mes = payload.get('mes')
	ano = payload.get('ano')
	lote_id = payload.get('lote_id')  # Opcional
	
	# Validações básicas
	if not unidade or not mes or not ano:
		return {'success': False, 'error': 'Campos obrigatórios ausentes: unidade, mes, ano'}
	
	try:
		mes = int(mes)
		ano = int(ano)
		if lote_id is not None:
			lote_id = int(lote_id)
	except Exception:
		return {'success': False, 'error': 'Valores inválidos para mes, ano ou lote_id'}
	
	if mes < 1 or mes > 12:
		return {'success': False, 'error': 'Mês deve estar entre 1 e 12'}
	
	# Buscar mapa existente usando sistema particionado
	mapas_existentes = _load_mapas_partitioned(mes, ano)
	if mapas_existentes is None:
		return {
			'success': False,
			'error': f'Nenhum mapa encontrado para {mes:02d}/{ano}. Não há dados para excluir.'
		}
	
	# Procurar mapa específico
	mapa_encontrado = None
	indice_mapa = None
	for i, m in enumerate(mapas_existentes):
		if not isinstance(m, dict):
			continue
		m_unidade = str(m.get('unidade', '')).strip()
		m_mes = m.get('mes')
		m_ano = m.get('ano')
		
		# Comparação básica: unidade, mês e ano
		try:
			if (m_unidade.lower() == unidade.lower() and 
				int(m_mes) == int(mes) and
				int(m_ano) == int(ano)):
				# Se lote_id foi fornecido, validar também
				if lote_id is not None:
					m_lote_id = m.get('lote_id')
					if int(m_lote_id) != int(lote_id):
						continue
				mapa_encontrado = m
				indice_mapa = i
				break
		except Exception:
			continue
	
	if mapa_encontrado is None:
		msg_extra = f' e Lote {lote_id}' if lote_id is not None else ''
		return {
			'success': False,
			'error': f'Mapa não encontrado para Unidade "{unidade}", período {mes:02d}/{ano}{msg_extra}.'
		}
	
	# Guardar informações para mensagem de confirmação
	mapa_id = mapa_encontrado.get('id')
	
	# Remover mapa do array
	mapas_existentes.pop(indice_mapa)
	
	# Salvar arquivo particionado atualizado (ou deletar se ficou vazio)
	if len(mapas_existentes) == 0:
		# Se não sobrou nenhum mapa, deletar o arquivo
		filepath = _get_mapas_filepath(mes, ano)
		try:
			if os.path.isfile(filepath):
				os.remove(filepath)
				print(f"🗑️ Arquivo vazio deletado: {filepath}")
		except Exception as e:
			print(f"⚠️ Aviso: Não foi possível deletar arquivo vazio: {e}")
			# Não é erro crítico, continuar
	else:
		# Ainda há mapas, salvar arquivo atualizado
		if not _save_mapas_partitioned(mapas_existentes, mes, ano):
			return {'success': False, 'error': 'Erro ao salvar dados após exclusão'}
	
	return {
		'success': True,
		'mensagem': f'Mapa {mapa_id} da unidade "{unidade}" ({mes:02d}/{ano}) excluído com sucesso.',
		'id': mapa_id
	}


# ----- Excel Export Helper -----
def int_to_roman(num):
	"""Converte número inteiro em algarismo romano"""
	val = [
		1000, 900, 500, 400,
		100, 90, 50, 40,
		10, 9, 5, 4,
		1
	]
	syms = [
		"M", "CM", "D", "CD",
		"C", "XC", "L", "XL",
		"X", "IX", "V", "IV",
		"I"
	]
	roman_num = ''
	i = 0
	while num > 0:
		for _ in range(num // val[i]):
			roman_num += syms[i]
			num -= val[i]
		i += 1
	return roman_num


def gerar_excel_exportacao(lote_id, unidades_list, data_inicio=None, data_fim=None):
	"""Gera arquivo Excel com dados do lote para exportação.
	
	Args:
		lote_id: ID do lote
		unidades_list: lista de nomes de unidades (vazia = todas)
		data_inicio: data de início do filtro (opcional)
		data_fim: data de fim do filtro (opcional)
	
	Retorna:
		dict com:
		- success: bool
		- error: str (se falha)
		- output: BytesIO (se sucesso)
		- filename: str (se sucesso)
	"""
	try:
		from openpyxl import load_workbook
		from copy import copy
		import io
		import re
		import os
		import glob
		import calendar
	except ImportError as e:
		return {'success': False, 'error': f'Biblioteca não instalada: {str(e)}'}
	
	try:
		# Carregar dados
		dashboard_data = carregar_lotes_para_dashboard()
		lotes = dashboard_data.get('lotes', [])
		
		# Buscar lote específico
		lote = None
		for l in lotes:
			try:
				if int(l.get('id')) == int(lote_id):
					lote = l
					break
			except Exception:
				continue
		
		if not lote:
			return {'success': False, 'error': 'Lote não encontrado'}
		
		precos = normalizar_precos(lote.get('precos', {}))

		# Buscar mapas (usando sistema particionado)
		mapas_filtrados = []
		base_dir = os.path.dirname(os.path.dirname(__file__))
		dados_dir = os.path.join(base_dir, 'dados')
		
		mapas_files = glob.glob(os.path.join(dados_dir, 'mapas_*.json'))
		
		for mapas_file in mapas_files:
			filename = os.path.basename(mapas_file)
			match = re.search(r'mapas_(\d{4})_(\d{2})\.json', filename)
			if match:
				ano_arquivo = int(match.group(1))
				mes_arquivo = int(match.group(2))
				
				mapas_mes = _load_mapas_partitioned(mes_arquivo, ano_arquivo)
				
				for m in mapas_mes:
					try:
						if int(m.get('lote_id')) != int(lote_id):
							continue
						
						if unidades_list:
							unidade_nome = (m.get('unidade') or '').strip()
							if unidade_nome not in unidades_list:
								continue
						
						mapas_filtrados.append(m)
					except Exception:
						continue

		if not mapas_filtrados:
			return {'success': False, 'error': 'Nenhum dado encontrado para os filtros selecionados'}

		# Carregar modelo Excel
		modelo_path = os.path.join(dados_dir, 'modelo.xlsx')
		
		if not os.path.exists(modelo_path):
			return {'success': False, 'error': 'Arquivo modelo.xlsx não encontrado'}

		wb = load_workbook(modelo_path)

		# Selecionar planilha COMPARATIVO
		if 'COMPARATIVO' in wb.sheetnames:
			ws1 = wb['COMPARATIVO']
		else:
			ws1 = wb.active
			ws1.title = 'COMPARATIVO'

		# Definir ordem dos preços (usado em RESUMO e COMPARATIVO)
		precos_ordem = [
			('cafe', 'interno'),
			('cafe', 'funcionario'),
			('almoco', 'interno'),
			('almoco', 'funcionario'),
			('lanche', 'interno'),
			('lanche', 'funcionario'),
			('jantar', 'interno'),
			('jantar', 'funcionario')
		]

		# ===== PREENCHER PLANILHA RESUMO =====
		if 'RESUMO' in wb.sheetnames:
			ws_resumo = wb['RESUMO']
			
			# Preencher B8 com número do contrato
			contrato_numero = lote.get('contrato', '')
			ws_resumo['B8'] = f"CONTRATO : {contrato_numero}"
			
			# Preencher B7 com texto dinâmico
			empresa_nome = lote.get('empresa', '')
			mes = None
			ano = None
			if mapas_filtrados:
				mes = mapas_filtrados[0].get('mes')
				ano = mapas_filtrados[0].get('ano')
			
			# Mês em português
			meses_pt = [
				'', 'JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO',
				'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO'
			]
			mes_nome = meses_pt[mes] if mes and 1 <= mes <= 12 else ''
			
			# Lote em romano
			lote_romano = int_to_roman(lote_id) if lote_id else ''
			texto_resumo = f"RESUMO FINAL LOTE {lote_romano} - EMPRESA {empresa_nome} - {mes_nome} {ano}".upper()
			ws_resumo['B7'] = texto_resumo
			
			# Desmesclar células mescladas na área de dados
			for merged_range in list(ws_resumo.merged_cells.ranges):
				min_col = merged_range.min_col
				max_col = merged_range.max_col
				min_row = merged_range.min_row
				max_row = merged_range.max_row
				if max_row >= 13 and min_col >= 2 and max_col <= 11:
					ws_resumo.unmerge_cells(str(merged_range))
			
			# Preparar lista de unidades
			if unidades_list:
				nomes_unidades = unidades_list
			else:
				nomes_unidades = list(set(m.get('unidade', '') for m in mapas_filtrados if m.get('unidade')))
				nomes_unidades.sort()
			
			quantidade_unidades = len(nomes_unidades)
			estilo_b11 = ws_resumo['B11']
			
			if quantidade_unidades == 1:
				# Caso de 1 unidade
				cell_b11 = ws_resumo['B11']
				cell_b11.value = 1
				cell_b11.font = copy(estilo_b11.font)
				cell_b11.border = copy(estilo_b11.border)
				cell_b11.alignment = copy(estilo_b11.alignment)
				cell_b11.number_format = estilo_b11.number_format
				cell_b11.protection = copy(estilo_b11.protection)
				
				cell_c11 = ws_resumo['C11']
				cell_c11.value = nomes_unidades[0]
				cell_c11.font = copy(estilo_b11.font)
				cell_c11.border = copy(estilo_b11.border)
				cell_c11.alignment = copy(estilo_b11.alignment)
				cell_c11.number_format = estilo_b11.number_format
				cell_c11.protection = copy(estilo_b11.protection)
				
				colunas_refeicoes = [
					('cafe_interno', 4),
					('cafe_funcionario', 5),
					('almoco_interno', 6),
					('almoco_funcionario', 7),
					('lanche_interno', 8),
					('lanche_funcionario', 9),
					('jantar_interno', 10),
					('jantar_funcionario', 11)
				]
				nome_unidade = nomes_unidades[0]
				mapa_unidade = next((m for m in mapas_filtrados if m.get('unidade') == nome_unidade), None)
				
				for campo, col_idx in colunas_refeicoes:
					valor_total = 0
					if mapa_unidade:
						valores = mapa_unidade.get(campo, [])
						valor_total = sum(int(v or 0) for v in valores)
					
					cell = ws_resumo.cell(row=11, column=col_idx, value=valor_total)
					cell.font = copy(estilo_b11.font)
					cell.border = copy(estilo_b11.border)
					cell.alignment = copy(estilo_b11.alignment)
					cell.number_format = estilo_b11.number_format
					cell.protection = copy(estilo_b11.protection)
				
				ws_resumo.merge_cells(start_row=13, start_column=2, end_row=16, end_column=2)
				
				linha_precos = 13
				for col_offset, (ref, tipo) in enumerate(precos_ordem):
					col_idx = 4 + col_offset
					valor_preco = precos.get(ref, {}).get(tipo, 0)
					cell = ws_resumo.cell(row=linha_precos, column=col_idx, value=valor_preco)
					cell_modelo = ws_resumo.cell(row=13, column=4)
					cell.font = copy(cell_modelo.font)
					cell.border = copy(cell_modelo.border)
					cell.alignment = copy(cell_modelo.alignment)
					cell.number_format = cell_modelo.number_format
					cell.protection = copy(cell_modelo.protection)
				
				for campo, col_idx in colunas_refeicoes:
					soma_total = 0
					for nome in nomes_unidades:
						mapa = next((m for m in mapas_filtrados if m.get('unidade') == nome), None)
						if mapa:
							valores = mapa.get(campo, [])
							soma_total += sum(int(v or 0) for v in valores)
					
					cell = ws_resumo.cell(row=14, column=col_idx, value=soma_total)
					cell_modelo = ws_resumo.cell(row=14, column=col_idx)
					cell.font = copy(cell_modelo.font)
					cell.border = copy(cell_modelo.border)
					cell.alignment = copy(cell_modelo.alignment)
					cell.number_format = cell_modelo.number_format
					cell.protection = copy(cell_modelo.protection)
				
				for col_idx in range(4, 12):
					preco = ws_resumo.cell(row=13, column=col_idx).value or 0
					soma = ws_resumo.cell(row=14, column=col_idx).value or 0
					produto = float(preco) * float(soma)
					
					cell = ws_resumo.cell(row=15, column=col_idx, value=produto)
					cell_modelo = ws_resumo.cell(row=15, column=4)
					cell.font = copy(cell_modelo.font)
					cell.border = copy(cell_modelo.border)
					cell.alignment = copy(cell_modelo.alignment)
					cell.number_format = cell_modelo.number_format
					cell.protection = copy(cell_modelo.protection)
				
				ws_resumo.merge_cells('D16:K16')
				soma_produtos = sum(ws_resumo.cell(row=15, column=col_idx).value or 0 for col_idx in range(4, 12))
				cell_total = ws_resumo['D16']
				cell_modelo = ws_resumo.cell(row=16, column=4)
				cell_total.value = soma_produtos
				cell_total.font = copy(cell_modelo.font)
				cell_total.border = copy(cell_modelo.border)
				cell_total.alignment = copy(cell_modelo.alignment)
				cell_total.number_format = cell_modelo.number_format
				cell_total.protection = copy(cell_modelo.protection)
				
			else:
				# Caso de múltiplas unidades
				ws_resumo.insert_rows(11, amount=quantidade_unidades - 1)
				
				for idx, nome_unidade in enumerate(nomes_unidades):
					linha_atual = 11 + idx
					
					cell_b = ws_resumo.cell(row=linha_atual, column=2, value=idx + 1)
					cell_b.font = copy(estilo_b11.font)
					cell_b.border = copy(estilo_b11.border)
					cell_b.alignment = copy(estilo_b11.alignment)
					cell_b.number_format = estilo_b11.number_format
					cell_b.protection = copy(estilo_b11.protection)
					
					cell_c = ws_resumo.cell(row=linha_atual, column=3, value=nome_unidade)
					cell_c.font = copy(estilo_b11.font)
					cell_c.border = copy(estilo_b11.border)
					cell_c.alignment = copy(estilo_b11.alignment)
					cell_c.number_format = estilo_b11.number_format
					cell_c.protection = copy(estilo_b11.protection)
					
					mapa_unidade = next((m for m in mapas_filtrados if m.get('unidade') == nome_unidade), None)
					
					colunas_refeicoes = [
						('cafe_interno', 4),
						('cafe_funcionario', 5),
						('almoco_interno', 6),
						('almoco_funcionario', 7),
						('lanche_interno', 8),
						('lanche_funcionario', 9),
						('jantar_interno', 10),
						('jantar_funcionario', 11)
					]
					
					for campo, col_idx in colunas_refeicoes:
						valor_total = 0
						if mapa_unidade:
							valores = mapa_unidade.get(campo, [])
							valor_total = sum(int(v or 0) for v in valores)
						
						cell = ws_resumo.cell(row=linha_atual, column=col_idx, value=valor_total)
						cell.font = copy(estilo_b11.font)
						cell.border = copy(estilo_b11.border)
						cell.alignment = copy(estilo_b11.alignment)
						cell.number_format = estilo_b11.number_format
						cell.protection = copy(estilo_b11.protection)
				
				linha_inicio_resumo = 11 + quantidade_unidades + 1
				
				ws_resumo.merge_cells(start_row=linha_inicio_resumo, start_column=2, 
									 end_row=linha_inicio_resumo + 3, end_column=2)
				
				linha_precos = linha_inicio_resumo
				for col_offset, (ref, tipo) in enumerate(precos_ordem):
					col_idx = 4 + col_offset
					valor_preco = precos.get(ref, {}).get(tipo, 0)
					cell = ws_resumo.cell(row=linha_precos, column=col_idx, value=valor_preco)
					cell_modelo = ws_resumo.cell(row=linha_inicio_resumo, column=4)
					cell.font = copy(cell_modelo.font)
					cell.border = copy(cell_modelo.border)
					cell.alignment = copy(cell_modelo.alignment)
					cell.number_format = cell_modelo.number_format
					cell.protection = copy(cell_modelo.protection)
				
				linha_somas = linha_inicio_resumo + 1
				colunas_refeicoes = [
					('cafe_interno', 4),
					('cafe_funcionario', 5),
					('almoco_interno', 6),
					('almoco_funcionario', 7),
					('lanche_interno', 8),
					('lanche_funcionario', 9),
					('jantar_interno', 10),
					('jantar_funcionario', 11)
				]
				
				for campo, col_idx in colunas_refeicoes:
					soma_total = 0
					for nome in nomes_unidades:
						mapa = next((m for m in mapas_filtrados if m.get('unidade') == nome), None)
						if mapa:
							valores = mapa.get(campo, [])
							soma_total += sum(int(v or 0) for v in valores)
					
					cell = ws_resumo.cell(row=linha_somas, column=col_idx, value=soma_total)
					cell_modelo = ws_resumo.cell(row=linha_somas, column=col_idx)
					cell.font = copy(cell_modelo.font)
					cell.border = copy(cell_modelo.border)
					cell.alignment = copy(cell_modelo.alignment)
					cell.number_format = cell_modelo.number_format
					cell.protection = copy(cell_modelo.protection)
				
				linha_produtos = linha_inicio_resumo + 2
				for col_idx in range(4, 12):
					preco = ws_resumo.cell(row=linha_precos, column=col_idx).value or 0
					soma = ws_resumo.cell(row=linha_somas, column=col_idx).value or 0
					produto = float(preco) * float(soma)
					
					cell = ws_resumo.cell(row=linha_produtos, column=col_idx, value=produto)
					cell_modelo = ws_resumo.cell(row=linha_produtos, column=4)
					cell.font = copy(cell_modelo.font)
					cell.border = copy(cell_modelo.border)
					cell.alignment = copy(cell_modelo.alignment)
					cell.number_format = cell_modelo.number_format
					cell.protection = copy(cell_modelo.protection)
				
				linha_total = linha_inicio_resumo + 3
				ws_resumo.merge_cells(start_row=linha_total, start_column=4, 
									 end_row=linha_total, end_column=11)
				soma_produtos = sum(ws_resumo.cell(row=linha_produtos, column=col_idx).value or 0 
								   for col_idx in range(4, 12))
				cell_total = ws_resumo.cell(row=linha_total, column=4)
				cell_modelo = ws_resumo.cell(row=linha_total, column=4)
				cell_total.value = soma_produtos
				cell_total.font = copy(cell_modelo.font)
				cell_total.border = copy(cell_modelo.border)
				cell_total.alignment = copy(cell_modelo.alignment)
				cell_total.number_format = cell_modelo.number_format
				cell_total.protection = copy(cell_modelo.protection)

		# ===== PREENCHER PLANILHA COMPARATIVO =====
		col_inicio = 13  # M = 13
		for idx, (ref, tipo) in enumerate(precos_ordem):
			col = col_inicio + idx
			valor_preco = precos.get(ref, {}).get(tipo, 0)
			cell_preco = ws1.cell(row=6, column=col, value=valor_preco)
			cell_modelo = ws1.cell(row=6, column=col)
			cell_preco.font = copy(cell_modelo.font)
			cell_preco.border = copy(cell_modelo.border)
			cell_preco.alignment = copy(cell_modelo.alignment)
			cell_preco.number_format = 'General'
			cell_preco.protection = copy(cell_modelo.protection)

		# Buscar cabeçalho 'LOCAÇÃO'
		header = None
		idx_locacao = None
		header_row = None
		for r in range(1, 21):
			row_values = [cell.value for cell in ws1[r]]
			if row_values and 'LOCAÇÃO' in row_values:
				header = row_values
				idx_locacao = row_values.index('LOCAÇÃO')
				header_row = r
				break

		if header is None:
			return {'success': False, 'error': 'Cabeçalho LOCAÇÃO não encontrado no modelo'}

		# Detectar índice da coluna UNIDADE
		idx_unidade = None
		for i, col_name in enumerate(header):
			if col_name and str(col_name).strip().upper() == 'UNIDADE':
				idx_unidade = i
				break

		# Preencher dados a partir da linha 12
		linha = 12
		lote_nome = f"LOTE {lote_id}"
		
		a12 = ws1.cell(row=12, column=1)
		style_a12 = {
			'font': copy(a12.font),
			'border': copy(a12.border),
			'alignment': copy(a12.alignment),
			'number_format': a12.number_format,
			'protection': copy(a12.protection)
		}

		tem_dados = False
		for mapa in mapas_filtrados:
			unidade_nome = (mapa.get('unidade') or '').strip()
			dados_siisp = mapa.get('dados_siisp', [])
			datas = mapa.get('datas', [])
			
			for i in range(len(datas)):
				if idx_locacao is not None:
					cell_locacao = ws1.cell(row=linha, column=idx_locacao+1, value=lote_nome)
					cell_locacao.font = style_a12['font']
					cell_locacao.border = style_a12['border']
					cell_locacao.alignment = style_a12['alignment']
					cell_locacao.number_format = style_a12['number_format']
					cell_locacao.protection = style_a12['protection']
				
				if idx_unidade is not None:
					cell_unidade = ws1.cell(row=linha, column=idx_unidade+1, value=unidade_nome)
					cell_unidade.font = style_a12['font']
					cell_unidade.border = style_a12['border']
					cell_unidade.alignment = style_a12['alignment']
					cell_unidade.number_format = style_a12['number_format']
					cell_unidade.protection = style_a12['protection']
				
				valor_siisp = dados_siisp[i] if i < len(dados_siisp) else 0
				cell_siisp = ws1.cell(row=linha, column=3, value=valor_siisp)
				cell_siisp.font = style_a12['font']
				cell_siisp.border = style_a12['border']
				cell_siisp.alignment = style_a12['alignment']
				cell_siisp.number_format = 'General'
				cell_siisp.protection = style_a12['protection']

				data_val = datas[i] if i < len(datas) else ''
				cell_data = ws1.cell(row=linha, column=4, value=data_val)
				cell_data.font = style_a12['font']
				cell_data.border = style_a12['border']
				cell_data.alignment = style_a12['alignment']
				cell_data.number_format = 'DD/MM/YYYY'
				cell_data.protection = style_a12['protection']

				colunas_refeicoes = [
					('cafe_interno', 5),
					('cafe_funcionario', 6),
					('almoco_interno', 7),
					('almoco_funcionario', 8),
					('lanche_interno', 9),
					('lanche_funcionario', 10),
					('jantar_interno', 11),
					('jantar_funcionario', 12)
				]
				for campo, col in colunas_refeicoes:
					valores = mapa.get(campo, [])
					valor_refeicao = valores[i] if i < len(valores) else 0
					cell_refeicao = ws1.cell(row=linha, column=col, value=valor_refeicao)
					cell_refeicao.font = style_a12['font']
					cell_refeicao.border = style_a12['border']
					cell_refeicao.alignment = style_a12['alignment']
					cell_refeicao.number_format = 'General'
					cell_refeicao.protection = style_a12['protection']

				linha += 1
				tem_dados = True

		if not tem_dados:
			return {'success': False, 'error': 'Nenhum dado para exportar'}

		# Copiar fórmulas de M12:T12 para as linhas preenchidas
		linhas_preenchidas = linha - 12
		for offset in range(1, linhas_preenchidas):
			target_row = 12 + offset
			for col in range(13, 21):
				formula_or_value = ws1.cell(row=12, column=col).value
				if ws1.cell(row=12, column=col).data_type == 'f':
					formula_ajustada = re.sub(r'(\D)12(\D|$)', lambda m: m.group(1)+str(target_row)+m.group(2), formula_or_value)
					ws1.cell(row=target_row, column=col, value=formula_ajustada)
					ws1.cell(row=target_row, column=col).data_type = 'f'
				else:
					ws1.cell(row=target_row, column=col, value=formula_or_value)
				
				cell_modelo = ws1.cell(row=12, column=col)
				cell_dest = ws1.cell(row=target_row, column=col)
				cell_dest.font = copy(cell_modelo.font)
				cell_dest.border = copy(cell_modelo.border)
				cell_dest.alignment = copy(cell_modelo.alignment)
				cell_dest.number_format = cell_modelo.number_format
				cell_dest.protection = copy(cell_modelo.protection)

		# Copiar regras de formatação condicional
		for col in range(13, 21):
			cell_coord = ws1.cell(row=12, column=col).coordinate
			regras_para_copiar = []
			for cf_rule in ws1.conditional_formatting:
				if cell_coord in cf_rule.cells:
					for rule in cf_rule.rules:
						regras_para_copiar.append(rule)
			for rule in regras_para_copiar:
				for target_row in range(13, linha):
					target_coord = ws1.cell(row=target_row, column=col).coordinate
					ws1.conditional_formatting.add(target_coord, rule)

		# Salvar em memória
		output = io.BytesIO()
		wb.save(output)
		output.seek(0)

		# Nome do arquivo
		nome_arquivo = f"tabela_lote_{lote_id}"
		if data_inicio and data_fim:
			nome_arquivo += f"_{data_inicio}_a_{data_fim}"
		nome_arquivo += ".xlsx"

		return {
			'success': True,
			'output': output,
			'filename': nome_arquivo
		}
	
	except Exception as e:
		import traceback
		traceback.print_exc()
		return {'success': False, 'error': f'Erro ao gerar arquivo: {str(e)}'}